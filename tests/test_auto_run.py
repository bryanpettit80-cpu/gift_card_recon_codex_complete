from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from gift_card_recon.auto_run import iso_week_period, run_weekly_reconciliations
from gift_card_recon.utils import parse_date


def test_iso_week_period_uses_report_end_date():
    assert iso_week_period(parse_date("2026-06-07")) == "2026-W23"


def test_auto_weekly_runner_infers_period_and_week_ending(tmp_path: Path):
    input_dir = tmp_path / "input" / "9355" / "weekly"
    activity_dir = input_dir / "activity"
    activity_dir.mkdir(parents=True)
    (input_dir / "summary").mkdir()

    create_activity(
        activity_dir / "06.07.2026 9355 Gift Card Activity.xlsx",
        begin="01-JUN-2026",
        end="07-JUN-2026",
        gross_activation=Decimal("150.00"),
        redemption=Decimal("-6657.73"),
    )
    (input_dir / "pos_controls.csv").write_text(
        "store,period,pos_gift_card_issue,pos_gift_card_payment\n9355,auto,150.00,6657.73\n",
        encoding="utf-8",
    )

    reports = run_weekly_reconciliations(input_root=tmp_path / "input", output_dir=tmp_path / "output")

    assert len(reports) == 1
    assert reports[0].status == "created"
    assert reports[0].period == "2026-W23"
    assert reports[0].period_end == parse_date("2026-06-07")

    output_path = tmp_path / "output" / "Gift_Card_Reconciliation_9355_2026-W23.xlsx"
    assert output_path.exists()
    wb = load_workbook(output_path, data_only=False)
    ws = wb["Reconciliation"]
    assert "Week Ending 06/07/2026" in ws["A1"].value
    assert "06/14/2026" not in ws["A1"].value
    assert ws["C6"].value == 150
    assert ws["C7"].value == 6657.73
    assert ws["A11"].value == "Gift Card Activity File Totals"
    assert ws["A13"].value == "06.07.2026 9355 Gift Card Activity.xlsx"
    assert ws["D13"].value == 150
    assert ws["I13"].value == -6657.73
    assert ws["J13"].value == -6507.73
    assert (input_dir / "pos_controls.csv").read_text(encoding="utf-8").splitlines() == [
        "store,period,pos_gift_card_issue,pos_gift_card_payment",
        "9355,auto,,",
    ]


def test_auto_weekly_runner_leaves_pos_controls_when_workbook_is_not_created(tmp_path: Path):
    input_dir = tmp_path / "input" / "9354" / "weekly"
    input_dir.mkdir(parents=True)
    original_controls = "store,period,pos_gift_card_issue,pos_gift_card_payment\n9354,auto,275.00,980.00\n"
    pos_path = input_dir / "pos_controls.csv"
    pos_path.write_text(original_controls, encoding="utf-8")

    reports = run_weekly_reconciliations(input_root=tmp_path / "input", output_dir=tmp_path / "output")

    assert len(reports) == 1
    assert reports[0].status == "skipped"
    assert pos_path.read_text(encoding="utf-8") == original_controls
    assert not (tmp_path / "output").exists()


def create_activity(path: Path, *, begin: str, end: str, gross_activation: Decimal, redemption: Decimal) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    ws.append([f"All GC Activity BY Rest Number and Date Range  BEGIN DATE: '{begin}', END DATE: '{end}'"])
    ws.append(["Response:000"])
    ws.append(["Rest Number:9355"])
    ws.append(["Card No", "Request", "Request Code Listing", "Business Date", "Corp Code", "Transaction No", "Amount SUM", "Promocode", "Authorization Code"])
    ws.append(["0001xxxx", 100, "Activation", "2026-06-07", None, 1, float(gross_activation), None, 111111])
    ws.append(["0002xxxx", 202, "Redemption No Nsf", "2026-06-07", None, 2, float(redemption), None, 222222])
    wb.save(path)
