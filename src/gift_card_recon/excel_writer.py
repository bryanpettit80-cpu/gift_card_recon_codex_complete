from __future__ import annotations

from copy import copy
from decimal import Decimal
from pathlib import Path
from typing import Any

from gift_card_recon.models import ReconciliationResult

MONEY_FMT = '$#,##0.00;($#,##0.00);-'
INT_FMT = '#,##0'
DATE_FMT = 'yyyy-mm-dd'


def write_reconciliation_workbook(result: ReconciliationResult, output_path: Path) -> Path:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write the reconciliation workbook. Run: pip install -r requirements.txt") from exc

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"
    _write_reconciliation_sheet(ws, result)
    _write_weekly_sheet(wb.create_sheet("Weekly Activity Detail"), result)
    _write_daily_sheet(wb.create_sheet("Daily Activity Detail"), result)
    _write_raw_sheet(wb.create_sheet("Raw Detail"), result)
    _write_source_files_sheet(wb.create_sheet("Source Files"), result)
    _write_exception_sheet(wb.create_sheet("Exception Log"), result)

    for sheet in wb.worksheets:
        _apply_freeze_and_filter(sheet)
        _auto_width(sheet)

    wb.save(output_path)
    return output_path


def _write_reconciliation_sheet(ws, result: ReconciliationResult) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    mode_label = result.mode.title()
    title = f"Gift Card Reconciliation - {mode_label} - Store {result.store} - {result.period}"
    if result.period_end:
        ending_label = "Week Ending" if result.mode == "weekly" else "Period Ending"
        title += f" - {ending_label} {result.period_end:%m/%d/%Y}"
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:H3")
    ws["A3"] = _build_conclusion(result)
    ws["A3"].fill = PatternFill("solid", fgColor="F2F2F2")
    ws["A3"].font = Font(italic=True, color="333333")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 42

    headers = ["Metric", "Summary Control", "GC Activity File Total", "Activity Variance", "POS Control", "POS Variance", "Status", "Review Note"]
    _write_row(ws, 5, headers)
    _style_header_row(ws, 5, 8)

    for row_idx, line in enumerate(result.lines, start=6):
        _write_row(
            ws,
            row_idx,
            [
                line.metric,
                _display_money(line.summary_value),
                _decimal_to_number(line.activity_value),
                _display_money(line.activity_variance),
                _decimal_to_number(line.pos_value),
                _decimal_to_number(line.pos_variance),
                line.status,
                line.note,
            ],
        )

    next_row = _write_activity_file_totals(ws, 11, result)

    system_title_row = next_row + 2
    system_header_row = system_title_row + 1
    system_data_row = system_header_row + 1
    _section_title(ws, system_title_row, "Gift Card System Detail")
    _write_row(ws, system_header_row, ["Metric", "Summary Value", "Activity-Derived Value", "Variance", "Status", "Logic"])
    _style_header_row(ws, system_header_row, 6)

    summary = result.summary
    activity_conversion = sum((r.conversion_redemptions for r in result.weekly_rollups), Decimal("0.00"))
    activity_non_conversion = sum((r.non_conversion_redemptions for r in result.weekly_rollups), Decimal("0.00"))
    summary_conversion = summary.conversion_redemptions if summary else None
    summary_non_conversion = summary.non_conversion_redemptions if summary else None
    conversion_var = (summary_conversion - activity_conversion) if summary_conversion is not None else None
    non_conversion_var = (summary_non_conversion - activity_non_conversion) if summary_non_conversion is not None else None

    detail_rows = [
        ["Conversion Promo Redemptions", _display_money(summary_conversion), _decimal_to_number(activity_conversion), _display_money(conversion_var), _status_for_variance(conversion_var), _summary_logic(summary, "conversion")],
        ["Non-Conversion Redemptions", _display_money(summary_non_conversion), _decimal_to_number(activity_non_conversion), _display_money(non_conversion_var), _status_for_variance(non_conversion_var), _summary_logic(summary, "non_conversion")],
        ["GCDR", _display_money(summary.gcdr if summary else None), "N/A", "N/A", "Info", _summary_logic(summary, "retained")],
        ["Payable Redemptions", _display_money(summary.payable_redemptions if summary else None), "N/A", "N/A", "Info", _summary_logic(summary, "retained")],
        ["Net Settlement", _display_money(summary.net_settlement if summary else None), "N/A", "N/A", "Info", _summary_logic(summary, "retained")],
    ]
    for idx, row in enumerate(detail_rows, start=system_data_row):
        _write_row(ws, idx, row)

    pos_title_row = system_data_row + len(detail_rows) + 3
    pos_header_row = pos_title_row + 1
    pos_data_row = pos_header_row + 1
    _section_title(ws, pos_title_row, "POS Controls Included on Reconciliation")
    _write_row(ws, pos_header_row, ["Control", "Amount", "Note"])
    _style_header_row(ws, pos_header_row, 3)
    _write_row(ws, pos_data_row, ["POS Gift Card Issue", _decimal_to_number(result.pos_controls.pos_gift_card_issue), "External POS control supplied for the period."])
    _write_row(ws, pos_data_row + 1, ["POS Gift Card Payment", _decimal_to_number(result.pos_controls.pos_gift_card_payment), "External POS control supplied for the period."])
    _write_row(ws, pos_data_row + 2, ["POS Net Impact", _decimal_to_number(result.pos_controls.net_impact), "Issue less payment. Negative means payment exceeded issue."])

    _format_currency(
        ws,
        [
            "B6:F8",
            f"D13:J{max(13, next_row - 1)}",
            f"B{system_data_row}:D{system_data_row + len(detail_rows) - 1}",
            f"B{pos_data_row}:B{pos_data_row + 2}",
        ],
    )
    _format_status(ws)
    _format_body(ws)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A5:H8"


def _write_activity_file_totals(ws, start_row: int, result: ReconciliationResult) -> int:
    _section_title(ws, start_row, "Gift Card Activity File Totals", max_col=10)
    headers = [
        "Source File",
        "Report Period",
        "Rows",
        "Gross Activations",
        "Void Activations",
        "Net Activations",
        "Gross Redemptions",
        "Void Redemptions",
        "Net Redemptions",
        "Net Activity",
    ]
    header_row = start_row + 1
    _write_row(ws, header_row, headers)
    _style_header_row(ws, header_row, len(headers))

    row_idx = header_row + 1
    total_rows = 0
    gross_activations = Decimal("0.00")
    void_activations = Decimal("0.00")
    net_activations = Decimal("0.00")
    gross_redemptions = Decimal("0.00")
    void_redemptions = Decimal("0.00")
    net_redemptions = Decimal("0.00")
    net_activity = Decimal("0.00")

    for rollup in result.weekly_rollups:
        total_rows += rollup.row_count
        gross_activations += rollup.gross_activations
        void_activations += rollup.void_activations
        net_activations += rollup.net_activations
        gross_redemptions += rollup.gross_redemptions
        void_redemptions += rollup.void_redemptions
        net_redemptions += rollup.net_redemptions
        net_activity += rollup.net_activity
        _write_row(
            ws,
            row_idx,
            [
                rollup.source_file,
                rollup.report_period,
                rollup.row_count,
                _decimal_to_number(rollup.gross_activations),
                _decimal_to_number(rollup.void_activations),
                _decimal_to_number(rollup.net_activations),
                _decimal_to_number(rollup.gross_redemptions),
                _decimal_to_number(rollup.void_redemptions),
                _decimal_to_number(rollup.net_redemptions),
                _decimal_to_number(rollup.net_activity),
            ],
        )
        row_idx += 1

    if len(result.weekly_rollups) > 1:
        _write_row(
            ws,
            row_idx,
            [
                "TOTAL",
                "",
                total_rows,
                _decimal_to_number(gross_activations),
                _decimal_to_number(void_activations),
                _decimal_to_number(net_activations),
                _decimal_to_number(gross_redemptions),
                _decimal_to_number(void_redemptions),
                _decimal_to_number(net_redemptions),
                _decimal_to_number(net_activity),
            ],
        )
        _style_total_row(ws, row_idx, len(headers))
        row_idx += 1

    for row in range(header_row + 1, row_idx):
        ws.cell(row=row, column=3).number_format = INT_FMT
    return row_idx


def _write_weekly_sheet(ws, result: ReconciliationResult) -> None:
    ws["A1"] = "Weekly Activity Detail"
    _style_title(ws, "A1:L1")
    headers = ["Source File", "Report Period", "Rows", "Gross Activations", "Void Activations", "Net Activations", "Gross Redemptions", "Void Redemptions", "Net Redemptions", "Conversion Promo Redemptions", "Non-Conversion Redemptions", "Net Activity"]
    _write_row(ws, 3, headers)
    _style_header_row(ws, 3, len(headers))
    for idx, r in enumerate(result.weekly_rollups, start=4):
        _write_row(ws, idx, [r.source_file, r.report_period, r.row_count, _decimal_to_number(r.gross_activations), _decimal_to_number(r.void_activations), _decimal_to_number(r.net_activations), _decimal_to_number(r.gross_redemptions), _decimal_to_number(r.void_redemptions), _decimal_to_number(r.net_redemptions), _decimal_to_number(r.conversion_redemptions), _decimal_to_number(r.non_conversion_redemptions), _decimal_to_number(r.net_activity)])
    total_row = 4 + len(result.weekly_rollups)
    if result.weekly_rollups:
        _write_row(ws, total_row, ["TOTAL", "", f"=SUM(C4:C{total_row-1})"] + [f"=SUM({col}4:{col}{total_row-1})" for col in "DEFGHIJKL"])
        _style_total_row(ws, total_row, 12)
    _format_currency(ws, [f"D4:L{max(total_row, 4)}"])
    for row in range(4, max(total_row, 4) + 1):
        ws.cell(row=row, column=3).number_format = INT_FMT


def _write_daily_sheet(ws, result: ReconciliationResult) -> None:
    ws["A1"] = "Daily Activity Detail"
    _style_title(ws, "A1:G1")
    headers = ["Business Date", "Source File", "Net Activations", "Net Redemptions", "Conversion Promo Redemptions", "Non-Conversion Redemptions", "Net Activity"]
    _write_row(ws, 3, headers)
    _style_header_row(ws, 3, len(headers))
    for idx, r in enumerate(result.daily_rollups, start=4):
        _write_row(ws, idx, [r.business_date, r.source_file, _decimal_to_number(r.net_activations), _decimal_to_number(r.net_redemptions), _decimal_to_number(r.conversion_redemptions), _decimal_to_number(r.non_conversion_redemptions), _decimal_to_number(r.net_activity)])
    _format_currency(ws, [f"C4:G{max(4, 3 + len(result.daily_rollups))}"])
    for cell in ws["A"]:
        if cell.row >= 4:
            cell.number_format = DATE_FMT


def _write_raw_sheet(ws, result: ReconciliationResult) -> None:
    ws["A1"] = "Raw Detail Parsed from Weekly Reports"
    _style_title(ws, "A1:I1")
    headers = ["Source File", "Card No", "Request", "Request Code Listing", "Business Date", "Transaction No", "Amount", "Promocode", "Authorization Code"]
    _write_row(ws, 3, headers)
    _style_header_row(ws, 3, len(headers))
    for idx, row in enumerate(result.raw_rows, start=4):
        _write_row(ws, idx, [row.source_file, row.card_no, row.request, row.request_code_listing, row.business_date, row.transaction_no, _decimal_to_number(row.amount), row.promocode, row.authorization_code])
    _format_currency(ws, [f"G4:G{max(4, 3 + len(result.raw_rows))}"])
    for cell in ws["E"]:
        if cell.row >= 4:
            cell.number_format = DATE_FMT


def _write_source_files_sheet(ws, result: ReconciliationResult) -> None:
    ws["A1"] = "Source File Audit Trail"
    _style_title(ws, "A1:F1")
    headers = ["File Name", "Type", "Size Bytes", "Modified At", "SHA-256", "Full Path"]
    _write_row(ws, 3, headers)
    _style_header_row(ws, 3, len(headers))
    for idx, audit in enumerate(result.source_files, start=4):
        _write_row(ws, idx, [audit.path.name, audit.file_type, audit.size_bytes, audit.modified_at, audit.sha256, str(audit.path)])
    for cell in ws["D"]:
        if cell.row >= 4:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"


def _write_exception_sheet(ws, result: ReconciliationResult) -> None:
    ws["A1"] = "Exception Log"
    _style_title(ws, "A1:B1")
    _write_row(ws, 3, ["Severity", "Message"])
    _style_header_row(ws, 3, 2)
    if result.exceptions:
        for idx, (severity, message) in enumerate(result.exceptions, start=4):
            _write_row(ws, idx, [severity, message])
    else:
        _write_row(ws, 4, ["OK", "No parsing or validation exceptions recorded."])
    _format_status(ws)


def _build_conclusion(result: ReconciliationResult) -> str:
    if result.mode == "weekly" and result.summary is None:
        pos_reviews = [line for line in result.lines if line.pos_variance is not None and abs(line.pos_variance) > Decimal("0.01")]
        if pos_reviews:
            review_text = "; ".join(f"{line.metric}: {line.pos_variance:+,.2f}" for line in pos_reviews)
            return f"Weekly mode. No weekly summary supplied; activity is reconciled directly to POS controls on this tab. POS variance review: {review_text}."
        return "Weekly mode. No weekly summary supplied; activity reconciles directly to POS controls within tolerance."
    if result.mode == "weekly":
        pos_reviews = [line for line in result.lines if line.pos_variance is not None and abs(line.pos_variance) > Decimal("0.01")]
        if pos_reviews:
            review_text = "; ".join(f"{line.metric}: {line.pos_variance:+,.2f}" for line in pos_reviews)
            return f"Weekly mode. Optional summary is included. POS controls are included on this tab. POS variance review: {review_text}."
        return "Weekly mode. Optional summary is included and POS controls are within tolerance."
    activity_clean = all(line.activity_variance is not None and abs(line.activity_variance) <= Decimal("0.01") for line in result.lines)
    pos_reviews = [line for line in result.lines if line.pos_variance is not None and abs(line.pos_variance) > Decimal("0.01")]
    if activity_clean and pos_reviews:
        review_text = "; ".join(f"{line.metric}: {line.pos_variance:+,.2f}" for line in pos_reviews)
        return f"Summary ties to weekly gift card activity. POS controls are included on this tab. POS variance review: {review_text}."
    if activity_clean:
        return "Summary ties to weekly gift card activity and POS controls are within tolerance."
    return "Review required: one or more summary-to-activity variances exists."


def _write_row(ws, row_idx: int, values: list[Any]) -> None:
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)


def _section_title(ws, row_idx: int, title: str, max_col: int = 8) -> None:
    from openpyxl.styles import Font, PatternFill
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
    cell = ws.cell(row=row_idx, column=1, value=title)
    cell.font = Font(bold=True, color="1F4E78")
    cell.fill = PatternFill("solid", fgColor="D9EAF7")


def _style_title(ws, merge_range: str) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    ws.merge_cells(merge_range)
    cell = ws[merge_range.split(":")[0]]
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center")


def _style_header_row(ws, row_idx: int, max_col: int) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    fill = PatternFill("solid", fgColor="5B9BD5")
    font = Font(bold=True, color="FFFFFF")
    side = Side(style="thin", color="D9EAF7")
    for col in range(1, max_col + 1):
        cell = ws.cell(row_idx, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=side, bottom=side, left=side, right=side)


def _style_total_row(ws, row_idx: int, max_col: int) -> None:
    from openpyxl.styles import Border, Font, PatternFill, Side
    fill = PatternFill("solid", fgColor="D9EAF7")
    side = Side(style="thin", color="BFBFBF")
    for col in range(1, max_col + 1):
        cell = ws.cell(row_idx, col)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = Border(top=side, bottom=side, left=side, right=side)


def _format_currency(ws, ranges: list[str]) -> None:
    for rng in ranges:
        for row in ws[rng]:
            for cell in row:
                cell.number_format = MONEY_FMT


def _format_body(ws) -> None:
    from openpyxl.styles import Alignment
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in ["A", "G"]:
        for cell in ws[col]:
            if cell.row > 1:
                new_font = copy(cell.font)
                new_font.bold = True
                cell.font = new_font


def _format_status(ws) -> None:
    from openpyxl.styles import PatternFill
    fills = {
        "OK": PatternFill("solid", fgColor="E2F0D9"),
        "Minor variance": PatternFill("solid", fgColor="FFF2CC"),
        "Review": PatternFill("solid", fgColor="FCE4D6"),
        "Info": PatternFill("solid", fgColor="E7E6E6"),
        "N/A": PatternFill("solid", fgColor="E7E6E6"),
    }
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in fills:
                cell.fill = fills[cell.value]


def _apply_freeze_and_filter(ws) -> None:
    if ws.auto_filter.ref:
        return
    if ws.max_row >= 3:
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{ws.cell(row=3, column=ws.max_column).coordinate}"


def _auto_width(ws) -> None:
    from openpyxl.utils import get_column_letter
    caps = {1: 34, 2: 32, 3: 18, 4: 28, 5: 18, 6: 18, 7: 18, 8: 46, 9: 22}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), caps.get(col_idx, 30))


def _decimal_to_number(value: Decimal | None) -> float | None:
    return None if value is None else float(value)


def _display_money(value: Decimal | None) -> float | str:
    return "N/A" if value is None else float(value)


def _status_for_variance(value: Decimal | None) -> str:
    if value is None:
        return "N/A"
    return "OK" if abs(value) <= Decimal("0.01") else "Review"


def _summary_logic(summary, kind: str) -> str:
    if summary is None:
        return "N/A - no weekly summary supplied."
    if kind == "conversion":
        return f"Promo codes from summary: {', '.join(sorted(summary.conversion_promo_codes)) or 'none found'}"
    if kind == "non_conversion":
        return "Total redemptions less summary-listed conversion promo code redemptions."
    return "Source summary value retained for accounting review."
