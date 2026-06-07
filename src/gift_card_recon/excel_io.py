from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any


class ExcelReadError(RuntimeError):
    pass


def workbook_values(path: Path) -> dict[str, list[list[Any]]]:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _xlsx_values(path)
    if suffix == ".xls":
        return _xls_values(path)
    raise ExcelReadError(f"Unsupported Excel file extension for {path.name!r}. Expected .xlsx or .xls.")


def first_sheet_values(path: Path) -> tuple[str, list[list[Any]]]:
    sheets = workbook_values(path)
    if not sheets:
        raise ExcelReadError(f"No worksheets found in {path}")
    name = next(iter(sheets))
    return name, sheets[name]


def _xlsx_values(path: Path) -> dict[str, list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExcelReadError("openpyxl is required to read .xlsx files. Run: pip install -r requirements.txt") from exc

    wb = load_workbook(path, data_only=True, read_only=True)
    result: dict[str, list[list[Any]]] = {}
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        result[ws.title] = _trim_matrix(rows)
    wb.close()
    return result


def _xls_values(path: Path) -> dict[str, list[list[Any]]]:
    try:
        import xlrd
        from xlrd.xldate import xldate_as_datetime
    except ImportError as exc:
        raise ExcelReadError("xlrd is required to read legacy .xls activity files. Run: pip install -r requirements.txt") from exc

    book = xlrd.open_workbook(str(path))
    result: dict[str, list[list[Any]]] = {}
    for sheet in book.sheets():
        rows: list[list[Any]] = []
        for r in range(sheet.nrows):
            row_values: list[Any] = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    dt = xldate_as_datetime(value, book.datemode)
                    value = dt.date() if dt.time() == datetime.min.time() else dt
                elif cell.ctype == xlrd.XL_CELL_NUMBER and isinstance(value, float) and value.is_integer():
                    value = int(value)
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    value = None
                row_values.append(value)
            rows.append(row_values)
        result[sheet.name] = _trim_matrix(rows)
    return result


def _trim_matrix(rows: list[list[Any]]) -> list[list[Any]]:
    def row_has_value(row: list[Any]) -> bool:
        return any(v is not None and v != "" for v in row)

    while rows and not row_has_value(rows[-1]):
        rows.pop()
    if not rows:
        return []
    max_col = 0
    for row in rows:
        for idx, value in enumerate(row, start=1):
            if value is not None and value != "":
                max_col = max(max_col, idx)
    return [row[:max_col] + [None] * max(0, max_col - len(row)) for row in rows]
