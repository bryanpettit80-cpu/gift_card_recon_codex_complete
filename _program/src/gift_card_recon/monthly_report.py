from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING, Iterable, Sequence

from gift_card_recon.close_assessment import (
    CloseAssessment,
    CloseStatus,
    ControlDisposition,
    ControlOutcome,
    REQUIRED_CLOSE_INTEGRITY_CODES,
)
from gift_card_recon.models import MonthlyCloseCertification, ReconciliationResult

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


ACCOUNTING_FMT = '$#,##0.00;($#,##0.00);$0.00'
DATE_FMT = "mm/dd/yyyy"

_FONT_NAME = "Arial"
_NAVY = "17365D"
_LIGHT_BLUE = "D9EAF7"
_GREEN = "E2F0D9"
_DARK_GREEN = "375623"
_AMBER = "FFF2CC"
_DARK_AMBER = "7F6000"
_RED = "F4CCCC"
_DARK_RED = "9C0006"
_GRAY = "F2F2F2"
_TEXT = "333333"

DEFAULT_EVIDENCE_LABELS = (
    "Gift Card Summary",
    "Weekly Gift Card Activity Reports",
    "Micros Daily System Totals",
    "Micros Tender Detail",
    "Darden Credit Memo",
)


@dataclass(frozen=True)
class WeeklyCloseReportRow:
    """One assessed week for page 2 of the monthly close report.

    The caller supplies ``disposition``. The renderer intentionally does not
    infer a status from the monetary values.
    """

    week_ending: date | None
    coverage: str
    pos_issue_variance: Decimal | None
    pos_payment_variance: Decimal | None
    pos_net_variance: Decimal | None
    tender_variance: Decimal | None
    disposition: ControlDisposition
    evidence_note: str = ""


@dataclass(frozen=True)
class MonthlyCloseReportData:
    """Presentation-only data for a two-page close certificate or diagnostic."""

    assessment: CloseAssessment
    period: str
    period_start: date
    period_end: date
    generated_at: datetime
    certification: MonthlyCloseCertification | None = None
    result: ReconciliationResult | None = None
    weekly_rows: tuple[WeeklyCloseReportRow, ...] = ()
    period_pos_net_variance: Decimal | None = None
    period_pos_disposition: ControlDisposition | None = None
    period_tender_variance: Decimal | None = None
    period_tender_disposition: ControlDisposition | None = None
    evidence_notes: tuple[str, ...] = ()
    source_labels: tuple[str, ...] = field(default_factory=lambda: DEFAULT_EVIDENCE_LABELS)
    explicit_exceptions: tuple[tuple[str, str], ...] = ()
    weekly_control_codes: frozenset[str] = frozenset()

    def __post_init__(self) -> None:
        store = self.assessment.store
        if self.period_end < self.period_start:
            raise ValueError("The monthly close report end date precedes its start date.")
        if self.certification is not None:
            if self.certification.store != store:
                raise ValueError("Certification store does not match the close assessment.")
            if self.certification.period != self.period:
                raise ValueError("Certification period does not match the report period.")
            if self.certification.darden_matched != self.assessment.darden_matched:
                raise ValueError(
                    "Certification Darden match state does not match the close assessment."
                )
        if self.result is not None:
            if self.result.store != store:
                raise ValueError("Reconciliation store does not match the close assessment.")
            if self.result.period != self.period:
                raise ValueError("Reconciliation period does not match the report period.")

    @property
    def exceptions(self) -> tuple[tuple[str, str], ...]:
        result_exceptions: Iterable[tuple[str, str]] = ()
        if self.result is not None:
            result_exceptions = self.result.exceptions
        return tuple(result_exceptions) + self.explicit_exceptions


def write_monthly_close_report_workbook(
    data: MonthlyCloseReportData,
    output_path: Path,
) -> Path:
    """Write a report-only workbook, including blocked-run diagnostics."""

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to write the monthly close report workbook."
        ) from exc

    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    write_monthly_close_report(workbook.active, data)
    workbook.save(destination)
    return destination


def write_monthly_close_report(ws: Worksheet, data: MonthlyCloseReportData) -> None:
    """Render the authoritative two-page monthly report into ``ws``.

    Overall and row-level dispositions are consumed from assessed inputs. This
    module performs presentation calculations only (for example, identifying
    the largest absolute weekly variance).
    """

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.pagebreak import Break

    ws.title = "Monthly Close Report"
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="B7C9DB")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    config = data.assessment.store_config
    heading = config.report_heading
    _set_document_properties(ws, data, heading)
    _write_page_heading(ws, 1, heading, data)

    ws.merge_cells("A4:H5")
    status_cell = ws["A4"]
    status_cell.value = data.assessment.status.value
    status_cell.font = Font(
        name=_FONT_NAME,
        bold=True,
        color=_status_text(data.assessment.status),
        size=16,
    )
    status_cell.fill = PatternFill("solid", fgColor=_status_fill(data.assessment.status))
    status_cell.alignment = Alignment(horizontal="center", vertical="center")
    status_cell.border = border
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 24

    _section_title(ws, 7, "Settlement Tie-Out")
    certification = data.certification
    summary_total = certification.summary_net_settlement if certification else None
    darden_total = certification.darden_credit_memo.total if certification else None
    darden_variance = certification.variance if certification else _darden_variance(data.assessment)
    darden_evaluated = certification is not None or darden_variance is not None
    darden_result = (
        "MATCHED"
        if data.assessment.darden_matched
        else "MISMATCHED" if darden_evaluated else "NOT EVALUATED"
    )
    cards = (
        ("Summary Settlement", summary_total, ACCOUNTING_FMT),
        ("Darden Credit Memo", darden_total, ACCOUNTING_FMT),
        ("Difference", darden_variance, ACCOUNTING_FMT),
        ("Match Status", darden_result, None),
    )
    for index, (label, value, number_format) in enumerate(cards):
        start_col = 1 + (index * 2)
        ws.merge_cells(start_row=8, start_column=start_col, end_row=8, end_column=start_col + 1)
        ws.merge_cells(start_row=9, start_column=start_col, end_row=10, end_column=start_col + 1)
        label_cell = ws.cell(8, start_col, label)
        value_cell = ws.cell(9, start_col, _excel_money(value) if number_format else value)
        label_cell.fill = PatternFill("solid", fgColor=_NAVY)
        label_cell.font = Font(name=_FONT_NAME, bold=True, color="FFFFFF", size=10)
        label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        label_cell.border = border
        value_cell.fill = PatternFill("solid", fgColor=_GRAY if index % 2 == 0 else "FFFFFF")
        value_cell.font = Font(name=_FONT_NAME, bold=True, color=_TEXT, size=11)
        value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value_cell.border = border
        if number_format:
            value_cell.number_format = number_format
        else:
            match_disposition = (
                ControlDisposition.PASS
                if data.assessment.darden_matched
                else ControlDisposition.BLOCK if darden_evaluated else None
            )
            if match_disposition is not None:
                value_cell.fill = PatternFill(
                    "solid", fgColor=_disposition_fill(match_disposition)
                )
                value_cell.font = Font(
                    name=_FONT_NAME,
                    bold=True,
                    color=_disposition_text_color(match_disposition),
                    size=11,
                )

    _section_title(ws, 12, "Close Controls")
    _write_cells(ws, 13, ("Control", "", "", "Status", "Variance", "Conclusion", "", ""))
    ws.merge_cells("A13:C13")
    ws.merge_cells("F13:H13")
    _style_header(ws, 13, 8)
    executive_controls = _executive_controls(data)
    control_start = 14
    for row_index, control in enumerate(executive_controls, start=control_start):
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
        ws.merge_cells(start_row=row_index, start_column=6, end_row=row_index, end_column=8)
        ws.cell(row_index, 1, control.label)
        ws.cell(row_index, 4, control.disposition.value)
        ws.cell(row_index, 5, _excel_money(control.variance))
        ws.cell(row_index, 6, _concise_control_message(control))
        ws.cell(row_index, 5).number_format = ACCOUNTING_FMT
        _style_assessed_row(ws, row_index, control.disposition, border)

    next_row = control_start + len(executive_controls) + 1
    _section_title(ws, next_row, "Open Items Summary")
    next_row += 1
    open_controls = tuple(control for control in data.assessment.controls if not control.passed)
    if open_controls:
        weekly_open = tuple(
            control
            for control in open_controls
            if _is_weekly_child_control(control, data.weekly_control_codes)
        )
        nonweekly_open = tuple(control for control in open_controls if control not in weekly_open)
        if weekly_open:
            disposition = _worst_disposition(weekly_open)
            affected_weeks = _weekly_open_groups(weekly_open)
            action_text = (
                f"{len(affected_weeks)} week(s) contain reviewed controls. "
                "See Weekly Variance Detail and Review Items on page 2."
            )
            ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=8)
            ws.cell(next_row, 1, disposition.value)
            ws.cell(next_row, 2, action_text)
            _style_action_row(ws, next_row, disposition, border)
            next_row += 1
        for control in nonweekly_open:
            ws.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=8)
            ws.cell(next_row, 1, control.disposition.value)
            ws.cell(next_row, 2, f"{control.label}: {_concise_control_message(control)}")
            _style_action_row(ws, next_row, control.disposition, border)
            next_row += 1
    else:
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=8)
        ws.cell(next_row, 1, "No open items. All close controls passed.")
        _style_neutral_merged_row(ws, next_row, border)
        next_row += 1

    page_two_start = next_row + 2
    ws.row_breaks.append(Break(id=page_two_start - 1))

    _write_page_heading(ws, page_two_start, heading, data)
    weekly_title = page_two_start + 3
    _section_title(ws, weekly_title, "Weekly Variance Detail")
    weekly_header = weekly_title + 1
    headers = (
        "Week Ending",
        "Coverage",
        "POS Issue",
        "POS Payment",
        "POS Net",
        "Tender",
        "Status",
        "Follow-up",
    )
    _write_cells(ws, weekly_header, headers)
    _style_header(ws, weekly_header, 8)
    row_cursor = weekly_header + 1
    if data.weekly_rows:
        for weekly in data.weekly_rows:
            follow_up = "-" if weekly.disposition is ControlDisposition.PASS else _weekly_followup(weekly)
            _write_cells(
                ws,
                row_cursor,
                (
                    weekly.week_ending,
                    weekly.coverage,
                    _excel_money(weekly.pos_issue_variance),
                    _excel_money(weekly.pos_payment_variance),
                    _excel_money(weekly.pos_net_variance),
                    _excel_money(weekly.tender_variance),
                    weekly.disposition.value,
                    follow_up,
                ),
            )
            ws.cell(row_cursor, 1).number_format = DATE_FMT
            for column in range(3, 7):
                ws.cell(row_cursor, column).number_format = ACCOUNTING_FMT
            _style_weekly_row(ws, row_cursor, weekly.disposition, border)
            row_cursor += 1
    else:
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=8)
        ws.cell(
            row_cursor,
            1,
            "No weekly detail was available. See Close Controls for the blocking assessment.",
        )
        ws.cell(row_cursor, 1).fill = PatternFill("solid", fgColor=_RED)
        ws.cell(row_cursor, 1).font = Font(
            name=_FONT_NAME, color=_DARK_RED, bold=True, size=10
        )
        ws.cell(row_cursor, 1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row_cursor, 1).border = border
        row_cursor += 1

    period_status = _period_status_text(data)
    _write_cells(
        ws,
        row_cursor,
        (
            "PERIOD",
            "Independent monthly controls",
            None,
            None,
            _excel_money(data.period_pos_net_variance),
            _excel_money(data.period_tender_variance),
            period_status,
            "Monthly controls assessed independently.",
        ),
    )
    for column in range(3, 7):
        ws.cell(row_cursor, column).number_format = ACCOUNTING_FMT
    _style_total_row(ws, row_cursor, border)
    row_cursor += 2

    _section_title(ws, row_cursor, "Variance Summary")
    row_cursor += 1
    largest = _largest_weekly_variance(data.weekly_rows)
    ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3)
    ws.merge_cells(start_row=row_cursor, start_column=5, end_row=row_cursor, end_column=8)
    ws.cell(row_cursor, 1, "Largest Weekly Variance")
    ws.cell(row_cursor, 4, _excel_money(largest[0]) if largest else None)
    ws.cell(row_cursor, 5, largest[1] if largest else "Not available")
    ws.cell(row_cursor, 4).number_format = ACCOUNTING_FMT
    _style_highlight_row(ws, row_cursor, border)
    row_cursor += 1
    ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3)
    ws.merge_cells(start_row=row_cursor, start_column=5, end_row=row_cursor, end_column=8)
    ws.cell(row_cursor, 1, "Period POS Net Variance")
    ws.cell(row_cursor, 4, _excel_money(data.period_pos_net_variance))
    ws.cell(row_cursor, 5, _disposition_text(data.period_pos_disposition))
    ws.cell(row_cursor, 4).number_format = ACCOUNTING_FMT
    _style_highlight_row(ws, row_cursor, border)
    row_cursor += 2

    _section_title(ws, row_cursor, "Review Items")
    row_cursor += 1
    review_items = _review_items(data)
    if review_items:
        for label, disposition, message in review_items:
            ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=2)
            ws.merge_cells(start_row=row_cursor, start_column=3, end_row=row_cursor, end_column=8)
            ws.cell(row_cursor, 1, label)
            ws.cell(row_cursor, 3, message)
            _style_review_item_row(ws, row_cursor, disposition, border)
            row_cursor += 1
    else:
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=8)
        ws.cell(row_cursor, 1, "No review items.")
        _style_neutral_merged_row(ws, row_cursor, border)
        row_cursor += 1

    row_cursor += 1
    _section_title(ws, row_cursor, "Evidence and Audit Trail")
    row_cursor += 1
    for label, note in _evidence_audit_rows(data):
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=2)
        ws.merge_cells(start_row=row_cursor, start_column=3, end_row=row_cursor, end_column=8)
        ws.cell(row_cursor, 1, label)
        ws.cell(row_cursor, 3, note)
        _style_evidence_row(ws, row_cursor, border)
        row_cursor += 1

    widths = {"A": 14, "B": 22, "C": 14, "D": 15, "E": 14, "F": 14, "G": 15, "H": 28}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    _replace_unicode_dashes(ws, row_cursor)
    _complete_merged_range_styles(ws, border)
    _apply_default_typography(ws, row_cursor)

    ws.freeze_panes = None
    ws.print_area = f"A1:H{row_cursor - 1}"
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 0
    ws.page_setup.fitToHeight = 0
    ws.page_setup.scale = 85
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
    ws.page_margins = PageMargins(
        left=0.2,
        right=0.2,
        top=0.3,
        bottom=0.38,
        header=0.12,
        footer=0.18,
    )
    ws.oddFooter.left.text = f"Generated {data.generated_at:%m/%d/%Y %I:%M %p}"
    ws.oddFooter.center.text = "Page &P of &N"
    ws.oddFooter.right.text = f"{config.location_name} | {data.period}"


def _write_page_heading(ws: Worksheet, row: int, heading: str, data: MonthlyCloseReportData) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    title = ws.cell(row, 1, heading)
    title.font = Font(name=_FONT_NAME, bold=True, color="FFFFFF", size=19)
    title.fill = PatternFill("solid", fgColor=_NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 32

    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=8)
    subtitle = ws.cell(
        row + 1,
        1,
        (
            f"{data.period} | {data.period_start:%B %d, %Y} to {data.period_end:%B %d, %Y} | "
            f"Generated {data.generated_at:%B %d, %Y at %I:%M %p}"
        ),
    )
    subtitle.font = Font(name=_FONT_NAME, color="44546A", italic=True, size=10)
    subtitle.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row + 1].height = 18


def _section_title(ws: Worksheet, row: int, title: str) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row, 1, title)
    cell.font = Font(name=_FONT_NAME, bold=True, color=_NAVY, size=11)
    cell.fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20


def _style_header(ws: Worksheet, row: int, max_column: int) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    side = Side(style="thin", color="C9D5E3")
    for column in range(1, max_column + 1):
        cell = ws.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.font = Font(name=_FONT_NAME, bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=side, bottom=side, left=side, right=side)
    ws.row_dimensions[row].height = 26


def _style_assessed_row(ws: Worksheet, row: int, disposition: ControlDisposition, border) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    for column in range(1, 9):
        cell = ws.cell(row, column)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.font = Font(name=_FONT_NAME, size=10, color=_TEXT)
    ws.cell(row, 1).font = Font(name=_FONT_NAME, bold=True, size=10, color=_TEXT)
    status_cell = ws.cell(row, 4)
    status_cell.fill = PatternFill("solid", fgColor=_disposition_fill(disposition))
    status_cell.font = Font(
        name=_FONT_NAME,
        bold=True,
        color=_disposition_text_color(disposition),
        size=10,
    )
    status_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24


def _style_action_row(ws: Worksheet, row: int, disposition: ControlDisposition, border) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    for column in range(1, 9):
        cell = ws.cell(row, column)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.font = Font(name=_FONT_NAME, size=10, bold=(column == 1), color=_TEXT)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=_disposition_fill(disposition))
    ws.cell(row, 1).font = Font(
        name=_FONT_NAME,
        bold=True,
        color=_disposition_text_color(disposition),
        size=10,
    )
    ws.row_dimensions[row].height = 24


def _style_weekly_row(ws: Worksheet, row: int, disposition: ControlDisposition, border) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    for column in range(1, 9):
        cell = ws.cell(row, column)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.font = Font(name=_FONT_NAME, size=10, color=_TEXT)
    status_cell = ws.cell(row, 7)
    status_cell.fill = PatternFill("solid", fgColor=_disposition_fill(disposition))
    status_cell.font = Font(
        name=_FONT_NAME,
        bold=True,
        color=_disposition_text_color(disposition),
        size=10,
    )
    status_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    note_length = len(str(ws.cell(row, 8).value or ""))
    ws.row_dimensions[row].height = min(40, 24 + (note_length // 70) * 8)


def _style_total_row(ws: Worksheet, row: int, border) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    for column in range(1, 9):
        cell = ws.cell(row, column)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
        cell.font = Font(name=_FONT_NAME, bold=True, size=10, color=_TEXT)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 26


def _style_highlight_row(ws: Worksheet, row: int, border) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    for column in range(1, 9):
        cell = ws.cell(row, column)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor=_GRAY if row % 2 else "FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.font = Font(
            name=_FONT_NAME,
            bold=(column == 1),
            size=10,
            color=_TEXT,
        )
    ws.row_dimensions[row].height = 24


def _style_review_item_row(
    ws: Worksheet,
    row: int,
    disposition: ControlDisposition,
    border,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    for column in range(1, 9):
        cell = ws.cell(row, column)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor="FFFFFF")
        cell.font = Font(name=_FONT_NAME, size=10, color=_TEXT)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    label = ws.cell(row, 1)
    label.fill = PatternFill("solid", fgColor=_disposition_fill(disposition))
    label.font = Font(
        name=_FONT_NAME,
        bold=True,
        size=10,
        color=_disposition_text_color(disposition),
    )
    ws.row_dimensions[row].height = 26


def _style_evidence_row(ws: Worksheet, row: int, border) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    for column in range(1, 9):
        cell = ws.cell(row, column)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor=_GRAY if row % 2 else "FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.font = Font(name=_FONT_NAME, size=9.5, color=_TEXT)
    ws.cell(row, 1).font = Font(
        name=_FONT_NAME,
        bold=True,
        size=9.5,
        color=_NAVY,
    )
    note_length = len(str(ws.cell(row, 3).value or ""))
    ws.row_dimensions[row].height = min(42, 25 + (note_length // 120) * 8)


def _style_neutral_merged_row(ws: Worksheet, row: int, border) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    for column in range(1, 9):
        cell = ws.cell(row, column)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor=_GRAY)
        cell.font = Font(name=_FONT_NAME, size=10, color=_TEXT)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 24


def _write_cells(ws: Worksheet, row: int, values: Sequence[object]) -> None:
    for column, value in enumerate(values, start=1):
        ws.cell(row, column, value)


def _set_document_properties(
    ws: Worksheet,
    data: MonthlyCloseReportData,
    heading: str,
) -> None:
    properties = ws.parent.properties
    properties.title = _ascii_dashes(f"{heading} {data.period} Monthly Close Report")
    properties.subject = "Gift Card Monthly Close Reconciliation"
    properties.creator = "Gift Card Reconciliation Close Control"
    properties.description = _ascii_dashes(
        f"Executive accounting close certificate for "
        f"{data.assessment.store_config.location_name}, {data.period}; "
        f"generated {data.generated_at:%B %d, %Y at %I:%M %p}."
    )


def _replace_unicode_dashes(ws: Worksheet, last_row: int) -> None:
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=8):
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = _ascii_dashes(cell.value)


def _complete_merged_range_styles(ws: Worksheet, border) -> None:
    """Apply the anchor style to every cell in each merged range."""

    for merged in tuple(ws.merged_cells.ranges):
        anchor = ws.cell(merged.min_row, merged.min_col)
        for row in ws.iter_rows(
            min_row=merged.min_row,
            max_row=merged.max_row,
            min_col=merged.min_col,
            max_col=merged.max_col,
        ):
            for cell in row:
                cell.border = copy(border)
                cell.alignment = copy(anchor.alignment)
                cell.fill = copy(anchor.fill)
                cell.font = copy(anchor.font)


def _apply_default_typography(ws: Worksheet, last_row: int) -> None:
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=8):
        for cell in row:
            font = copy(cell.font)
            font.name = _FONT_NAME
            if font.sz is None:
                font.sz = 10
            cell.font = font


def _ascii_dashes(value: str) -> str:
    return str(value).translate(
        str.maketrans(
            {
                "\u2010": "-",
                "\u2011": "-",
                "\u2012": "-",
                "\u2013": "-",
                "\u2014": "-",
                "\u2212": "-",
            }
        )
    )


def _concise_control_message(control: ControlOutcome) -> str:
    message = re.sub(r"\s+", " ", _ascii_dashes(control.message)).strip()
    if len(message) <= 180:
        return message
    return f"{message[:177].rstrip()}..."


def _weekly_open_groups(controls: Sequence[ControlOutcome]) -> tuple[str, ...]:
    groups: list[str] = []
    for control in controls:
        label = _weekly_group_label(control.label)
        if label not in groups:
            groups.append(label)
    return tuple(groups)


def _weekly_group_label(label: str) -> str:
    text = _ascii_dashes(label)
    match = re.search(r"week ending\s+(\d{1,2}/\d{1,2}/\d{4})", text, re.IGNORECASE)
    return f"Week ending {match.group(1)}" if match else "Weekly controls"


def _weekly_followup(row: WeeklyCloseReportRow) -> str:
    metrics = (
        ("POS issue", row.pos_issue_variance),
        ("POS payment", row.pos_payment_variance),
        ("POS net", row.pos_net_variance),
        ("Tender", row.tender_variance),
    )
    values = [
        f"{label} {_format_signed_money(value)}"
        for label, value in metrics
        if value is not None and value != 0
    ]
    if values:
        return "; ".join(values)
    note = re.sub(r"\s+", " ", _ascii_dashes(row.evidence_note)).strip()
    return note or f"{row.disposition.value.title()} assessed weekly control."


def _format_signed_money(value: Decimal) -> str:
    sign = "-" if value < 0 else "+"
    return f"{sign}${abs(value):,.2f}"


def _review_items(
    data: MonthlyCloseReportData,
) -> tuple[tuple[str, ControlDisposition, str], ...]:
    items: list[tuple[str, ControlDisposition, str]] = []
    seen: set[str] = set()

    def add(label: str, disposition: ControlDisposition, message: str) -> None:
        normalized = _normalized_review_text(f"{label}: {message}")
        if normalized not in seen:
            seen.add(normalized)
            items.append((_ascii_dashes(label), disposition, _ascii_dashes(message)))

    weekly_dates: set[str] = set()
    for weekly in data.weekly_rows:
        if weekly.disposition is ControlDisposition.PASS:
            continue
        label = (
            f"Week ending {weekly.week_ending:%m/%d/%Y}"
            if weekly.week_ending
            else "Unknown week"
        )
        weekly_dates.add(label.casefold())
        add(label, weekly.disposition, _weekly_followup(weekly))

    weekly_controls = tuple(
        control
        for control in data.assessment.controls
        if not control.passed
        and _is_weekly_child_control(control, data.weekly_control_codes)
    )
    for group in _weekly_open_groups(weekly_controls):
        if group.casefold() in weekly_dates:
            continue
        grouped = tuple(
            control
            for control in weekly_controls
            if _weekly_group_label(control.label) == group
        )
        messages = []
        for control in grouped:
            message = _concise_control_message(control)
            if message not in messages:
                messages.append(message)
        add(group, _worst_disposition(grouped), "; ".join(messages))

    for control in data.assessment.controls:
        if control.passed or _is_weekly_child_control(control, data.weekly_control_codes):
            continue
        add(control.label, control.disposition, _concise_control_message(control))

    assessment_exception_keys = {
        _normalized_review_text(f"{control.label}: {control.message}")
        for control in data.assessment.controls
        if not control.passed
    }
    for severity, message in data.exceptions:
        if _normalized_review_text(str(message)) in assessment_exception_keys:
            continue
        add("Exception", _disposition_from_severity(severity), str(message))
    return tuple(items)


def _normalized_review_text(value: str) -> str:
    return re.sub(r"\s+", " ", _ascii_dashes(str(value))).strip().casefold()


def _evidence_audit_rows(data: MonthlyCloseReportData) -> tuple[tuple[str, str], ...]:
    source_package = "; ".join(_ascii_dashes(label) for label in data.source_labels)
    notes: list[str] = []
    for note in data.evidence_notes:
        normalized = re.sub(r"\s+", " ", _ascii_dashes(str(note))).strip()
        if normalized and normalized not in notes:
            notes.append(normalized)
    archive_note = " ".join(notes) or (
        "Source evidence and report hashes are recorded in the close manifest."
    )
    close_policy = (
        "Overall status comes from the centralized close assessment. "
        "Darden settlement matching is one control, not the overall close decision."
    )
    return (
        ("Source package", source_package),
        ("Archive and hash", archive_note),
        ("Close policy", close_policy),
    )


def _excel_money(value: Decimal | None) -> float | None:
    return None if value is None else float(value)


def _darden_variance(assessment: CloseAssessment) -> Decimal | None:
    return next(
        (control.variance for control in assessment.controls if control.code == "darden_summary_match"),
        None,
    )


def _executive_controls(data: MonthlyCloseReportData) -> tuple[ControlOutcome, ...]:
    # Page 1 is an executive certificate, while page 2 carries every weekly
    # child control. Keep a stable set of core controls and add any non-weekly
    # exception so a failure can never disappear from the matrix.
    core_codes = {
        "darden_summary_match",
        "summary_identity",
        "activity_coverage",
        "micros_coverage",
        "tender_evidence",
        "archive_integrity",
        "summary_activity_net_gift_card_impact",
        "period_pos_period_pos_net_gift_card_impact",
        "period_tender_period_tender",
    }
    return tuple(
        control
        for control in data.assessment.controls
        if control.code in core_codes
        or (
            not control.passed
            and not _is_weekly_child_control(control, data.weekly_control_codes)
        )
    )


def _worst_disposition(controls: Sequence[ControlOutcome]) -> ControlDisposition:
    dispositions = {control.disposition for control in controls}
    if ControlDisposition.BLOCK in dispositions:
        return ControlDisposition.BLOCK
    if ControlDisposition.REVIEW in dispositions:
        return ControlDisposition.REVIEW
    return ControlDisposition.PASS


def _is_weekly_child_control(
    control: ControlOutcome,
    explicit_weekly_codes: frozenset[str],
) -> bool:
    if explicit_weekly_codes:
        return control.code in explicit_weekly_codes
    if control.code in REQUIRED_CLOSE_INTEGRITY_CODES:
        return False
    if control.code == "darden_summary_match" or control.code.startswith("summary_activity_"):
        return False
    text = f"{control.code} {control.label}".lower()
    if "period" in text or "monthly" in text:
        return False
    if not control.code.startswith(("pos_", "tender_")):
        return False
    return any(token in text for token in ("week", "weekly", "week_ending", "2026", "2027"))


def _status_fill(status: CloseStatus) -> str:
    return {
        CloseStatus.CLOSED: _GREEN,
        CloseStatus.CLOSED_WITH_REVIEW: _AMBER,
        CloseStatus.REVIEW_REQUIRED: _RED,
    }[status]


def _status_text(status: CloseStatus) -> str:
    return {
        CloseStatus.CLOSED: _DARK_GREEN,
        CloseStatus.CLOSED_WITH_REVIEW: _DARK_AMBER,
        CloseStatus.REVIEW_REQUIRED: _DARK_RED,
    }[status]


def _disposition_fill(disposition: ControlDisposition) -> str:
    return {
        ControlDisposition.PASS: _GREEN,
        ControlDisposition.REVIEW: _AMBER,
        ControlDisposition.BLOCK: _RED,
    }[disposition]


def _disposition_text_color(disposition: ControlDisposition) -> str:
    return {
        ControlDisposition.PASS: _DARK_GREEN,
        ControlDisposition.REVIEW: _DARK_AMBER,
        ControlDisposition.BLOCK: _DARK_RED,
    }[disposition]


def _disposition_text(disposition: ControlDisposition | None) -> str:
    return "NOT AVAILABLE" if disposition is None else disposition.value


def _period_status_text(data: MonthlyCloseReportData) -> str:
    return (
        f"POS: {_disposition_text(data.period_pos_disposition)}; "
        f"Tender: {_disposition_text(data.period_tender_disposition)}"
    )


def _largest_weekly_variance(
    weekly_rows: Sequence[WeeklyCloseReportRow],
) -> tuple[Decimal, str] | None:
    candidates: list[tuple[Decimal, str]] = []
    for row in weekly_rows:
        week = row.week_ending.strftime("%m/%d/%Y") if row.week_ending else "Unknown week"
        values = (
            ("POS issue", row.pos_issue_variance),
            ("POS payment", row.pos_payment_variance),
            ("POS net", row.pos_net_variance),
            ("Tender", row.tender_variance),
        )
        for label, value in values:
            if value is not None:
                candidates.append((abs(value), f"{label} | Week ending {week} | Reported {value:+,.2f}"))
    if not candidates:
        return None
    largest = max(candidates, key=lambda item: item[0])
    if largest[0] == 0:
        return Decimal("0.00"), "No weekly variance"
    return largest


def _disposition_from_severity(severity: str) -> ControlDisposition:
    normalized = str(severity).strip().upper()
    if normalized in {"BLOCK", "ERROR", "REVIEW REQUIRED", "FAIL", "FAILED"}:
        return ControlDisposition.BLOCK
    if normalized in {"REVIEW", "WARNING", "WARN", "MINOR VARIANCE"}:
        return ControlDisposition.REVIEW
    return ControlDisposition.PASS
