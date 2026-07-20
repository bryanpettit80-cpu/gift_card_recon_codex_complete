from __future__ import annotations

import os
import re
import tempfile
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Mapping

from gift_card_recon.store_config import REVIEW_VARIANCE_LIMIT


EXPLANATION_SHEET_NAME = "Variance Explanation"
EXPLANATION_SHEET = EXPLANATION_SHEET_NAME
IDENTITY_SHEET_NAME = "_identity"
EXPLANATION_INPUT_CELL = "B15"
MAX_EXPLANATION_LENGTH = 500

SCHEMA_VERSION = 1
DOCUMENT_TYPE = "gift_card_weekly_variance_explanation"
ACCOUNTING_FORMAT = '$#,##0.00;($#,##0.00);$0.00'

_MONEY_CENTS = Decimal("0.01")
_STORE_PATTERN = re.compile(r"^[0-9]{4}$")
_ILLEGAL_EXCEL_TEXT = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
_VISIBLE_IDENTITY_CELLS = {
    "store": "B4",
    "week_start": "D4",
    "week_end": "F4",
}
_CONTROL_ROWS = {
    "pos_issue_variance": ("POS gift card issue", 8),
    "pos_payment_variance": ("POS gift card payment", 9),
    "pos_net_variance": ("POS net", 10),
    "tender_variance": ("Tender detail", 11),
}
_IDENTITY_FIELDS = (
    "schema_version",
    "document_type",
    "store",
    "week_start",
    "week_end",
    "pos_issue_variance",
    "pos_payment_variance",
    "pos_net_variance",
    "tender_variance",
    "input_sheet",
    "input_cell",
)


@dataclass(frozen=True, slots=True)
class WeeklyVarianceExplanation:
    store: str
    week_start: date
    week_end: date
    pos_issue_variance: Decimal
    pos_payment_variance: Decimal
    pos_net_variance: Decimal
    tender_variance: Decimal
    explanation: str = ""

    def __post_init__(self) -> None:
        store = _normalize_store(self.store)
        week_start = _normalize_date(self.week_start, label="week_start")
        week_end = _normalize_date(self.week_end, label="week_end")
        _validate_week(week_start, week_end)
        object.__setattr__(self, "store", store)
        object.__setattr__(self, "week_start", week_start)
        object.__setattr__(self, "week_end", week_end)
        for field_name in _CONTROL_ROWS:
            object.__setattr__(
                self,
                field_name,
                _normalize_money(getattr(self, field_name), label=field_name),
            )
        object.__setattr__(self, "explanation", _normalize_explanation(self.explanation))

    @property
    def controls(self) -> Mapping[str, Decimal]:
        return {
            field_name: getattr(self, field_name)
            for field_name in _CONTROL_ROWS
        }

    @property
    def requires_explanation(self) -> bool:
        return any(abs(value) > REVIEW_VARIANCE_LIMIT for value in self.controls.values())


def variance_explanation_path(
    monthly_period_dir: Path,
    store: str | int,
    week_end: date,
) -> Path:
    """Return the deterministic editable-input path for one store/week."""

    normalized_store = _normalize_store(store)
    normalized_end = _normalize_date(week_end, label="week_end")
    if normalized_end.weekday() != 6:
        raise ValueError("week_end must be a Sunday.")
    filename = (
        f"Weekly_Variance_{normalized_store}_"
        f"{normalized_end:%Y-%m-%d}.xlsx"
    )
    return Path(monthly_period_dir) / "Variance Explanations" / filename


def variance_control_mismatch_details(
    data: WeeklyVarianceExplanation,
    *,
    pos_issue_variance: Decimal,
    pos_payment_variance: Decimal,
    pos_net_variance: Decimal,
    tender_variance: Decimal,
) -> tuple[str, ...]:
    """Describe exact-cent differences between a form and current controls."""

    expected = {
        "pos_issue_variance": _normalize_money(
            pos_issue_variance,
            label="expected pos_issue_variance",
        ),
        "pos_payment_variance": _normalize_money(
            pos_payment_variance,
            label="expected pos_payment_variance",
        ),
        "pos_net_variance": _normalize_money(
            pos_net_variance,
            label="expected pos_net_variance",
        ),
        "tender_variance": _normalize_money(
            tender_variance,
            label="expected tender_variance",
        ),
    }
    details: list[str] = []
    for field_name, (label, _row) in _CONTROL_ROWS.items():
        form_value = getattr(data, field_name)
        current_value = expected[field_name]
        if form_value != current_value:
            details.append(
                f"{label}: form {form_value:+,.2f}, current {current_value:+,.2f}"
            )
    return tuple(details)


def write_variance_explanation_workbook(
    data: WeeklyVarianceExplanation,
    output_path: Path,
    *,
    overwrite: bool = False,
) -> Path:
    """Create the editable companion workbook without touching weekly evidence."""

    if not isinstance(data, WeeklyVarianceExplanation):
        raise TypeError("data must be a WeeklyVarianceExplanation.")
    output = Path(output_path)
    if output.suffix.casefold() != ".xlsx":
        raise ValueError(f"Variance explanation workbook must end in .xlsx: {output}")
    if output.exists() and not overwrite:
        raise FileExistsError(f"Variance explanation workbook already exists: {output}")
    if _would_be_excel_formula(data.explanation):
        raise ValueError("Explanation text cannot be an Excel formula.")

    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:  # pragma: no cover - declared runtime dependency
        raise RuntimeError(
            "openpyxl is required to create variance explanation workbooks."
        ) from exc

    workbook = Workbook()
    visible = workbook.active
    visible.title = EXPLANATION_SHEET_NAME
    visible.sheet_view.showGridLines = False
    identity = workbook.create_sheet(IDENTITY_SHEET_NAME)

    navy = "17365D"
    light_blue = "D9EAF7"
    pale_red = "F4CCCC"
    dark_red = "9C0006"
    bright_yellow = "FFFF00"
    gray = "F2F2F2"
    text = "333333"
    thin_side = Side(style="thin", color="B7C9DB")
    border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    visible.merge_cells("A1:F2")
    title = visible["A1"]
    title.value = "WEEKLY VARIANCE EXPLANATION"
    title.fill = PatternFill("solid", fgColor=navy)
    title.font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal="center", vertical="center")
    for row in range(1, 3):
        visible.row_dimensions[row].height = 24

    for label_cell, value_cell, label, value in (
        ("A4", "B4", "Store", data.store),
        ("C4", "D4", "Week Start", data.week_start),
        ("E4", "F4", "Week End", data.week_end),
    ):
        visible[label_cell] = label
        visible[value_cell] = value
        visible[label_cell].fill = PatternFill("solid", fgColor=navy)
        visible[label_cell].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        visible[value_cell].fill = PatternFill("solid", fgColor=gray)
        visible[value_cell].font = Font(name="Arial", size=10, bold=True, color=text)
        visible[label_cell].alignment = Alignment(horizontal="center", vertical="center")
        visible[value_cell].alignment = Alignment(horizontal="center", vertical="center")
        visible[label_cell].border = border
        visible[value_cell].border = border
    visible["D4"].number_format = "mm/dd/yyyy"
    visible["F4"].number_format = "mm/dd/yyyy"
    visible.row_dimensions[4].height = 24

    visible.merge_cells("A6:F6")
    visible["A6"] = "Weekly Control Summary"
    visible["A6"].fill = PatternFill("solid", fgColor=light_blue)
    visible["A6"].font = Font(name="Arial", size=11, bold=True, color=navy)
    visible["A6"].alignment = Alignment(vertical="center")
    visible["A6"].border = border

    visible.merge_cells("A7:C7")
    visible.merge_cells("D7:E7")
    visible["A7"] = "Control"
    visible["D7"] = "Variance"
    visible["F7"] = f"Over ${REVIEW_VARIANCE_LIMIT:.2f}?"
    for column in range(1, 7):
        cell = visible.cell(7, column)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    visible.row_dimensions[7].height = 26

    for field_name, (label, row) in _CONTROL_ROWS.items():
        value = getattr(data, field_name)
        visible.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        visible.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        visible.cell(row, 1, label)
        visible.cell(row, 4, float(value))
        visible.cell(row, 6, "YES" if abs(value) > REVIEW_VARIANCE_LIMIT else "No")
        visible.cell(row, 4).number_format = ACCOUNTING_FORMAT
        triggered = abs(value) > REVIEW_VARIANCE_LIMIT
        row_fill = pale_red if triggered else "FFFFFF"
        for column in range(1, 7):
            cell = visible.cell(row, column)
            cell.fill = PatternFill("solid", fgColor=row_fill)
            cell.font = Font(
                name="Arial",
                size=10,
                bold=(column in {1, 6}),
                color=dark_red if triggered else text,
            )
            cell.alignment = Alignment(
                horizontal="center" if column >= 4 else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border
        visible.row_dimensions[row].height = 23

    visible.merge_cells("A13:F13")
    visible["A13"] = (
        "Enter a plain-language explanation in the yellow box. "
        "The explanation documents the variance; it does not approve or clear it."
    )
    visible["A13"].fill = PatternFill("solid", fgColor=light_blue)
    visible["A13"].font = Font(name="Arial", size=10, bold=True, color=navy)
    visible["A13"].alignment = Alignment(wrap_text=True, vertical="center")
    visible["A13"].border = border
    visible.row_dimensions[13].height = 34

    visible["A15"] = "Explanation"
    visible["A15"].fill = PatternFill("solid", fgColor=navy)
    visible["A15"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    visible["A15"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    visible["A15"].border = border
    visible.merge_cells("B15:F19")
    input_cell = visible[EXPLANATION_INPUT_CELL]
    input_cell.value = data.explanation or None
    input_cell.fill = PatternFill("solid", fgColor=bright_yellow)
    input_cell.font = Font(name="Arial", size=11, color="000000")
    input_cell.alignment = Alignment(wrap_text=True, vertical="top")
    input_cell.border = border
    input_cell.protection = Protection(locked=False)
    for row in range(15, 20):
        visible.row_dimensions[row].height = 24

    validation = DataValidation(
        type="textLength",
        operator="lessThanOrEqual",
        formula1=str(MAX_EXPLANATION_LENGTH),
        allow_blank=True,
    )
    validation.error = f"Explanation cannot exceed {MAX_EXPLANATION_LENGTH} characters."
    validation.errorTitle = "Explanation too long"
    validation.prompt = (
        f"Enter up to {MAX_EXPLANATION_LENGTH} characters. Press Alt+Enter for a new line."
    )
    validation.promptTitle = "Variance explanation"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    visible.add_data_validation(validation)
    validation.add(input_cell)

    visible.merge_cells("A21:F22")
    visible["A21"] = (
        "This companion workbook is a monthly-close input. The completed weekly report "
        "and its archived evidence package remain unchanged and hash-verifiable."
    )
    visible["A21"].fill = PatternFill("solid", fgColor=gray)
    visible["A21"].font = Font(name="Arial", size=9, italic=True, color=text)
    visible["A21"].alignment = Alignment(wrap_text=True, vertical="center")
    visible["A21"].border = border

    # Highlight any future edit that pushes a displayed variance above the shared limit.
    for _field_name, (_label, row) in _CONTROL_ROWS.items():
        visible.conditional_formatting.add(
            f"D{row}",
            CellIsRule(
                operator="greaterThan",
                formula=[str(REVIEW_VARIANCE_LIMIT)],
                fill=PatternFill("solid", fgColor=pale_red),
            ),
        )
        visible.conditional_formatting.add(
            f"D{row}",
            CellIsRule(
                operator="lessThan",
                formula=[str(-REVIEW_VARIANCE_LIMIT)],
                fill=PatternFill("solid", fgColor=pale_red),
            ),
        )

    visible.column_dimensions["A"].width = 18
    visible.column_dimensions["B"].width = 18
    visible.column_dimensions["C"].width = 18
    visible.column_dimensions["D"].width = 17
    visible.column_dimensions["E"].width = 17
    visible.column_dimensions["F"].width = 16
    visible.freeze_panes = None
    visible.print_area = "A1:F22"
    visible.print_options.horizontalCentered = True
    visible.page_setup.orientation = "portrait"
    visible.page_setup.paperSize = visible.PAPERSIZE_LETTER
    visible.page_setup.fitToWidth = 1
    visible.page_setup.fitToHeight = 1
    visible.sheet_properties.pageSetUpPr.fitToPage = True
    visible.sheet_properties.pageSetUpPr.autoPageBreaks = False
    visible.oddFooter.center.text = "Page &P of &N"
    visible.oddFooter.right.text = f"Store {data.store} | Week ending {data.week_end:%m/%d/%Y}"
    visible.protection.sheet = True
    visible.protection.selectLockedCells = False
    visible.protection.selectUnlockedCells = True

    identity_values: dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "document_type": DOCUMENT_TYPE,
        "store": data.store,
        "week_start": data.week_start.isoformat(),
        "week_end": data.week_end.isoformat(),
        "pos_issue_variance": _decimal_text(data.pos_issue_variance),
        "pos_payment_variance": _decimal_text(data.pos_payment_variance),
        "pos_net_variance": _decimal_text(data.pos_net_variance),
        "tender_variance": _decimal_text(data.tender_variance),
        "input_sheet": EXPLANATION_SHEET_NAME,
        "input_cell": EXPLANATION_INPUT_CELL,
    }
    for row, field_name in enumerate(_IDENTITY_FIELDS, start=1):
        identity.cell(row, 1, field_name)
        identity.cell(row, 2, identity_values[field_name])
    identity.sheet_state = "veryHidden"
    workbook.active = workbook.sheetnames.index(EXPLANATION_SHEET_NAME)
    workbook.security.lockStructure = True
    workbook.properties.title = (
        f"Store {data.store} Week Ending {data.week_end:%Y-%m-%d} Variance Explanation"
    )
    workbook.properties.subject = "Weekly Gift Card Variance Explanation Input"
    workbook.properties.creator = "Gift Card Reconciliation"
    workbook.properties.description = (
        "Editable explanation input that accompanies, but does not modify, immutable weekly evidence."
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    try:
        handle, temp_name = tempfile.mkstemp(
            prefix=".variance-explanation-",
            suffix=".xlsx",
            dir=output.parent,
        )
        os.close(handle)
        temporary = Path(temp_name)
        workbook.save(temporary)
        if output.exists() and not overwrite:
            raise FileExistsError(f"Variance explanation workbook already exists: {output}")
        os.replace(temporary, output)
        temporary = None
    finally:
        workbook.close()
        if temporary is not None:
            temporary.unlink(missing_ok=True)
    return output


def read_variance_explanation_workbook(
    path: Path,
    *,
    expected_store: str | int | None = None,
    expected_week_start: date | None = None,
    expected_week_end: date | None = None,
    require_text: bool = True,
) -> WeeklyVarianceExplanation:
    """Read and strictly validate one companion explanation workbook."""

    workbook_path = Path(path)
    if workbook_path.suffix.casefold() != ".xlsx":
        raise ValueError(
            f"Variance explanation workbook must end in .xlsx: {workbook_path}"
        )
    if not workbook_path.is_file():
        raise ValueError(f"Variance explanation workbook is missing: {workbook_path}")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - declared runtime dependency
        raise RuntimeError(
            "openpyxl is required to read variance explanation workbooks."
        ) from exc

    try:
        workbook = load_workbook(
            workbook_path,
            data_only=False,
            read_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise ValueError(
            f"Could not open variance explanation workbook {workbook_path}: {exc}"
        ) from exc

    try:
        _reject_formulas(workbook, workbook_path)
        if EXPLANATION_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Missing required worksheet {EXPLANATION_SHEET_NAME!r}: {workbook_path}"
            )
        if IDENTITY_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Missing required identity worksheet {IDENTITY_SHEET_NAME!r}: {workbook_path}"
            )
        visible = workbook[EXPLANATION_SHEET_NAME]
        identity_sheet = workbook[IDENTITY_SHEET_NAME]
        if identity_sheet.sheet_state != "veryHidden":
            raise ValueError("Variance explanation identity worksheet is not very hidden.")
        identity = _read_identity(identity_sheet)
        if identity["schema_version"] != SCHEMA_VERSION:
            raise ValueError(
                "Unsupported variance explanation schema version: "
                f"{identity['schema_version']!r}."
            )
        if identity["document_type"] != DOCUMENT_TYPE:
            raise ValueError("Workbook is not a weekly variance explanation input.")
        if identity["input_sheet"] != EXPLANATION_SHEET_NAME:
            raise ValueError("Variance explanation input-sheet identity does not match the schema.")
        if identity["input_cell"] != EXPLANATION_INPUT_CELL:
            raise ValueError("Variance explanation input-cell identity does not match the schema.")

        store = _normalize_store(identity["store"])
        week_start = _parse_identity_date(identity["week_start"], label="week_start")
        week_end = _parse_identity_date(identity["week_end"], label="week_end")
        _validate_week(week_start, week_end)
        variances = {
            field_name: _normalize_money(identity[field_name], label=field_name)
            for field_name in _CONTROL_ROWS
        }

        visible_store = _normalize_store(visible[_VISIBLE_IDENTITY_CELLS["store"]].value)
        visible_start = _normalize_date(
            visible[_VISIBLE_IDENTITY_CELLS["week_start"]].value,
            label="visible week_start",
        )
        visible_end = _normalize_date(
            visible[_VISIBLE_IDENTITY_CELLS["week_end"]].value,
            label="visible week_end",
        )
        if visible_store != store or visible_start != week_start or visible_end != week_end:
            raise ValueError("Visible store/week values do not match the protected workbook identity.")
        for field_name, (_label, row) in _CONTROL_ROWS.items():
            visible_value = _normalize_money(
                visible.cell(row, 4).value,
                label=f"visible {field_name}",
            )
            if visible_value != variances[field_name]:
                raise ValueError(
                    f"Visible {field_name} does not match the protected workbook identity."
                )

        if expected_store is not None and store != _normalize_store(expected_store):
            raise ValueError(
                f"Variance explanation store mismatch: expected {expected_store}, found {store}."
            )
        if expected_week_start is not None:
            expected_start = _normalize_date(expected_week_start, label="expected_week_start")
            if week_start != expected_start:
                raise ValueError(
                    "Variance explanation week-start mismatch: "
                    f"expected {expected_start.isoformat()}, found {week_start.isoformat()}."
                )
        if expected_week_end is not None:
            expected_end = _normalize_date(expected_week_end, label="expected_week_end")
            if week_end != expected_end:
                raise ValueError(
                    "Variance explanation week-ending mismatch: "
                    f"expected {expected_end.isoformat()}, found {week_end.isoformat()}."
                )

        raw_explanation = visible[EXPLANATION_INPUT_CELL].value
        if raw_explanation is None:
            explanation = ""
        elif not isinstance(raw_explanation, str):
            raise ValueError("Variance explanation must be entered as text.")
        else:
            explanation = _normalize_explanation(raw_explanation)
        if require_text and not explanation:
            raise ValueError(
                f"Variance explanation is required in {EXPLANATION_SHEET_NAME}!"
                f"{EXPLANATION_INPUT_CELL}."
            )

        return WeeklyVarianceExplanation(
            store=store,
            week_start=week_start,
            week_end=week_end,
            explanation=explanation,
            **variances,
        )
    finally:
        workbook.close()


def _read_identity(sheet: Any) -> dict[str, Any]:
    identity: dict[str, Any] = {}
    for row in range(1, sheet.max_row + 1):
        key = sheet.cell(row, 1).value
        if key is None:
            continue
        if not isinstance(key, str) or key not in _IDENTITY_FIELDS:
            raise ValueError(f"Unexpected variance explanation identity field: {key!r}.")
        if key in identity:
            raise ValueError(f"Duplicate variance explanation identity field: {key!r}.")
        identity[key] = sheet.cell(row, 2).value
    missing = [field_name for field_name in _IDENTITY_FIELDS if field_name not in identity]
    if missing:
        raise ValueError(
            "Variance explanation identity is missing required fields: "
            + ", ".join(missing)
            + "."
        )
    raw_schema = identity["schema_version"]
    if isinstance(raw_schema, bool):
        raise ValueError("Variance explanation schema version is invalid.")
    try:
        identity["schema_version"] = int(raw_schema)
    except (TypeError, ValueError) as exc:
        raise ValueError("Variance explanation schema version is invalid.") from exc
    return identity


def _reject_formulas(workbook: Any, path: Path) -> None:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise ValueError(
                        "Formulas are not allowed in variance explanation workbooks: "
                        f"{path} ({sheet.title}!{cell.coordinate})."
                    )


def _normalize_store(value: object) -> str:
    if isinstance(value, bool):
        raise ValueError("Store must be a four-digit restaurant number.")
    store = str(value).strip()
    if not _STORE_PATTERN.fullmatch(store):
        raise ValueError("Store must be a four-digit restaurant number.")
    return store


def _normalize_date(value: object, *, label: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raise ValueError(f"{label} must be an Excel or Python date.")


def _parse_identity_date(value: object, *, label: str) -> date:
    if not isinstance(value, str):
        raise ValueError(f"Variance explanation identity {label} must be an ISO date.")
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(
            f"Variance explanation identity {label} is not a valid ISO date: {value!r}."
        ) from exc


def _validate_week(week_start: date, week_end: date) -> None:
    if week_start.weekday() != 0 or week_end.weekday() != 6:
        raise ValueError("Variance explanation dates must cover a Monday-Sunday week.")
    if (week_end - week_start).days != 6:
        raise ValueError("Variance explanation dates must cover exactly seven calendar days.")


def _normalize_money(value: object, *, label: str) -> Decimal:
    if isinstance(value, bool) or value is None:
        raise ValueError(f"{label} must be a finite money value.")
    try:
        amount = value if isinstance(value, Decimal) else Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{label} must be a finite money value.") from exc
    if not amount.is_finite():
        raise ValueError(f"{label} must be a finite money value.")
    return amount.quantize(_MONEY_CENTS, rounding=ROUND_HALF_UP)


def _normalize_explanation(value: object) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        raise ValueError("Variance explanation must be text.")
    normalized = value.replace("\r\n", "\n").replace("\r", "\n").strip()
    if _ILLEGAL_EXCEL_TEXT.search(normalized):
        raise ValueError("Variance explanation contains unsupported control characters.")
    if len(normalized) > MAX_EXPLANATION_LENGTH:
        raise ValueError(
            f"Variance explanation exceeds the {MAX_EXPLANATION_LENGTH}-character limit."
        )
    return normalized


def _would_be_excel_formula(value: str) -> bool:
    for character in value:
        if character.isspace():
            continue
        return character == "="
    return False


def _decimal_text(value: Decimal) -> str:
    return format(value, ".2f")
