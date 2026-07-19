from __future__ import annotations

import pytest
from openpyxl import Workbook

from gift_card_recon.excel_writer import _TrustedFormula, _write_row


@pytest.mark.parametrize(
    "value",
    [
        '=WEBSERVICE("https://example.invalid")',
        "+1+1",
        "-2+3",
        "@SUM(A1:A2)",
        ' \t=HYPERLINK("https://example.invalid")',
        "\u200b@SUM(A1:A2)",
    ],
)
def test_write_row_neutralizes_untrusted_formula_text(value: str) -> None:
    ws = Workbook().active

    _write_row(ws, 1, [value])

    assert ws["A1"].value == f"'{value}"
    assert ws["A1"].data_type == "s"


def test_write_row_preserves_only_explicitly_trusted_formulas() -> None:
    ws = Workbook().active

    _write_row(ws, 1, [_TrustedFormula("=SUM(A2:A3)"), "=SUM(B2:B3)"])

    assert ws["A1"].value == "=SUM(A2:A3)"
    assert ws["A1"].data_type == "f"
    assert ws["B1"].value == "'=SUM(B2:B3)"
    assert ws["B1"].data_type == "s"
