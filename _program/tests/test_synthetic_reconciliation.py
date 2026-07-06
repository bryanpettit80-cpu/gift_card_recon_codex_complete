from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from gift_card_recon.cli import main
import pytest

from gift_card_recon.parsers import ParseError, discover_input_files, parse_activity_file, parse_pos_controls, parse_summary
from gift_card_recon.reconcile import build_reconciliation
from gift_card_recon.utils import parse_date


WEEKLY = [
    ("05.03.2026 9354 Gift Card Activity.xlsx", "27-APR-2026", "03-MAY-2026", Decimal("2755.00"), Decimal("-25.00"), Decimal("-925.00"), Decimal("-6521.47"), Decimal("0.00")),
    ("05.10.2026 9354 Gift Card Activity.xlsx", "04-MAY-2026", "10-MAY-2026", Decimal("1815.00"), Decimal("-185.00"), Decimal("-1882.51"), Decimal("-10411.44"), Decimal("0.00")),
    ("05.17.2026 9354 Gift Card Activity.xlsx", "11-MAY-2026", "17-MAY-2026", Decimal("975.00"), Decimal("0.00"), Decimal("-1366.77"), Decimal("-8036.30"), Decimal("50.00")),
    ("05.24.2026 9354 Gift Card Activity.xlsx", "18-MAY-2026", "24-MAY-2026", Decimal("3997.00"), Decimal("0.00"), Decimal("-2710.59"), Decimal("-9502.64"), Decimal("0.00")),
    ("05.31.2026 9354 Gift Card Activity.xlsx", "25-MAY-2026", "31-MAY-2026", Decimal("2175.00"), Decimal("0.00"), Decimal("-1339.53"), Decimal("-7221.23"), Decimal("0.00")),
]

REPO_ROOT = Path(__file__).resolve().parents[2]


def test_full_reconciliation_synthetic(tmp_path: Path):
    input_dir = tmp_path / "input" / "9354" / "2026-05"
    summary_dir = input_dir / "summary"
    activity_dir = input_dir / "activity"
    summary_dir.mkdir(parents=True)
    activity_dir.mkdir(parents=True)

    summary_path = summary_dir / "05.31.2026 9354 Gift Card Summary.xlsx"
    create_summary(summary_path)
    for item in WEEKLY:
        create_activity(activity_dir / item[0], *item[1:])
    (input_dir / "pos_controls.csv").write_text("store,period,pos_gift_card_issue,pos_gift_card_payment\n9354,2026-05,11642.00,49869.75\n", encoding="utf-8")

    summary = parse_summary(summary_path, "9354")
    activities = [parse_activity_file(path, summary.conversion_promo_codes) for path in sorted(activity_dir.glob("*.xlsx"))]
    pos = parse_pos_controls(input_dir / "pos_controls.csv", "9354", "2026-05")
    result = build_reconciliation(store="9354", period="2026-05", period_end=parse_date("2026-05-31"), summary=summary, activities=activities, pos_controls=pos)

    assert result.activity_total_activations == Decimal("11507.00")
    assert result.activity_total_redemptions == Decimal("-49867.48")
    assert result.lines[0].pos_variance == Decimal("135.00")
    assert result.lines[1].pos_variance == Decimal("2.27")
    assert result.lines[2].pos_variance == Decimal("132.73")

    output_path = tmp_path / "out.xlsx"
    main([
        "--store", "9354",
        "--period", "2026-05",
        "--period-end", "2026-05-31",
        "--input-dir", str(input_dir),
        "--output-file", str(output_path),
    ])
    assert output_path.exists()
    wb = load_workbook(output_path, data_only=False)
    assert wb.sheetnames == ["Reconciliation", "Weekly Activity Detail", "Daily Activity Detail", "Raw Detail", "Source Files", "Exception Log"]
    assert wb["Reconciliation"]["B6"].value == 11507
    assert wb["Reconciliation"]["C5"].value == "GC Activity File Total"
    assert wb["Reconciliation"]["E6"].value == 11642
    assert wb["Reconciliation"]["A11"].value == "Gift Card Activity File Totals"
    assert wb["Reconciliation"]["A18"].value == "TOTAL"
    assert wb["Reconciliation"]["F18"].value == 11507
    assert wb["Reconciliation"]["I18"].value == -49867.48


def test_weekly_reconciliation_without_summary(tmp_path: Path):
    input_dir = tmp_path / "input" / "9354" / "2026-W22"
    activity_dir = input_dir / "activity"
    activity_dir.mkdir(parents=True)
    create_activity(activity_dir / WEEKLY[0][0], *WEEKLY[0][1:])
    (input_dir / "pos_controls.csv").write_text("store,period,pos_gift_card_issue,pos_gift_card_payment\n9354,2026-W22,2730.00,7446.47\n", encoding="utf-8")

    output_path = tmp_path / "weekly-no-summary.xlsx"
    main([
        "--mode", "weekly",
        "--store", "9354",
        "--period", "2026-W22",
        "--period-end", "2026-05-03",
        "--input-dir", str(input_dir),
        "--output-file", str(output_path),
    ])

    wb = load_workbook(output_path, data_only=False)
    assert wb.sheetnames == ["Reconciliation", "Weekly Activity Detail", "Daily Activity Detail", "Raw Detail", "Source Files", "Exception Log"]
    ws = wb["Reconciliation"]
    assert "Weekly" in ws["A1"].value
    assert ws["B6"].value == "N/A"
    assert ws["C6"].value == 2730
    assert ws["E6"].value == 2730
    assert ws["F6"].value == 0
    assert ws["B8"].value == "N/A"
    assert ws.auto_filter.ref == "A5:H8"


def test_weekly_reconciliation_with_optional_summary(tmp_path: Path):
    input_dir = tmp_path / "input" / "9354" / "2026-W22"
    summary_dir = input_dir / "summary"
    activity_dir = input_dir / "activity"
    summary_dir.mkdir(parents=True)
    activity_dir.mkdir(parents=True)
    create_summary(summary_dir / "05.31.2026 9354 Gift Card Summary.xlsx")
    for item in WEEKLY:
        create_activity(activity_dir / item[0], *item[1:])

    summary_path, activity_paths, _pos_path = discover_input_files(input_dir, mode="weekly")
    assert summary_path is not None
    summary = parse_summary(summary_path, "9354")
    activities = [parse_activity_file(path, summary.conversion_promo_codes) for path in activity_paths]
    pos = parse_pos_controls_from_values("9354", "2026-W22", "11642.00", "49869.75")
    result = build_reconciliation(store="9354", period="2026-W22", period_end=parse_date("2026-05-31"), summary=summary, activities=activities, pos_controls=pos, mode="weekly")

    assert result.mode == "weekly"
    assert result.lines[0].summary_value == Decimal("11507.00")
    assert result.lines[0].activity_variance == Decimal("0.00")
    assert result.lines[2].pos_variance == Decimal("132.73")


def test_monthly_mode_still_requires_summary(tmp_path: Path):
    input_dir = tmp_path / "input" / "9354" / "2026-05"
    activity_dir = input_dir / "activity"
    activity_dir.mkdir(parents=True)
    create_activity(activity_dir / WEEKLY[0][0], *WEEKLY[0][1:])

    with pytest.raises(ParseError, match="Expected exactly one Gift Card Summary"):
        discover_input_files(input_dir, mode="monthly")


def test_weekly_mode_fails_when_activity_files_are_missing(tmp_path: Path):
    input_dir = tmp_path / "input" / "9354" / "2026-W22"
    input_dir.mkdir(parents=True)
    (input_dir / "pos_controls.csv").write_text("store,period,pos_gift_card_issue,pos_gift_card_payment\n9354,2026-W22,2730.00,7446.47\n", encoding="utf-8")

    with pytest.raises(ParseError, match="No Gift Card Activity"):
        discover_input_files(input_dir, mode="weekly")


def test_weekly_mode_fails_when_pos_controls_are_missing(tmp_path: Path):
    input_dir = tmp_path / "input" / "9354" / "2026-W22"
    activity_dir = input_dir / "activity"
    activity_dir.mkdir(parents=True)
    create_activity(activity_dir / WEEKLY[0][0], *WEEKLY[0][1:])

    with pytest.raises(SystemExit, match="POS controls missing"):
        main([
            "--mode", "weekly",
            "--store", "9354",
            "--period", "2026-W22",
            "--input-dir", str(input_dir),
            "--output-file", str(tmp_path / "should-not-exist.xlsx"),
        ])


def test_pos_controls_reject_malformed_values(tmp_path: Path):
    pos_path = tmp_path / "pos_controls.csv"
    pos_path.write_text("store,period,pos_gift_card_issue,pos_gift_card_payment\n9354,2026-W22,not-money,7446.47\n", encoding="utf-8")

    with pytest.raises(ParseError, match="malformed value"):
        parse_pos_controls(pos_path, "9354", "2026-W22")


def test_click_runner_uses_auto_weekly_mode():
    script = (REPO_ROOT / "Run-Gift-Card-Reconciliation.cmd").read_text(encoding="utf-8")
    assert "gift_card_recon.auto_run" in script
    assert "Run-Weekly-Reconciliation.ps1" not in script
    assert "run_recon.ps1" not in script


def create_summary(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([None] * 16)
    ws.append([1, "Cards ACTIVATED PRE-Conversion", None, "A", "B", "C", "D = A - B - C ", "E = -(D) *10%"])
    ws.append([None, "Franchise Partner", "Store Number", "Redemption of Pre-Conversion Card", "Redemption of own activated card", "Redemption of card activated at another RC Location", "Redemptions subject to 10% Fee", "GCDR Total"])
    ws.append([None, "Sorensen", 9354, -7063.27, -2829.99, -147.94, -4085.34, 408.534])
    ws.append([])
    ws.append([2, "Cards ACTIVATED POST-Conversion", None, "A", "B", "C", "D = A - B - C ", "E = -(D) *10%"])
    ws.append([None, "Franchise Partner", "Store Number", "Redemption of Post-Conversion Card", "Redemption of own activated card", "Redemption of card activated at another RC Location", "Redemptions subject to 10% Fee", "GCDR"])
    ws.append([None, "Sorensen", 9354, -1161.13, -335, 0, -826.13, 82.613])
    ws.append([])
    ws.append([3, "In-Restaurant Activations and Cash Collections", None, "H", "I"])
    ws.append([None, "Franchise Partner", "Store Number", "Activations", "Activated in restaurant pre-conversion and Redeemed at another Ruth's Chris in current month"])
    ws.append([None, "Sorensen", 9354, 0, 358.21])
    ws.append([])
    ws.append([4, "Non-Conversion Promo Card Calculation", None, "H", "A", "D", "E = -(D) *10%", None, "RC Conversion Promo Codes"])
    ws.append([None, "Franchise Partner", "Store Number", "Total Activations", "Total Redemptions", "GCDR Redemptions", "GCDR", None, "RC Conversion Promo Codes"])
    ws.append([None, "Sorensen", 9354, 11507, -41643.08, -22008.26, 2200.826, None, 8106677, "PIN"])
    for code in [8106682, 8107472, 8107507, 8238836]:
        ws.append([None, None, None, None, None, None, None, None, code])
    ws.append([])
    ws.append([5, "SUMMARY", None, None, "H", "I", "A", "J", "E", "K= H + I + A + J + E "])
    ws.append([None, "Franchise Partner", "Store Number", "Gift Card Franchise Fee Rate", "Total Activations", "Activated in restaurant pre-conversion and Redeemed at another Ruth's Chris in current month", "Total Redemptions", "Payable Redemptions", "GCDR", "Net Settlement"])
    ws.append([None, "Sorensen", 9354, 0.1, 11507, 358.21, -49867.48, -47037.49, 2691.973, -32480.307])
    wb.save(path)


def parse_pos_controls_from_values(store: str, period: str, issue: str, payment: str):
    from gift_card_recon.parsers import pos_controls_from_args

    return pos_controls_from_args(store, period, issue, payment)


def create_activity(path: Path, begin: str, end: str, gross_activation: Decimal, void_activation: Decimal, conversion_redemption: Decimal, non_conversion_redemption_gross: Decimal, void_redemption: Decimal) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    ws.append([f"All GC Activity BY Rest Number and Date Range  Date Printed:  01-JUN-26  BEGIN DATE: '{begin}', END DATE: '{end}', Rest Number Parameter 1: '9354'"])
    ws.append(["Response:000"])
    ws.append(["Rest Number:9354"])
    ws.append(["Card No", "Request", "Request Code Listing", "Business Date", "Corp Code", "Transaction No", "Amount SUM", "Promocode", "Authorization Code"])
    ws.append(["0001xxxx", 100, "Activation", "2026-05-01", None, 1, float(gross_activation), 8056682, 111111])
    if void_activation != 0:
        ws.append(["0002xxxx", 102, "Void Of Activation", "2026-05-01", None, 2, float(void_activation), 8056682, 222222])
    ws.append(["0003xxxx", 202, "Redemption No Nsf", "2026-05-01", None, 3, float(conversion_redemption), 8106677, 333333])
    ws.append(["0004xxxx", 202, "Redemption No Nsf", "2026-05-01", None, 4, float(non_conversion_redemption_gross), 8056682, 444444])
    if void_redemption != 0:
        ws.append(["0005xxxx", 203, "Void Of Redemption", "2026-05-01", None, 5, float(void_redemption), 8056682, 555555])
    wb.save(path)
