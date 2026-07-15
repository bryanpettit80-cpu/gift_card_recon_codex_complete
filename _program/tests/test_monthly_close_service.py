from __future__ import annotations

import json
from dataclasses import replace
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from gift_card_recon.close_assessment import CloseStatus
from gift_card_recon.evidence_archive import ArchiveError
from gift_card_recon.fiscal_calendar import fiscal_period_for_label
from gift_card_recon.models import DardenCreditMemo
from gift_card_recon.monthly_close_service import (
    CloseBlockedError,
    canonical_output_paths,
    review_output_paths,
    run_monthly_close_service,
)
from gift_card_recon.pdf_export import PdfExportError
from gift_card_recon.parsers import ParseError
from gift_card_recon.store_config import get_store_config


@pytest.mark.parametrize(
    ("store", "payment_variances", "expected_status", "expected_period_net"),
    [
        (
            "9354",
            [Decimal("-0.01"), Decimal("0.00"), Decimal("2.44"), Decimal("0.00"), Decimal("0.00")],
            CloseStatus.CLOSED_WITH_REVIEW,
            Decimal("-2.43"),
        ),
        (
            "9355",
            [Decimal("0.00")] * 5,
            CloseStatus.CLOSED,
            Decimal("0.00"),
        ),
    ],
)
def test_realistic_five_week_close_status_and_report_flow(
    tmp_path: Path,
    store: str,
    payment_variances: list[Decimal],
    expected_status: CloseStatus,
    expected_period_net: Decimal,
) -> None:
    setup = _build_period(tmp_path, store=store, payment_variances=payment_variances)

    run = run_monthly_close_service(
        store=store,
        period=setup["period"].period_key,
        input_dir=setup["input_dir"],
        micros_path=setup["micros_dir"],
        micros_work_dir=tmp_path / "extract",
        archive_root=setup["archive_root"],
        output_root=setup["output_root"],
        darden_report=setup["darden"],
        fiscal_period=setup["period"],
        cleanup_sources=False,
        allow_unconfigured_micros=True,
        generated_at=datetime(2026, 7, 6, 9, 30),
        pdf_exporter=_fake_pdf_exporter,
    )

    assert run.assessment.status is expected_status
    assert sum((row.net_variance for row in run.weekly_variances), Decimal("0.00")) == expected_period_net
    assert run.workbook_path.exists()
    assert run.pdf_path.exists()
    expected_xlsx, expected_pdf = canonical_output_paths(
        setup["output_root"],
        config=get_store_config(store),
        fiscal_period=setup["period"],
    )
    assert (run.workbook_path, run.pdf_path) == (expected_xlsx, expected_pdf)

    workbook = load_workbook(run.workbook_path, data_only=False)
    report = workbook["Monthly Close Report"]
    assert report["A1"].value == get_store_config(store).report_heading
    assert report["A4"].value == expected_status.value
    assert report["G9"].value == "MATCHED"
    assert len(report.row_breaks.brk) == 1
    assert report.page_setup.fitToHeight == 0
    assert report.page_setup.scale == 85
    reconciliation_sheet = workbook["Reconciliation"]
    assert reconciliation_sheet["A3"].value.startswith(
        f"Authoritative monthly close status: {expected_status.value}."
    )
    assert "does not determine overall close status" in reconciliation_sheet["A3"].value
    assert any(
        reconciliation_sheet.cell(row, 1).value == "Darden Settlement Control"
        for row in range(1, reconciliation_sheet.max_row + 1)
    )
    source_sheet = workbook["Source Files"]
    source_paths = [
        source_sheet.cell(row, 6).value for row in range(4, source_sheet.max_row + 1)
    ]
    normalized_source_paths = [str(value).replace("\\", "/") for value in source_paths]
    assert all(value.startswith(f"Monthly Close/{store}") for value in normalized_source_paths)
    assert any(str(value).endswith("DLYSYSTT.TXT") for value in source_paths)
    assert any(str(value).endswith("TENDER_DETAIL.TXT") for value in source_paths)
    exception_sheet = workbook["Exception Log"]
    if expected_status is CloseStatus.CLOSED_WITH_REVIEW:
        assert exception_sheet["A4"].value == "REVIEW"
        assert "No parsing or validation exceptions" not in str(exception_sheet["B4"].value)

    manifest = json.loads(run.manifest_path.read_text(encoding="utf-8"))
    assert manifest["status"] == expected_status.value
    assert len(manifest["sources"]) == 9
    assert {item["role"] for item in manifest["artifacts"]} == {"workbook", "pdf"}
    assert all(len(item["sha256"]) == 64 for item in manifest["sources"])


def test_review_output_paths_use_monthly_specific_operator_folder(tmp_path: Path) -> None:
    output_root = tmp_path / "03 Finished Reports"
    workbook, pdf = review_output_paths(
        output_root,
        config=get_store_config("9355"),
        fiscal_period=fiscal_period_for_label("FY27-M01"),
    )

    expected_folder = output_root / "Monthly Close - Review Required"
    assert workbook == expected_folder / "Virginia_Beach_9355_FY27-M01_Review_Required.xlsx"
    assert pdf == expected_folder / "Virginia_Beach_9355_FY27-M01_Review_Required.pdf"
    assert workbook.parent != output_root / "Review Required"


def test_larger_variance_creates_only_review_required_artifacts(tmp_path: Path) -> None:
    setup = _build_period(
        tmp_path,
        store="9355",
        issue_variances=[Decimal("7.00"), Decimal("0.00"), Decimal("0.00"), Decimal("0.00"), Decimal("0.00")],
    )

    with pytest.raises(CloseBlockedError) as exc_info:
        _run(setup, cleanup_sources=True)

    error = exc_info.value
    assert error.assessment.status is CloseStatus.REVIEW_REQUIRED
    assert error.review_workbook is not None and error.review_workbook.exists()
    assert error.review_pdf is not None and error.review_pdf.exists()
    expected_review_folder = setup["output_root"] / "Monthly Close - Review Required"
    assert error.review_workbook.parent == expected_review_folder
    assert error.review_pdf.parent == expected_review_folder
    assert not (setup["output_root"] / "Review Required").exists()
    report = load_workbook(error.review_workbook)["Monthly Close Report"]
    assert report["A4"].value == "REVIEW REQUIRED"
    canonical, canonical_pdf = canonical_output_paths(
        setup["output_root"],
        config=get_store_config("9355"),
        fiscal_period=setup["period"],
    )
    assert not canonical.exists()
    assert not canonical_pdf.exists()
    assert not (setup["archive_root"] / "Monthly Close").exists()
    assert setup["summary_path"].exists()
    assert setup["darden"].source_file.exists()


def test_missing_tender_evidence_is_blocking_and_never_archived(tmp_path: Path) -> None:
    setup = _build_period(tmp_path, store="9355")
    (setup["micros_dir"] / "TENDER_DETAIL.TXT").unlink()

    with pytest.raises(CloseBlockedError, match="TENDER_DETAIL.TXT") as exc_info:
        _run(setup)

    assert exc_info.value.review_workbook is not None
    assert not (setup["archive_root"] / "Monthly Close").exists()
    assert setup["summary_path"].exists()


def test_pdf_export_failure_blocks_publication_and_archive(tmp_path: Path) -> None:
    setup = _build_period(tmp_path, store="9355")
    review_xlsx, review_pdf = review_output_paths(
        setup["output_root"],
        config=get_store_config("9355"),
        fiscal_period=setup["period"],
    )
    review_xlsx.parent.mkdir(parents=True)
    review_xlsx.write_bytes(b"older diagnostic workbook")
    review_pdf.write_bytes(b"older diagnostic PDF")

    def fail_export(**_kwargs):
        raise PdfExportError("simulated Excel PDF export failure")

    with pytest.raises(CloseBlockedError, match="simulated Excel PDF export failure") as exc_info:
        _run(setup, pdf_exporter=fail_export)

    canonical, canonical_pdf = canonical_output_paths(
        setup["output_root"],
        config=get_store_config("9355"),
        fiscal_period=setup["period"],
    )
    assert not canonical.exists() and not canonical_pdf.exists()
    assert not (setup["archive_root"] / "Monthly Close").exists()
    assert exc_info.value.review_workbook is not None
    assert exc_info.value.review_pdf is None
    assert exc_info.value.review_pdf_error == "simulated Excel PDF export failure"
    assert exc_info.value.review_workbook == review_xlsx
    assert review_xlsx.read_bytes() != b"older diagnostic workbook"
    assert not review_pdf.exists()
    assert setup["summary_path"].exists()


def test_locked_stale_diagnostic_preserves_old_pair_without_masking_blocker(
    tmp_path: Path,
    monkeypatch,
) -> None:
    setup = _build_period(
        tmp_path,
        store="9355",
        issue_variances=[Decimal("7.00"), Decimal("0.00"), Decimal("0.00"), Decimal("0.00"), Decimal("0.00")],
    )
    review_xlsx, review_pdf = review_output_paths(
        setup["output_root"],
        config=get_store_config("9355"),
        fiscal_period=setup["period"],
    )
    review_xlsx.parent.mkdir(parents=True)
    review_xlsx.write_bytes(b"older diagnostic workbook")
    review_pdf.write_bytes(b"older diagnostic PDF")

    from gift_card_recon import monthly_close_service as service

    original = service._assert_publishable

    def locked_pdf(path: Path) -> None:
        if Path(path) == review_pdf:
            raise PermissionError("simulated locked diagnostic PDF")
        original(path)

    def fail_export(**_kwargs):
        raise PdfExportError("simulated diagnostic PDF failure")

    monkeypatch.setattr(service, "_assert_publishable", locked_pdf)
    with pytest.raises(CloseBlockedError) as exc_info:
        _run(setup, pdf_exporter=fail_export)

    error = exc_info.value
    assert "7.00" in str(error)
    assert error.review_workbook is None
    assert error.review_pdf is None
    assert error.review_publication_error == "simulated locked diagnostic PDF"
    assert review_xlsx.read_bytes() == b"older diagnostic workbook"
    assert review_pdf.read_bytes() == b"older diagnostic PDF"


def test_archive_failure_blocks_publication_and_preserves_sources(
    tmp_path: Path,
    monkeypatch,
) -> None:
    setup = _build_period(tmp_path, store="9355")

    def fail_archive(_records):
        raise ArchiveError("simulated archive failure")

    monkeypatch.setattr(
        "gift_card_recon.monthly_close_service.execute_archive_plan",
        fail_archive,
    )
    with pytest.raises(CloseBlockedError, match="simulated archive failure"):
        _run(setup, cleanup_sources=True)

    canonical, canonical_pdf = canonical_output_paths(
        setup["output_root"],
        config=get_store_config("9355"),
        fiscal_period=setup["period"],
    )
    assert not canonical.exists() and not canonical_pdf.exists()
    assert setup["summary_path"].exists()
    assert setup["darden"].source_file.exists()


def test_locked_canonical_output_fails_without_alternate_or_archiving(
    tmp_path: Path,
    monkeypatch,
) -> None:
    setup = _build_period(tmp_path, store="9355")
    canonical, _ = canonical_output_paths(
        setup["output_root"],
        config=get_store_config("9355"),
        fiscal_period=setup["period"],
    )
    canonical.parent.mkdir(parents=True)
    canonical.write_bytes(b"older canonical workbook")

    from gift_card_recon import monthly_close_service as service

    original = service._assert_publishable

    def locked_only(path: Path) -> None:
        if Path(path) == canonical:
            raise PermissionError("simulated locked canonical output")
        original(path)

    monkeypatch.setattr(service, "_assert_publishable", locked_only)
    with pytest.raises(CloseBlockedError, match="simulated locked canonical output"):
        _run(setup, cleanup_sources=True)

    assert canonical.read_bytes() == b"older canonical workbook"
    assert not list(canonical.parent.glob("*with_weekly_variance*"))
    assert not (setup["archive_root"] / "Monthly Close").exists()
    assert setup["summary_path"].exists()


def test_successful_close_removes_only_live_inputs_after_verified_archive(tmp_path: Path) -> None:
    setup = _build_period(tmp_path, store="9355")

    run = _run(setup, cleanup_sources=True)

    assert run.workbook_path.exists() and run.pdf_path.exists()
    assert not setup["summary_path"].exists()
    assert not setup["darden"].source_file.exists()
    assert not list((setup["input_dir"] / "activity").glob("*.xlsx"))
    assert (setup["micros_dir"] / "DLYSYSTT.TXT").exists()
    assert (setup["micros_dir"] / "TENDER_DETAIL.TXT").exists()
    archive_base = (
        setup["archive_root"]
        / "Monthly Close"
        / "9355"
        / setup["period"].folder_name
    )
    assert len(list((archive_base / "activity").glob("*.xlsx"))) == 5
    assert (archive_base / "summary" / setup["summary_path"].name).exists()
    assert (archive_base / "darden" / setup["darden"].source_file.name).exists()


def test_summary_period_mismatch_is_blocking(tmp_path: Path) -> None:
    setup = _build_period(tmp_path, store="9355")
    wrong = setup["summary_path"].with_name("06.28.2026 9355 Gift Card Summary.xlsx")
    setup["summary_path"].replace(wrong)

    with pytest.raises(CloseBlockedError, match="report end 2026-06-28"):
        _run(setup)


@pytest.mark.parametrize(
    ("wrong_store", "period_start", "period_end", "message"),
    [
        ("9354", date(2026, 6, 1), date(2026, 7, 5), "expected store 9355"),
        ("9355", date(2026, 6, 2), date(2026, 7, 5), "does not match FY27-M01"),
    ],
)
def test_wrong_darden_identity_is_blocking(
    tmp_path: Path,
    wrong_store: str,
    period_start: date,
    period_end: date,
    message: str,
) -> None:
    setup = _build_period(tmp_path, store="9355")
    setup["darden"] = DardenCreditMemo(
        source_file=setup["darden"].source_file,
        store=wrong_store,
        period_start=period_start,
        period_end=period_end,
        total=Decimal("-200.00"),
    )

    with pytest.raises(CloseBlockedError, match=message):
        _run(setup)


def test_invalid_workbook_extension_is_rejected(tmp_path: Path) -> None:
    setup = _build_period(tmp_path, store="9355")

    with pytest.raises(ParseError, match="must end in .xlsx"):
        _run(setup, output_path=tmp_path / "Output" / "wrong.pdf")


def test_supplied_fiscal_period_object_must_match_canonical_calendar(tmp_path: Path) -> None:
    setup = _build_period(tmp_path, store="9355")
    altered = replace(setup["period"], start_date=date(2026, 6, 2))

    with pytest.raises(ParseError, match="does not match the supplied period object"):
        _run(setup, fiscal_period=altered)


def test_manifest_failure_rolls_back_new_archive_copies_and_canonical_outputs(
    tmp_path: Path,
    monkeypatch,
) -> None:
    setup = _build_period(tmp_path, store="9355")

    def fail_manifest(*_args, **_kwargs):
        raise ArchiveError("simulated manifest failure")

    monkeypatch.setattr(
        "gift_card_recon.monthly_close_service.write_close_manifest_atomic",
        fail_manifest,
    )
    with pytest.raises(CloseBlockedError, match="simulated manifest failure"):
        _run(setup, cleanup_sources=True)

    canonical, canonical_pdf = canonical_output_paths(
        setup["output_root"],
        config=get_store_config("9355"),
        fiscal_period=setup["period"],
    )
    assert not canonical.exists() and not canonical_pdf.exists()
    assert not (setup["archive_root"] / "Monthly Close").exists()
    assert setup["summary_path"].exists()
    assert setup["darden"].source_file.exists()


def test_evidence_change_during_calculation_blocks_archive_and_publication(
    tmp_path: Path,
    monkeypatch,
) -> None:
    setup = _build_period(tmp_path, store="9355")
    from gift_card_recon import monthly_close_service as service

    original_plan = service.plan_evidence_archive

    def mutate_then_plan(items, *, archive_root):
        setup["summary_path"].write_bytes(b"changed after parsing")
        return original_plan(items, archive_root=archive_root)

    monkeypatch.setattr(service, "plan_evidence_archive", mutate_then_plan)
    with pytest.raises(CloseBlockedError, match="Evidence changed while the close was being calculated"):
        _run(setup)

    assert not (setup["archive_root"] / "Monthly Close").exists()


def test_success_archives_and_removes_superseded_review_artifact(tmp_path: Path) -> None:
    setup = _build_period(tmp_path, store="9355")
    review = (
        setup["output_root"]
        / "Monthly Close - Review Required"
        / "Virginia_Beach_9355_FY27-M01_Review_Required.xlsx"
    )
    review.parent.mkdir(parents=True)
    review.write_bytes(b"older diagnostic")

    _run(setup)

    assert not review.exists()
    archived = (
        setup["archive_root"]
        / "Generated Reports"
        / "Diagnostics"
        / "9355"
        / "FY27-M01"
        / review.name
    )
    assert archived.exists()


def _run(setup: dict[str, object], **overrides):
    arguments = {
        "store": setup["store"],
        "period": setup["period"].period_key,
        "input_dir": setup["input_dir"],
        "micros_path": setup["micros_dir"],
        "micros_work_dir": setup["root"] / "extract",
        "archive_root": setup["archive_root"],
        "output_root": setup["output_root"],
        "darden_report": setup["darden"],
        "fiscal_period": setup["period"],
        "cleanup_sources": False,
        "allow_unconfigured_micros": True,
        "generated_at": datetime(2026, 7, 6, 9, 30),
        "pdf_exporter": _fake_pdf_exporter,
    }
    arguments.update(overrides)
    return run_monthly_close_service(**arguments)


def _build_period(
    root: Path,
    *,
    store: str,
    issue_variances: list[Decimal] | None = None,
    payment_variances: list[Decimal] | None = None,
) -> dict[str, object]:
    fiscal_period = fiscal_period_for_label("FY27-M01")
    input_dir = root / "Monthly Close" / store / fiscal_period.folder_name
    summary_path = input_dir / "summary" / f"07.05.2026 {store} Gift Card Summary.xlsx"
    activity_dir = input_dir / "activity"
    micros_dir = root / f"micros-{store}"
    summary_path.parent.mkdir(parents=True)
    activity_dir.mkdir()
    micros_dir.mkdir()
    _write_summary(summary_path, store=store)

    issue_variances = issue_variances or [Decimal("0.00")] * 5
    payment_variances = payment_variances or [Decimal("0.00")] * 5
    system_rows: list[tuple[date, Decimal, Decimal]] = []
    for index, week_end in enumerate(fiscal_period.expected_week_endings):
        week_start = week_end - timedelta(days=6)
        activation = Decimal("100.00") if index == 0 else Decimal("0.00")
        redemption = Decimal("-300.00") if index == 0 else Decimal("0.00")
        _write_activity(
            activity_dir / f"{week_end:%m.%d.%Y} {store} Gift Card Activity.xlsx",
            store=store,
            begin=week_start,
            end=week_end,
            transaction_date=week_start,
            activation=activation,
            redemption=redemption,
        )
        for business_date in _dates(week_start, week_end):
            system_rows.append((business_date, Decimal("0.00"), Decimal("0.00")))
        issue_total = activation + issue_variances[index]
        payment_total = abs(redemption) + payment_variances[index]
        first_row_index = len(system_rows) - 7
        system_rows[first_row_index] = (week_start, issue_total, payment_total)
    _write_micros(micros_dir, system_rows)

    darden_path = input_dir / "darden" / f"FY27-M01 {store} Darden Credit Memo.pdf"
    darden_path.parent.mkdir()
    darden_path.write_bytes(b"synthetic Darden evidence")
    darden = DardenCreditMemo(
        source_file=darden_path,
        store=store,
        period_start=fiscal_period.start_date,
        period_end=fiscal_period.end_date,
        total=Decimal("-200.00"),
        invoice_number="FY27 M01 synthetic",
        invoice_date=date(2026, 7, 6),
    )
    return {
        "root": root,
        "store": store,
        "period": fiscal_period,
        "input_dir": input_dir,
        "summary_path": summary_path,
        "micros_dir": micros_dir,
        "darden": darden,
        "archive_root": root / "Archive - Old Files",
        "output_root": root / "Output",
    }


def _write_summary(path: Path, *, store: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["SUMMARY"])
    sheet.append(
        [
            "Franchise Partner",
            "Store Number",
            "Gift Card Franchise Fee Rate",
            "Total Activations",
            "Total Redemptions",
            "Payable Redemptions",
            "GCDR",
            "Net Settlement",
        ]
    )
    sheet.append(["Sorensen", int(store), 0.1, 100, -300, -300, 0, -200])
    workbook.save(path)


def _write_activity(
    path: Path,
    *,
    store: str,
    begin: date,
    end: date,
    transaction_date: date,
    activation: Decimal,
    redemption: Decimal,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet 1"
    sheet.append(
        [
            f"All GC Activity BEGIN DATE: '{begin:%d-%b-%Y}', END DATE: "
            f"'{end:%d-%b-%Y}', Rest Number Parameter 1: '{store}'"
        ]
    )
    sheet.append(
        [
            "Card No",
            "Request",
            "Request Code Listing",
            "Business Date",
            "Transaction No",
            "Amount SUM",
            "Promocode",
            "Authorization Code",
        ]
    )
    sheet.append(["0001", 100, "Activation", transaction_date, 1, float(activation), None, 111])
    sheet.append(["0002", 202, "Redemption No Nsf", transaction_date, 2, float(redemption), None, 222])
    workbook.save(path)


def _write_micros(folder: Path, rows: list[tuple[date, Decimal, Decimal]]) -> None:
    config = get_store_config("9355")
    system_lines: list[str] = []
    tender_lines: list[str] = []
    for business_date, issue, payment in rows:
        row = ["0"] * 132
        row[0] = f"{business_date:%Y-%m-%d} 00:00:00.000"
        row[config.micros_issue_column_index] = f"{issue:.2f}"
        row[config.micros_payment_column_index] = f"{payment:.2f}"
        system_lines.append(",".join(row))
        tender_lines.append(
            f"'{business_date:%Y-%m-%d}',{payment:.2f},350,'G C Payment','T'"
        )
    (folder / "DLYSYSTT.TXT").write_text("\n".join(system_lines), encoding="utf-8")
    (folder / "TENDER_DETAIL.TXT").write_text("\n".join(tender_lines), encoding="utf-8")


def _dates(start: date, end: date):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def _fake_pdf_exporter(*, workbook_path: Path, pdf_path: Path, expected_location_label: str):
    assert Path(workbook_path).is_file()
    assert expected_location_label in {
        "RICHMOND - STORE 9354",
        "VIRGINIA BEACH - STORE 9355",
    }
    Path(pdf_path).write_bytes(b"%PDF-1.4 mocked two-page Excel export")
    return Path(pdf_path)
