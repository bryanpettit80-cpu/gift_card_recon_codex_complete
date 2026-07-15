from __future__ import annotations

import csv
import errno
import json
import os
import shutil
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from gift_card_recon import weekly_service
from gift_card_recon.auto_run import iso_week_period, run_weekly_reconciliations
from gift_card_recon.store_config import get_store_config
from gift_card_recon.utils import parse_date, sha256_file


def test_iso_week_period_uses_report_end_date():
    assert iso_week_period(parse_date("2026-06-07")) == "2026-W23"


def test_auto_weekly_runner_derives_controls_and_publishes_verified_package(tmp_path: Path):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9355 Virginia Beach" / "activity" / "06.07.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 1),
        end=date(2026, 6, 7),
        issue=Decimal("150.00"),
        payment=Decimal("6657.73"),
    )
    create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 1),
        issue=Decimal("150.00"),
        payment=Decimal("6657.73"),
    )

    reports = run(paths)

    assert [(report.store, report.status, report.close_status) for report in reports] == [
        ("9355", "created", "PASS")
    ], reports[0].message
    output_path = (
        paths["output"]
        / "9355 Virginia Beach"
        / "2026"
        / "Gift_Card_Reconciliation_9355_2026-W23.xlsx"
    )
    package = paths["archive"] / "9355 Virginia Beach" / "2026" / "2026-W23"
    monthly_path = (
        paths["monthly"]
        / "9355 Virginia Beach"
        / "FY27 M01 - Fiscal June"
        / "activity"
        / activity_path.name
    )
    assert output_path.is_file()
    assert monthly_path.is_file()
    assert not activity_path.exists()
    assert sha256_file(output_path) == sha256_file(package / "report" / output_path.name)
    assert sha256_file(monthly_path) == sha256_file(package / "activity" / activity_path.name)

    evidence_rows = list(csv.DictReader((package / "pos" / "weekly_pos_tender_evidence.csv").open(encoding="utf-8")))
    assert len(evidence_rows) == 7
    assert sum(Decimal(row["pos_gift_card_issue"]) for row in evidence_rows) == Decimal("150.00")
    assert sum(Decimal(row["pos_gift_card_payment"]) for row in evidence_rows) == Decimal("6657.73")

    manifest = json.loads((package / "weekly_manifest.json").read_text(encoding="utf-8"))
    assert manifest["schema_version"] == 1
    assert manifest["status"] == "PASS"
    assert len(manifest["daily_pos_tender"]) == 7
    assert manifest["artifacts"]["canonical_workbook"]["sha256"] == sha256_file(output_path)
    assert manifest["artifacts"]["weekly_pos_tender_evidence"]["sha256"] == sha256_file(
        package / "pos" / "weekly_pos_tender_evidence.csv"
    )

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["Reconciliation"]
    assert sheet["C6"].value == 150
    assert sheet["C7"].value == 6657.73


def test_richmond_accepts_only_zero_scheduled_monday_as_complete(tmp_path: Path):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9354 Richmond" / "activity" / "07.05.2026 9354 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9354",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("275.00"),
        payment=Decimal("980.00"),
    )
    create_micros_week(
        paths["operations"],
        store="9354",
        week_start=date(2026, 6, 29),
        issue=Decimal("275.00"),
        payment=Decimal("980.00"),
        omit_dates={date(2026, 6, 29)},
    )

    reports = run(paths)

    assert reports[0].status == "created"
    assert reports[0].close_status == "PASS"
    package = paths["archive"] / "9354 Richmond" / "2026" / "2026-W27"
    manifest = json.loads((package / "weekly_manifest.json").read_text(encoding="utf-8"))
    monday = manifest["daily_pos_tender"][0]
    assert monday["business_date"] == "2026-06-29"
    assert monday["accepted_scheduled_closure"] is True
    assert monday["pos_gift_card_issue"] == "0.00"


def test_nonzero_pos_or_tender_variance_publishes_review(tmp_path: Path):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9355 Virginia Beach" / "activity" / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("275.00"),
        payment=Decimal("980.00"),
    )
    create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("276.00"),
        payment=Decimal("980.00"),
        tender_payment=Decimal("979.00"),
    )

    reports = run(paths)

    assert reports[0].status == "created"
    assert reports[0].close_status == "REVIEW"
    package = paths["archive"] / "9355 Virginia Beach" / "2026" / "2026-W27"
    manifest = json.loads((package / "weekly_manifest.json").read_text(encoding="utf-8"))
    assert manifest["status"] == "REVIEW"
    assert manifest["totals"]["tender_variance"] == "1.00"
    assert len(manifest["review_items"]) == 2
    assert any("POS variance" in item for item in manifest["review_items"])
    assert any("tender-detail variance" in item for item in manifest["review_items"])


def test_store_failures_are_independent_and_empty_inbox_is_no_op(tmp_path: Path):
    paths = make_layout(tmp_path, stores=("9354 Richmond", "9355 Virginia Beach"))
    wrong_store_path = paths["input"] / "9354 Richmond" / "activity" / "07.05.2026 wrong Gift Card Activity.xlsx"
    create_activity(
        wrong_store_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("1.00"),
        payment=Decimal("2.00"),
    )
    create_micros_week(
        paths["operations"],
        store="9354",
        week_start=date(2026, 6, 29),
        issue=Decimal("1.00"),
        payment=Decimal("2.00"),
    )

    reports = run(paths)

    assert [(report.store, report.status) for report in reports] == [
        ("9354", "skipped"),
        ("9355", "no-op"),
    ]
    assert "identifies store 9355; expected 9354" in reports[0].message
    assert wrong_store_path.exists()
    assert not paths["output"].exists()
    assert not paths["archive"].exists()


def test_missing_micros_coverage_blocks_without_any_publication(tmp_path: Path):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9355 Virginia Beach" / "activity" / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
        omit_dates={date(2026, 7, 1)},
    )

    reports = run(paths)

    assert reports[0].status == "skipped"
    assert "2026-07-01 is missing" in reports[0].message
    assert activity_path.exists()
    assert not paths["output"].exists()
    assert not paths["archive"].exists()
    assert not paths["monthly"].exists()


def test_exact_duplicate_uses_durable_archive_without_monthly_staging_or_live_micros(tmp_path: Path):
    paths = make_layout(tmp_path)
    activity_dir = paths["input"] / "9355 Virginia Beach" / "activity"
    activity_path = activity_dir / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    micros_dir = create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    assert run(paths)[0].status == "created"

    archived_activity = (
        paths["archive"]
        / "9355 Virginia Beach"
        / "2026"
        / "2026-W27"
        / "activity"
        / activity_path.name
    )
    shutil.rmtree(paths["monthly"])
    shutil.rmtree(micros_dir)
    shutil.copy2(archived_activity, activity_path)
    duplicate_report = run(paths)[0]
    assert duplicate_report.status == "duplicate"
    quarantine = paths["review"] / "duplicate-inputs" / "9355" / "2026-W27"
    assert len(list(quarantine.glob("*Gift Card Activity.xlsx"))) == 1
    assert not activity_path.exists()

    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("11.00"),
        payment=Decimal("20.00"),
    )
    conflict_report = run(paths)[0]
    assert conflict_report.status == "skipped"
    assert "different Activity report already exists" in conflict_report.message
    assert activity_path.exists()


@pytest.mark.parametrize(
    "artifact_key",
    ("archived_activity", "weekly_pos_tender_evidence", "archived_workbook", "canonical_workbook"),
)
def test_exact_duplicate_rejects_any_tampered_durable_or_canonical_artifact(
    tmp_path: Path,
    artifact_key: str,
):
    paths = make_layout(tmp_path)
    activity_path = (
        paths["input"]
        / "9355 Virginia Beach"
        / "activity"
        / "07.05.2026 9355 Gift Card Activity.xlsx"
    )
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    micros_dir = create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    assert run(paths)[0].status == "created"

    package = paths["archive"] / "9355 Virginia Beach" / "2026" / "2026-W27"
    manifest = json.loads((package / "weekly_manifest.json").read_text(encoding="utf-8"))
    archived_activity = package / manifest["artifacts"]["archived_activity"]["relative_path"]
    shutil.copy2(archived_activity, activity_path)
    record = manifest["artifacts"][artifact_key]
    target = Path(record["path"]) if artifact_key == "canonical_workbook" else package / record["relative_path"]
    with target.open("ab") as stream:
        stream.write(b"tampered")
    shutil.rmtree(paths["monthly"])
    shutil.rmtree(micros_dir)

    report = run(paths)[0]

    assert report.status == "skipped"
    assert "Existing weekly evidence package is incomplete or invalid" in report.message
    assert activity_path.exists()


def test_manifest_source_hash_cannot_make_a_different_activity_an_exact_duplicate(tmp_path: Path):
    paths = make_layout(tmp_path)
    activity_path = (
        paths["input"]
        / "9355 Virginia Beach"
        / "activity"
        / "07.05.2026 9355 Gift Card Activity.xlsx"
    )
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    assert run(paths)[0].status == "created"

    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("11.00"),
        payment=Decimal("20.00"),
    )
    package = paths["archive"] / "9355 Virginia Beach" / "2026" / "2026-W27"
    manifest_path = package / "weekly_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["source_files"]["activity"]["size_bytes"] = activity_path.stat().st_size
    manifest["source_files"]["activity"]["sha256"] = sha256_file(activity_path)
    manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    report = run(paths)[0]

    assert report.status == "skipped"
    assert "manifest Activity source does not match" in report.message
    assert activity_path.exists()


@pytest.mark.parametrize("redirect_kind", ("absolute", "traversal"))
def test_manifest_cannot_redirect_archived_activity_outside_package(
    tmp_path: Path,
    redirect_kind: str,
):
    paths = make_layout(tmp_path)
    activity_path = (
        paths["input"]
        / "9355 Virginia Beach"
        / "activity"
        / "07.05.2026 9355 Gift Card Activity.xlsx"
    )
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    micros_dir = create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    assert run(paths)[0].status == "created"

    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("11.00"),
        payment=Decimal("20.00"),
    )
    package = paths["archive"] / "9355 Virginia Beach" / "2026" / "2026-W27"
    manifest_path = package / "weekly_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    incoming_size = activity_path.stat().st_size
    incoming_hash = sha256_file(activity_path)
    for key in ("source_files", "artifacts"):
        record = (
            manifest[key]["activity"]
            if key == "source_files"
            else manifest[key]["archived_activity"]
        )
        record["size_bytes"] = incoming_size
        record["sha256"] = incoming_hash
    manifest["artifacts"]["archived_activity"]["relative_path"] = (
        str(activity_path.resolve())
        if redirect_kind == "absolute"
        else os.path.relpath(activity_path, package)
    )
    manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    shutil.rmtree(micros_dir)

    report = run(paths)[0]

    assert report.status == "skipped"
    assert "relative_path is outside its expected package folder" in report.message
    assert activity_path.exists()


def test_manifest_cannot_redirect_canonical_workbook_to_archived_copy(tmp_path: Path):
    paths = make_layout(tmp_path)
    activity_path = (
        paths["input"]
        / "9355 Virginia Beach"
        / "activity"
        / "07.05.2026 9355 Gift Card Activity.xlsx"
    )
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    micros_dir = create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    assert run(paths)[0].status == "created"

    package = paths["archive"] / "9355 Virginia Beach" / "2026" / "2026-W27"
    manifest_path = package / "weekly_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    archived_activity = package / manifest["artifacts"]["archived_activity"]["relative_path"]
    archived_workbook = package / manifest["artifacts"]["archived_workbook"]["relative_path"]
    shutil.copy2(archived_activity, activity_path)
    manifest["artifacts"]["canonical_workbook"]["path"] = str(archived_workbook.resolve())
    manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    shutil.rmtree(micros_dir)

    report = run(paths)[0]

    assert report.status == "skipped"
    assert "canonical_workbook path does not match" in report.message
    assert activity_path.exists()


def test_existing_canonical_workbook_blocks_and_preserves_input(tmp_path: Path):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9355 Virginia Beach" / "activity" / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    output_path = (
        paths["output"]
        / "9355 Virginia Beach"
        / "2026"
        / "Gift_Card_Reconciliation_9355_2026-W27.xlsx"
    )
    output_path.parent.mkdir(parents=True)
    output_path.write_bytes(b"locked canonical placeholder")

    report = run(paths)[0]

    assert report.status == "skipped"
    assert "canonical weekly report already exists" in report.message
    assert activity_path.exists()
    assert output_path.read_bytes() == b"locked canonical placeholder"
    assert not paths["archive"].exists()
    assert not any(path.is_file() for path in paths["monthly"].rglob("*"))


def test_transient_dropbox_package_lock_is_retried_without_residue(tmp_path: Path, monkeypatch):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9355 Virginia Beach" / "activity" / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    package_path = paths["archive"] / "9355 Virginia Beach" / "2026" / "2026-W27"
    output_path = (
        paths["output"]
        / "9355 Virginia Beach"
        / "2026"
        / "Gift_Card_Reconciliation_9355_2026-W27.xlsx"
    )
    monthly_path = (
        paths["monthly"]
        / "9355 Virginia Beach"
        / "FY27 M01 - Fiscal June"
        / "activity"
        / activity_path.name
    )
    real_replace = weekly_service.os.replace
    real_mkdtemp = weekly_service.tempfile.mkdtemp
    package_attempts = 0
    retry_delays: list[float] = []
    stage_roots: list[Path] = []
    runtime_temp = tmp_path / "runtime-temp"
    runtime_temp.mkdir()

    def lock_package_once(source, destination):
        nonlocal package_attempts
        if Path(destination) == package_path:
            package_attempts += 1
            assert Path(source).name == "package"
            if package_attempts == 1:
                exc = PermissionError(
                    errno.EACCES,
                    "simulated transient Dropbox lock",
                    str(destination),
                )
                exc.winerror = 32
                raise exc
        return real_replace(source, destination)

    def capture_stage_root(*args, **kwargs):
        if kwargs.get("dir") is None:
            kwargs["dir"] = runtime_temp
        path = Path(real_mkdtemp(*args, **kwargs))
        stage_roots.append(path)
        return str(path)

    monkeypatch.setattr(weekly_service.os, "replace", lock_package_once)
    monkeypatch.setattr(weekly_service.tempfile, "mkdtemp", capture_stage_root)
    monkeypatch.setattr(weekly_service.time, "sleep", retry_delays.append)

    report = run(paths)[0]

    assert report.status == "created", report.message
    assert report.close_status == "PASS"
    assert package_attempts == 2
    assert retry_delays == [weekly_service._TRANSIENT_FILE_RETRY_DELAYS[0]]
    assert len(stage_roots) == 1
    assert paths["operations"] not in stage_roots[0].parents
    assert not stage_roots[0].exists()
    assert package_path.is_dir()
    assert output_path.is_file()
    assert monthly_path.is_file()
    assert not activity_path.exists()
    assert sha256_file(output_path) == sha256_file(package_path / "report" / output_path.name)
    assert sha256_file(monthly_path) == sha256_file(package_path / "activity" / activity_path.name)
    assert not list(paths["archive"].glob(".weekly-staging-*"))
    assert not list(paths["output"].rglob(".gc-*.tmp"))
    assert not list(paths["monthly"].rglob(".gc-*.tmp"))


def test_nonempty_destination_error_is_not_retried(monkeypatch):
    attempts = 0
    retry_delays: list[float] = []

    def fail_with_nonempty_destination():
        nonlocal attempts
        attempts += 1
        exc = OSError(errno.EACCES, "simulated nonempty destination")
        exc.winerror = 145
        raise exc

    monkeypatch.setattr(weekly_service.time, "sleep", retry_delays.append)

    with pytest.raises(OSError, match="simulated nonempty destination"):
        weekly_service._retry_transient_file_operation(fail_with_nonempty_destination)

    assert attempts == 1
    assert retry_delays == []


def test_commit_failure_rolls_back_archive_output_and_monthly_copy(tmp_path: Path, monkeypatch):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9355 Virginia Beach" / "activity" / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    output_path = (
        paths["output"]
        / "9355 Virginia Beach"
        / "2026"
        / "Gift_Card_Reconciliation_9355_2026-W27.xlsx"
    )
    real_replace = weekly_service.os.replace
    canonical_attempts = 0
    retry_delays: list[float] = []

    def fail_canonical_commit(source, destination):
        nonlocal canonical_attempts
        if Path(destination) == output_path:
            canonical_attempts += 1
            raise OSError(errno.ENOSPC, "injected canonical commit failure", str(destination))
        return real_replace(source, destination)

    monkeypatch.setattr(weekly_service.os, "replace", fail_canonical_commit)
    monkeypatch.setattr(weekly_service.time, "sleep", retry_delays.append)

    report = run(paths)[0]

    assert report.status == "skipped"
    assert "committing canonical workbook" in report.message
    assert canonical_attempts == 1
    assert retry_delays == []
    assert activity_path.exists()
    assert not output_path.exists()
    assert not (paths["archive"] / "9355 Virginia Beach" / "2026" / "2026-W27").exists()
    assert not any(path.is_file() for path in paths["monthly"].rglob("*"))


def test_exhausted_dropbox_monthly_lock_rolls_back_every_destination(tmp_path: Path, monkeypatch):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9355 Virginia Beach" / "activity" / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    package_path = paths["archive"] / "9355 Virginia Beach" / "2026" / "2026-W27"
    output_path = (
        paths["output"]
        / "9355 Virginia Beach"
        / "2026"
        / "Gift_Card_Reconciliation_9355_2026-W27.xlsx"
    )
    monthly_path = (
        paths["monthly"]
        / "9355 Virginia Beach"
        / "FY27 M01 - Fiscal June"
        / "activity"
        / activity_path.name
    )
    monthly_attempts = 0
    retry_delays: list[float] = []
    stage_roots: list[Path] = []
    runtime_temp = tmp_path / "runtime-temp"
    runtime_temp.mkdir()
    real_replace = weekly_service.os.replace
    real_mkdtemp = weekly_service.tempfile.mkdtemp

    def keep_monthly_locked(source, destination):
        nonlocal monthly_attempts
        if Path(destination) == monthly_path:
            monthly_attempts += 1
            exc = PermissionError(errno.EACCES, "simulated persistent Dropbox lock", str(destination))
            exc.winerror = 32
            raise exc
        return real_replace(source, destination)

    def capture_stage_root(*args, **kwargs):
        if kwargs.get("dir") is None:
            kwargs["dir"] = runtime_temp
        path = Path(real_mkdtemp(*args, **kwargs))
        stage_roots.append(path)
        return str(path)

    monkeypatch.setattr(weekly_service.os, "replace", keep_monthly_locked)
    monkeypatch.setattr(weekly_service.tempfile, "mkdtemp", capture_stage_root)
    monkeypatch.setattr(weekly_service.time, "sleep", retry_delays.append)

    report = run(paths)[0]

    assert report.status == "skipped"
    assert "committing monthly-close activity" in report.message
    assert monthly_attempts == len(weekly_service._TRANSIENT_FILE_RETRY_DELAYS) + 1
    assert retry_delays == list(weekly_service._TRANSIENT_FILE_RETRY_DELAYS)
    assert len(stage_roots) == 1
    assert not stage_roots[0].exists()
    assert activity_path.exists()
    assert not package_path.exists()
    assert not output_path.exists()
    assert not monthly_path.exists()
    assert not list(paths["archive"].glob(".weekly-staging-*"))


def test_source_mutation_during_render_blocks_manifest_and_publication(tmp_path: Path, monkeypatch):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9355 Virginia Beach" / "activity" / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    micros_dir = create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    tender_path = micros_dir / "TENDER_DETAIL.TXT"
    real_writer = weekly_service.write_reconciliation_workbook

    def mutate_source_after_render(*args, **kwargs):
        result = real_writer(*args, **kwargs)
        with tender_path.open("a", encoding="utf-8") as stream:
            stream.write("\n2026-07-05,1.00,,G C Payment")
        return result

    monkeypatch.setattr(weekly_service, "write_reconciliation_workbook", mutate_source_after_render)

    report = run(paths)[0]

    assert report.status == "skipped"
    assert "Source evidence changed while weekly reconciliation was running" in report.message
    assert "TENDER_DETAIL.TXT" in report.message
    assert activity_path.exists()
    assert not paths["output"].exists()
    assert not paths["archive"].exists()
    assert not paths["monthly"].exists()


def test_one_cent_pos_variance_is_overall_review_in_workbook(tmp_path: Path):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9355 Virginia Beach" / "activity" / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.01"),
        payment=Decimal("20.00"),
    )

    report = run(paths)[0]

    assert report.status == "created"
    assert report.close_status == "REVIEW"
    workbook = load_workbook(report.output_path, data_only=True)
    conclusion = str(workbook["Reconciliation"]["A3"].value)
    assert "Weekly overall status: REVIEW" in conclusion
    assert "+0.01" in conclusion


def test_tender_only_variance_is_overall_review_in_workbook(tmp_path: Path):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9355 Virginia Beach" / "activity" / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
        tender_payment=Decimal("19.99"),
    )

    report = run(paths)[0]

    assert report.status == "created"
    assert report.close_status == "REVIEW"
    workbook = load_workbook(report.output_path, data_only=True)
    conclusion = str(workbook["Reconciliation"]["A3"].value)
    assert "Weekly overall status: REVIEW" in conclusion
    assert "tender-detail variance is +0.01" in conclusion


def test_rollback_reports_retained_path_when_cleanup_fails(tmp_path: Path, monkeypatch):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9355 Virginia Beach" / "activity" / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 29),
        end=date(2026, 7, 5),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    create_micros_week(
        paths["operations"],
        store="9355",
        week_start=date(2026, 6, 29),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )
    output_path = (
        paths["output"]
        / "9355 Virginia Beach"
        / "2026"
        / "Gift_Card_Reconciliation_9355_2026-W27.xlsx"
    )
    monthly_path = (
        paths["monthly"]
        / "9355 Virginia Beach"
        / "FY27 M01 - Fiscal June"
        / "activity"
        / activity_path.name
    )
    real_replace = weekly_service.os.replace
    real_unlink = Path.unlink

    def fail_monthly_commit(source, destination):
        if Path(destination) == monthly_path:
            raise OSError("injected monthly commit failure")
        return real_replace(source, destination)

    def fail_output_cleanup(self, *args, **kwargs):
        if self == output_path:
            raise OSError("injected output cleanup failure")
        return real_unlink(self, *args, **kwargs)

    monkeypatch.setattr(weekly_service.os, "replace", fail_monthly_commit)
    monkeypatch.setattr(Path, "unlink", fail_output_cleanup)

    report = run(paths)[0]

    assert report.status == "skipped"
    assert "Rollback incomplete" in report.message
    assert f"retained path {output_path}" in report.message
    assert output_path.exists()
    assert activity_path.exists()
    assert not (paths["archive"] / "9355 Virginia Beach" / "2026" / "2026-W27").exists()


def test_renamed_excel_file_is_parsed_and_blocked_not_silently_ignored(tmp_path: Path):
    paths = make_layout(tmp_path)
    renamed_path = paths["input"] / "9355 Virginia Beach" / "activity" / "renamed.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "not a Gift Card Activity report"
    workbook.save(renamed_path)

    report = run(paths)[0]

    assert report.status == "skipped"
    assert "renamed.xlsx" in report.message
    assert renamed_path.exists()
    assert not paths["output"].exists()
    assert not paths["archive"].exists()


def test_non_monday_sunday_activity_range_blocks_before_micros_read(tmp_path: Path):
    paths = make_layout(tmp_path)
    activity_path = paths["input"] / "9355 Virginia Beach" / "activity" / "07.06.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin=date(2026, 6, 30),
        end=date(2026, 7, 6),
        issue=Decimal("10.00"),
        payment=Decimal("20.00"),
    )

    report = run(paths)[0]

    assert report.status == "skipped"
    assert "expected an exact Monday-Sunday week" in report.message
    assert activity_path.exists()
    assert not paths["output"].exists()
    assert not paths["archive"].exists()


def make_layout(tmp_path: Path, *, stores: tuple[str, ...] = ("9355 Virginia Beach",)) -> dict[str, Path]:
    operations = tmp_path / "Gift Card Reconciliation"
    input_root = operations / "01 Weekly Gift Card Activity Reports"
    for store in stores:
        (input_root / store / "activity").mkdir(parents=True)
    return {
        "operations": operations,
        "input": input_root,
        "output": operations / "03 Finished Reports" / "Weekly",
        "monthly": operations / "02 Monthly Close Inputs",
        "archive": operations / "04 Archive" / "Weekly Reconciliation",
        "review": operations / "_automation_runs" / "review",
    }


def run(paths: dict[str, Path]):
    return run_weekly_reconciliations(
        operations_root=paths["operations"],
        input_root=paths["input"],
        output_dir=paths["output"],
        monthly_close_root=paths["monthly"],
        archive_root=paths["archive"],
        review_root=paths["review"],
    )


def create_activity(
    path: Path,
    *,
    store: str,
    begin: date,
    end: date,
    issue: Decimal,
    payment: Decimal,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet 1"
    sheet.append(
        [
            f"All GC Activity BY Rest Number and Date Range  "
            f"BEGIN DATE: '{begin:%d-%b-%Y}', END DATE: '{end:%d-%b-%Y}'"
        ]
    )
    sheet.append(["Response:000"])
    sheet.append([f"Rest Number:{store}"])
    sheet.append(
        [
            "Card No",
            "Request",
            "Request Code Listing",
            "Business Date",
            "Corp Code",
            "Transaction No",
            "Amount SUM",
            "Promocode",
            "Authorization Code",
        ]
    )
    sheet.append(["0001xxxx", 100, "Activation", end.isoformat(), None, 1, float(issue), None, 111111])
    sheet.append(["0002xxxx", 202, "Redemption No Nsf", end.isoformat(), None, 2, float(-payment), None, 222222])
    workbook.save(path)


def create_micros_week(
    operations_root: Path,
    *,
    store: str,
    week_start: date,
    issue: Decimal,
    payment: Decimal,
    tender_payment: Decimal | None = None,
    omit_dates: set[date] | None = None,
) -> Path:
    config = get_store_config(store)
    micros_dir = (operations_root / config.micros_default_path).resolve()
    micros_dir.mkdir(parents=True, exist_ok=True)
    omit_dates = omit_dates or set()
    tender_payment = payment if tender_payment is None else tender_payment
    system_rows: list[list[str]] = []
    tender_rows: list[list[str]] = []
    for offset in range(7):
        business_date = week_start + timedelta(days=offset)
        if business_date in omit_dates:
            continue
        is_final = offset == 6
        row = [""] * 121
        row[0] = business_date.isoformat()
        row[config.micros_issue_column_index] = str(issue if is_final else Decimal("0.00"))
        row[config.micros_payment_column_index] = str(payment if is_final else Decimal("0.00"))
        system_rows.append(row)
        tender_rows.append(
            [
                business_date.isoformat(),
                str(tender_payment if is_final else Decimal("0.00")),
                "",
                "G C Payment",
            ]
        )
    with (micros_dir / "DLYSYSTT.TXT").open("w", encoding="utf-8", newline="") as stream:
        csv.writer(stream).writerows(system_rows)
    with (micros_dir / "TENDER_DETAIL.TXT").open("w", encoding="utf-8", newline="") as stream:
        csv.writer(stream).writerows(tender_rows)
    return micros_dir
