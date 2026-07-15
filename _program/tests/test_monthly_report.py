from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from gift_card_recon.close_assessment import (
    CloseStatus,
    ControlDisposition,
    ControlOutcome,
    build_close_assessment,
)
from gift_card_recon.models import DardenCreditMemo, MonthlyCloseCertification
from gift_card_recon.monthly_report import (
    MonthlyCloseReportData,
    WeeklyCloseReportRow,
    write_monthly_close_report,
    write_monthly_close_report_workbook,
)


def make_certification(store: str) -> MonthlyCloseCertification:
    memo = DardenCreditMemo(
        source_file=Path("Darden June FY27.pdf"),
        store=store,
        period_start=date(2026, 6, 1),
        period_end=date(2026, 7, 5),
        total=Decimal("-26722.95"),
    )
    return MonthlyCloseCertification(
        store=store,
        period="FY27-M01",
        period_start=date(2026, 6, 1),
        period_end=date(2026, 7, 5),
        summary_net_settlement=Decimal("-26722.95"),
        darden_credit_memo=memo,
        variance=Decimal("0.00"),
    )


def make_assessment(store: str, disposition: ControlDisposition):
    control = ControlOutcome(
        code="pos_period_net",
        label="Period POS net",
        disposition=disposition,
        message=(
            "Period POS net has no variance."
            if disposition is ControlDisposition.PASS
            else "Period POS net variance is +2.43; review before sign-off."
        ),
        variance=Decimal("0.00") if disposition is ControlDisposition.PASS else Decimal("2.43"),
    )
    return build_close_assessment(
        store=store,
        darden_variance=Decimal("0.00"),
        controls=(control,),
    )


def find_row(ws, value: str) -> int | None:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == value:
            return row
    return None


def test_report_uses_assessment_for_amber_status_and_has_two_intentional_pages():
    assessment = make_assessment("9354", ControlDisposition.REVIEW)
    generated_at = datetime(2026, 7, 10, 14, 30)
    weekly = WeeklyCloseReportRow(
        week_ending=date(2026, 6, 28),
        coverage="Complete; scheduled Monday closed with zero evidence",
        pos_issue_variance=Decimal("0.00"),
        pos_payment_variance=Decimal("2.44"),
        pos_net_variance=Decimal("-2.44"),
        tender_variance=Decimal("0.00"),
        disposition=ControlDisposition.REVIEW,
        evidence_note="Document the small POS payment variance.",
    )
    data = MonthlyCloseReportData(
        assessment=assessment,
        period="FY27-M01",
        period_start=date(2026, 6, 1),
        period_end=date(2026, 7, 5),
        generated_at=generated_at,
        certification=make_certification("9354"),
        weekly_rows=(weekly,),
        period_pos_net_variance=Decimal("2.43"),
        period_pos_disposition=ControlDisposition.REVIEW,
        period_tender_variance=Decimal("0.00"),
        period_tender_disposition=ControlDisposition.PASS,
    )

    wb = Workbook()
    ws = wb.active
    write_monthly_close_report(ws, data)

    assert assessment.status is CloseStatus.CLOSED_WITH_REVIEW
    assert ws["A1"].value == "RICHMOND - STORE 9354"
    assert ws["A1"].font.name == "Arial"
    assert ws["A1"].font.sz == 19
    assert ws["A2"].font.name == "Arial"
    assert ws["A2"].font.sz == 10
    assert ws["A4"].value == "CLOSED WITH REVIEW"
    assert ws["A4"].fill.fgColor.rgb.endswith("FFF2CC")
    assert ws["G8"].value == "Match Status"
    assert ws["G9"].value == "MATCHED"
    assert ws["A9"].number_format == '$#,##0.00;($#,##0.00);$0.00'
    assert "Red" not in ws["A9"].number_format
    assert ws["A9"].fill.fgColor.rgb.endswith("F2F2F2")
    assert ws["C9"].fill.fgColor.rgb.endswith("FFFFFF")
    assert "Generated July 10, 2026 at 02:30 PM" in ws["A2"].value
    assert wb.properties.title == "RICHMOND - STORE 9354 FY27-M01 Monthly Close Report"
    assert wb.properties.subject == "Gift Card Monthly Close Reconciliation"
    assert wb.properties.creator == "Gift Card Reconciliation Close Control"
    assert "Executive accounting close certificate" in wb.properties.description

    assert len(ws.row_breaks.brk) == 1
    page_two_start = ws.row_breaks.brk[0].id + 1
    assert ws.cell(page_two_start, 1).value == "RICHMOND - STORE 9354"
    assert ws.cell(page_two_start + 1, 1).value.startswith("FY27-M01")
    assert ws.page_setup.fitToWidth == 0
    assert ws.page_setup.fitToHeight == 0
    assert ws.page_setup.scale == 85
    assert ws.oddFooter.center.text == "Page &P of &N"
    assert ws.oddFooter.left.text == "Generated 07/10/2026 02:30 PM"
    assert ws.oddFooter.right.text == "Richmond | FY27-M01"

    assert find_row(ws, "Settlement Tie-Out") == 7
    assert ws["A7"].font.name == "Arial"
    assert ws["A7"].font.sz == 11
    assert find_row(ws, "Close Controls") is not None
    assert find_row(ws, "Open Items Summary") is not None
    weekly_title = find_row(ws, "Weekly Variance Detail")
    assert weekly_title is not None
    assert ws.cell(weekly_title + 1, 2).value == "Coverage"
    assert ws.cell(weekly_title + 1, 7).value == "Status"
    assert ws.cell(weekly_title + 2, 6).number_format.startswith("$#,##0.00")
    assert ws.cell(weekly_title + 2, 7).value == "REVIEW"
    assert ws.cell(weekly_title + 2, 8).value == (
        "POS payment +$2.44; POS net -$2.44"
    )

    assert find_row(ws, "Variance Summary") is not None
    highlight = find_row(ws, "Largest Weekly Variance")
    assert highlight is not None
    assert ws.cell(highlight, 4).value == 2.44
    assert "POS payment" in ws.cell(highlight, 5).value
    assert ws.cell(highlight + 1, 1).value == "Period POS Net Variance"
    assert ws.cell(highlight + 1, 4).value == 2.43
    assert ws.cell(highlight + 1, 5).value == "REVIEW"

    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert "No review items." not in values
    assert any("review before sign-off" in str(value) for value in values)
    assert "Evidence and Audit Trail" in values
    evidence_row = find_row(ws, "Source package")
    assert evidence_row is not None
    assert ws.cell(evidence_row, 1).font.sz == 9.5
    assert all(
        "—" not in str(value) and "–" not in str(value)
        for value in values
    )
    for merged in ws.merged_cells.ranges:
        for row in ws.iter_rows(
            min_row=merged.min_row,
            max_row=merged.max_row,
            min_col=merged.min_col,
            max_col=merged.max_col,
        ):
            for cell in row:
                assert cell.border.left.style == "thin"
                assert cell.border.right.style == "thin"
                assert cell.border.top.style == "thin"
                assert cell.border.bottom.style == "thin"


def test_green_report_uses_virginia_beach_heading_and_no_exception_message():
    assessment = make_assessment("9355", ControlDisposition.PASS)
    data = MonthlyCloseReportData(
        assessment=assessment,
        period="FY27-M01",
        period_start=date(2026, 6, 1),
        period_end=date(2026, 7, 5),
        generated_at=datetime(2026, 7, 10, 9, 0),
        certification=make_certification("9355"),
        weekly_rows=(
            WeeklyCloseReportRow(
                week_ending=date(2026, 7, 5),
                coverage="Complete",
                pos_issue_variance=Decimal("0.00"),
                pos_payment_variance=Decimal("0.00"),
                pos_net_variance=Decimal("0.00"),
                tender_variance=Decimal("0.00"),
                disposition=ControlDisposition.PASS,
            ),
        ),
        period_pos_net_variance=Decimal("0.00"),
        period_pos_disposition=ControlDisposition.PASS,
        period_tender_variance=Decimal("0.00"),
        period_tender_disposition=ControlDisposition.PASS,
    )

    wb = Workbook()
    ws = wb.active
    write_monthly_close_report(ws, data)

    assert ws["A1"].value == "VIRGINIA BEACH - STORE 9355"
    assert ws["A4"].value == "CLOSED"
    assert ws["A4"].fill.fgColor.rgb.endswith("E2F0D9")
    assert find_row(ws, "No open items. All close controls passed.") is not None
    assert find_row(ws, "No review items.") is not None
    weekly_title = find_row(ws, "Weekly Variance Detail")
    assert weekly_title is not None
    assert ws.cell(weekly_title + 2, 8).value == "-"
    highlight = find_row(ws, "Largest Weekly Variance")
    assert highlight is not None
    assert ws.cell(highlight, 4).value == 0.0
    assert ws.cell(highlight, 5).value == "No weekly variance"


def test_unevaluated_darden_is_not_reported_as_a_mismatch():
    assessment = build_close_assessment(
        store="9355",
        darden_variance=None,
        controls=(
            ControlOutcome(
                "evidence_failure",
                "Evidence failure",
                ControlDisposition.BLOCK,
                "Tender evidence is missing.",
            ),
        ),
    )
    data = MonthlyCloseReportData(
        assessment=assessment,
        period="FY27-M01",
        period_start=date(2026, 6, 1),
        period_end=date(2026, 7, 5),
        generated_at=datetime(2026, 7, 10, 12, 0),
    )
    wb = Workbook()
    ws = wb.active

    write_monthly_close_report(ws, data)

    assert ws["G9"].value == "NOT EVALUATED"


def test_report_data_rejects_conflicting_darden_match_states():
    assessment = build_close_assessment(
        store="9354",
        darden_variance=Decimal("1.00"),
        controls=(),
    )

    with pytest.raises(
        ValueError,
        match="Certification Darden match state does not match",
    ):
        MonthlyCloseReportData(
            assessment=assessment,
            period="FY27-M01",
            period_start=date(2026, 6, 1),
            period_end=date(2026, 7, 5),
            generated_at=datetime(2026, 7, 10, 12, 0),
            certification=make_certification("9354"),
        )


def test_blocking_control_is_red_even_when_darden_is_matched(tmp_path: Path):
    assessment = make_assessment("9354", ControlDisposition.BLOCK)
    data = MonthlyCloseReportData(
        assessment=assessment,
        period="FY27-M01",
        period_start=date(2026, 6, 1),
        period_end=date(2026, 7, 5),
        generated_at=datetime(2026, 7, 10, 10, 0),
        certification=make_certification("9354"),
    )

    wb = Workbook()
    ws = wb.active
    write_monthly_close_report(ws, data)

    assert ws["A4"].value == "REVIEW REQUIRED"
    assert ws["A4"].fill.fgColor.rgb.endswith("F4CCCC")
    assert ws["G9"].value == "MATCHED"
    assert ws["G9"].fill.fgColor.rgb.endswith("E2F0D9")
    assert any(
        "No weekly detail was available" in str(cell.value)
        for row in ws.iter_rows()
        for cell in row
    )

    diagnostic_path = tmp_path / "Monthly Close - Review Required" / "Richmond_diagnostic.xlsx"
    assert write_monthly_close_report_workbook(data, diagnostic_path) == diagnostic_path
    assert diagnostic_path.is_file()


def test_realistic_many_control_assessment_remains_two_pages_and_keeps_all_followups():
    controls: list[ControlOutcome] = []
    integrity_codes = (
        "summary_identity",
        "activity_identity",
        "activity_coverage",
        "darden_identity",
        "micros_source",
        "micros_coverage",
        "tender_evidence",
    )
    for code in integrity_codes:
        controls.append(
            ControlOutcome(
                code=code,
                label=code.replace("_", " ").title(),
                disposition=ControlDisposition.PASS,
                message="Passed.",
            )
        )
    for index in range(3):
        controls.append(
            ControlOutcome(
                code=f"summary_activity_metric_{index}",
                label=f"Summary to activity metric {index}",
                disposition=ControlDisposition.PASS,
                message="Matches to the cent.",
                variance=Decimal("0.00"),
            )
        )

    weekly_codes: set[str] = set()
    weekly_rows: list[WeeklyCloseReportRow] = []
    for week_index, day in enumerate((7, 14, 21, 28, 5), start=1):
        month = 6 if week_index < 5 else 7
        week_ending = date(2026, month, day)
        weekly_disposition = ControlDisposition.REVIEW if week_index in {2, 4} else ControlDisposition.PASS
        weekly_rows.append(
            WeeklyCloseReportRow(
                week_ending=week_ending,
                coverage="Complete",
                pos_issue_variance=Decimal("0.00"),
                pos_payment_variance=Decimal("2.43") if weekly_disposition is ControlDisposition.REVIEW else Decimal("0.00"),
                pos_net_variance=Decimal("-2.43") if weekly_disposition is ControlDisposition.REVIEW else Decimal("0.00"),
                tender_variance=Decimal("0.00"),
                disposition=weekly_disposition,
            )
        )
        for metric in ("POS issue", "POS payment", "POS net", "Tender"):
            code = f"{'tender' if metric == 'Tender' else 'pos'}_week_{week_index}_{metric.lower().replace(' ', '_')}"
            weekly_codes.add(code)
            disposition = (
                ControlDisposition.REVIEW
                if weekly_disposition is ControlDisposition.REVIEW and metric in {"POS payment", "POS net"}
                else ControlDisposition.PASS
            )
            label = f"Week ending {week_ending:%m/%d/%Y} {metric}"
            controls.append(
                ControlOutcome(
                    code=code,
                    label=label,
                    disposition=disposition,
                    message="Review the small weekly variance." if disposition is ControlDisposition.REVIEW else "No variance.",
                    variance=Decimal("2.43") if disposition is ControlDisposition.REVIEW else Decimal("0.00"),
                )
            )

    controls.extend(
        (
            ControlOutcome("pos_period_issue", "Period POS issue", ControlDisposition.PASS, "No variance.", Decimal("0.00")),
            ControlOutcome("pos_period_payment", "Period POS payment", ControlDisposition.REVIEW, "Review +2.43.", Decimal("2.43")),
            ControlOutcome("pos_period_net", "Period POS net", ControlDisposition.REVIEW, "Review +2.43.", Decimal("2.43")),
            ControlOutcome("tender_period", "Period tender", ControlDisposition.PASS, "No variance.", Decimal("0.00")),
        )
    )
    assessment = build_close_assessment(
        store="9354",
        darden_variance=Decimal("0.00"),
        controls=controls,
    )
    data = MonthlyCloseReportData(
        assessment=assessment,
        period="FY27-M01",
        period_start=date(2026, 6, 1),
        period_end=date(2026, 7, 5),
        generated_at=datetime(2026, 7, 10, 12, 0),
        certification=make_certification("9354"),
        weekly_rows=tuple(weekly_rows),
        period_pos_net_variance=Decimal("2.43"),
        period_pos_disposition=ControlDisposition.REVIEW,
        period_tender_variance=Decimal("0.00"),
        period_tender_disposition=ControlDisposition.PASS,
        explicit_exceptions=tuple(
            (
                control.disposition.value,
                f"{control.label}: {control.message}",
            )
            for control in assessment.controls
            if not control.passed
        ),
        weekly_control_codes=frozenset(weekly_codes),
    )

    wb = Workbook()
    ws = wb.active
    write_monthly_close_report(ws, data)

    assert len(assessment.controls) == 35
    assert len(ws.row_breaks.brk) == 1
    assert ws.page_setup.fitToWidth == 0
    assert ws.page_setup.fitToHeight == 0
    assert ws.page_setup.scale == 85
    page_one_end = ws.row_breaks.brk[0].id
    assert ws["A1"].value == "RICHMOND - STORE 9354"

    matrix_start = find_row(ws, "Close Controls") + 2
    actions_start = find_row(ws, "Open Items Summary")
    matrix_text = " ".join(
        str(ws.cell(row, column).value or "")
        for row in range(matrix_start, actions_start)
        for column in range(1, 9)
    )
    assert "Week ending" not in matrix_text
    assert "Period POS net" in matrix_text

    page_one_text = " ".join(
        str(ws.cell(row, column).value or "")
        for row in range(1, page_one_end + 1)
        for column in range(1, 9)
    )
    assert "week(s) contain reviewed controls" in page_one_text
    assert "Period POS payment" in page_one_text
    assert "Period POS net" in page_one_text
    review_start = find_row(ws, "Review Items")
    evidence_start = find_row(ws, "Evidence and Audit Trail")
    review_text = " ".join(
        str(ws.cell(row, column).value or "")
        for row in range(review_start + 1, evidence_start)
        for column in range(1, 9)
    )
    assert review_text.count("Week ending 06/14/2026") == 1
    assert review_text.count("Week ending 06/28/2026") == 1
    assert "POS payment +$2.43; POS net -$2.43" in review_text
    assert "Period POS payment" in review_text
    assert "Period POS net" in review_text
    assert "Exception" not in review_text
