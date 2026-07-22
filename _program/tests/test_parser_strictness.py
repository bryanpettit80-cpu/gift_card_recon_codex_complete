from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
import pytest

from gift_card_recon.parsers import ParseError, parse_activity_file, parse_summary


def test_summary_never_falls_back_to_a_wrong_store_row(tmp_path: Path) -> None:
    path = tmp_path / "summary.xlsx"
    _write_summary(path, rows=[["Sorensen", 9354, 100, -300, -200]])

    with pytest.raises(ParseError, match="does not contain a row for store 9355"):
        parse_summary(path, store="9355")


def test_summary_requires_exactly_one_matching_store_row(tmp_path: Path) -> None:
    path = tmp_path / "summary.xlsx"
    _write_summary(
        path,
        rows=[
            ["Sorensen", 9355, 100, -300, -200],
            ["Sorensen duplicate", 9355, 100, -300, -200],
        ],
    )

    with pytest.raises(ParseError, match="multiple rows for store 9355"):
        parse_summary(path, store="9355")


def test_summary_rejects_malformed_required_money(tmp_path: Path) -> None:
    path = tmp_path / "summary.xlsx"
    _write_summary(path, rows=[["Sorensen", 9355, "not money", -300, -200]])

    with pytest.raises(ParseError, match="Total Activations"):
        parse_summary(path, store="9355")


def test_summary_required_money_does_not_use_a_near_match_header(tmp_path: Path) -> None:
    path = tmp_path / "summary.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["SUMMARY"])
    sheet.append(
        [
            "Franchise Partner",
            "Store Number",
            "Total Activations",
            "Total Redemptions",
            "Adjusted Net Settlement",
        ]
    )
    sheet.append(["Sorensen", 9355, 100, -300, -200])
    workbook.save(path)

    with pytest.raises(ParseError, match="Net Settlement"):
        parse_summary(path, store="9355")


def test_activity_transaction_row_with_blank_amount_is_blocking(tmp_path: Path) -> None:
    path = tmp_path / "06.07.2026 9355 Gift Card Activity.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "BEGIN DATE: '01-JUN-2026', END DATE: '07-JUN-2026', "
            "Rest Number Parameter 1: '9355'"
        ]
    )
    sheet.append(
        [
            "Card No",
            "Request",
            "Request Code Listing",
            "Business Date",
            "Transaction No",
            "Amount SUM",
        ]
    )
    sheet.append(["0001", 100, "Activation", "2026-06-02", 1, None])
    workbook.save(path)

    with pytest.raises(ParseError, match="missing required field.*Amount"):
        parse_activity_file(path)


def test_activity_transaction_row_with_blank_business_date_is_blocking(tmp_path: Path) -> None:
    path = tmp_path / "06.07.2026 9355 Gift Card Activity.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "BEGIN DATE: '01-JUN-2026', END DATE: '07-JUN-2026', "
            "Rest Number Parameter 1: '9355'"
        ]
    )
    sheet.append(
        [
            "Card No",
            "Request",
            "Request Code Listing",
            "Business Date",
            "Transaction No",
            "Amount SUM",
        ]
    )
    sheet.append(["0001", 100, "Activation", None, 1, 25])
    workbook.save(path)

    with pytest.raises(ParseError, match="missing required field.*Business Date"):
        parse_activity_file(path)


def _write_summary(path: Path, *, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["SUMMARY"])
    sheet.append(
        [
            "Franchise Partner",
            "Store Number",
            "Total Activations",
            "Total Redemptions",
            "Net Settlement",
        ]
    )
    for row in rows:
        sheet.append(row)
    workbook.save(path)
