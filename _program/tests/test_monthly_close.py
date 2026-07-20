from __future__ import annotations

from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from gift_card_recon.darden import build_monthly_close_certification, parse_darden_credit_memo_text
from gift_card_recon.fiscal_calendar import fiscal_period_for_date, fiscal_period_for_label
from gift_card_recon.monthly_close import (
    ISSUE_AMOUNT_INDEX,
    PAYMENT_AMOUNT_INDEX,
    build_weekly_pos_variances,
    cleanup_monthly_close_sources,
    format_monthly_close_preflight,
    monthly_activity_week_endings,
    parse_micros_daily_pos_controls,
    prepare_monthly_close_inputs,
    run_monthly_close,
    stage_darden_credit_memo,
    validate_tender_payment_totals,
)
from gift_card_recon.models import DardenCreditMemo
from gift_card_recon.parsers import ParseError
from gift_card_recon.parsers import parse_activity_file

REPO_ROOT = Path(__file__).resolve().parents[2]


def test_darden_fiscal_calendar_maps_july_5_2026_to_fiscal_june():
    period = fiscal_period_for_date(date(2026, 7, 5))

    assert period.period_key == "FY27-M01"
    assert period.folder_name == "FY27 M01 - Fiscal June"
    assert period.start_date == date(2026, 6, 1)
    assert period.end_date == date(2026, 7, 5)
    assert period.expected_week_endings == [
        date(2026, 6, 7),
        date(2026, 6, 14),
        date(2026, 6, 21),
        date(2026, 6, 28),
        date(2026, 7, 5),
    ]


def test_darden_fiscal_calendar_accepts_legacy_year_month_label():
    period = fiscal_period_for_label("2026-06")

    assert period.period_key == "FY27-M01"
    assert period.folder_name == "FY27 M01 - Fiscal June"


def test_darden_credit_memo_text_parses_location_period_and_signed_total(tmp_path: Path):
    report = parse_darden_credit_memo_text(
        """
        CREDIT MEMO
        INVOICE NUMBER Jun FY27 - Sorensen 2
        INVOICE DATE 7/6/2026
        06/01/2026-07/05/2026 Darden SV Gift Cards Activity
        Location 9355 4,840.00 278.01 (34,372.36) 2,531.40 (26,722.95)$
        TOTAL (26,722.95)$
        """,
        source_file=tmp_path / "Jun FY27 - Sorensen 2.pdf",
    )

    assert report.store == "9355"
    assert report.period_start == date(2026, 6, 1)
    assert report.period_end == date(2026, 7, 5)
    assert report.total == Decimal("-26722.95")
    assert report.invoice_number == "Jun FY27 - Sorensen 2"
    assert report.invoice_date == date(2026, 7, 6)


def test_darden_certification_rounds_summary_to_cents_and_preserves_sign(tmp_path: Path):
    summary_path = tmp_path / "07.05.2026 9355 Gift Card Summary.xlsx"
    create_summary(summary_path, store="9355", activations=Decimal("0.00"), redemptions=Decimal("-26722.949"))
    from gift_card_recon.parsers import parse_summary

    summary = parse_summary(summary_path, store="9355")
    report = create_darden_report(tmp_path / "memo.pdf", store="9355", total=Decimal("-26722.95"))
    certification = build_monthly_close_certification(
        store="9355",
        period="FY27-M01",
        period_start=date(2026, 6, 1),
        period_end=date(2026, 7, 5),
        summary=summary,
        darden_credit_memo=report,
    )

    assert certification.darden_matched
    assert certification.variance == Decimal("0.00")

    opposite_sign_report = create_darden_report(tmp_path / "opposite.pdf", store="9355", total=Decimal("26722.95"))
    opposite = build_monthly_close_certification(
        store="9355",
        period="FY27-M01",
        period_start=date(2026, 6, 1),
        period_end=date(2026, 7, 5),
        summary=summary,
        darden_credit_memo=opposite_sign_report,
    )
    assert not opposite.darden_matched
    assert opposite.variance == Decimal("53445.90")


def _legacy_monthly_close_generates_standard_workbook_with_weekly_pos_variance(tmp_path: Path):
    input_dir = tmp_path / "Monthly Close" / "9355" / "FY27 M01 - Fiscal June"
    summary_dir = input_dir / "summary"
    activity_dir = input_dir / "activity"
    micros_dir = tmp_path / "micros"
    summary_dir.mkdir(parents=True)
    activity_dir.mkdir()
    micros_dir.mkdir()

    create_summary(summary_dir / "07.05.2026 9355 Gift Card Summary.xlsx", store="9355", activations=Decimal("100.00"), redemptions=Decimal("-300.00"))
    create_activity(
        activity_dir / "06.07.2026 9355 Gift Card Activity.xlsx",
        store="9355",
        begin="01-JUN-2026",
        end="07-JUN-2026",
        activation=Decimal("100.00"),
        redemption=Decimal("-300.00"),
    )
    write_micros_exports(micros_dir, date(2026, 6, 1), [Decimal("15.00")] * 7, [Decimal("40.00")] * 6 + [Decimal("70.00")])

    output_path = tmp_path / "Output" / "Gift_Card_Reconciliation_9355_FY27-M01.xlsx"
    darden_report = create_darden_report(input_dir / "darden" / "Jun FY27 - Sorensen 2.pdf", store="9355", total=Decimal("-200.00"))
    saved_path, result, weekly_rows = run_monthly_close(
        store="9355",
        period="FY27-M01",
        period_start=date(2026, 6, 1),
        period_end=date(2026, 7, 5),
        input_dir=input_dir,
        output_path=output_path,
        micros_path=micros_dir,
        micros_work_dir=tmp_path / "extract",
        darden_report=darden_report,
    )

    assert saved_path == output_path
    assert result.pos_controls.pos_gift_card_issue == Decimal("105.00")
    assert result.pos_controls.pos_gift_card_payment == Decimal("310.00")
    assert weekly_rows[0].issue_variance == Decimal("5.00")
    assert weekly_rows[0].payment_variance == Decimal("10.00")
    assert weekly_rows[0].net_variance == Decimal("-5.00")

    wb = load_workbook(saved_path, data_only=False)
    assert wb.sheetnames == [
        "Monthly Close Report",
        "Reconciliation",
        "Weekly Activity Detail",
        "Daily Activity Detail",
        "Raw Detail",
        "Source Files",
        "Exception Log",
    ]
    report_ws = wb["Monthly Close Report"]
    assert report_ws["A4"].value.startswith("☑ CLOSED")
    assert report_ws["A14"].value == -200
    assert report_ws["C14"].value == -200
    assert report_ws["E14"].value == 0
    assert report_ws["G14"].value == "CLOSED"
    assert report_ws["A19"].value == "Stage"
    assert report_ws["C19"].value == "Evidence"
    assert report_ws["F19"].value == "Result"
    assert report_ws["H19"].value == "Owner / Next Action"
    ws = wb["Reconciliation"]
    section_row = find_row(ws, "Weekly POS Variance Detail")
    assert section_row is not None
    assert section_row > find_row(ws, "POS Controls Included on Reconciliation")
    assert ws.cell(section_row + 1, 1).value == "Week Ending"
    assert ws.cell(section_row + 2, 1).value.date() == date(2026, 6, 7)
    assert ws.cell(section_row + 2, 2).value == 100
    assert ws.cell(section_row + 2, 3).value == 105
    assert ws.cell(section_row + 2, 7).value == 10
    assert ws.cell(section_row + 3, 8).value == -5
    final_row = find_row(ws, "Darden Final Close Certification")
    assert final_row is not None
    assert ws.cell(final_row + 2, 1).value == "☑"
    assert ws.cell(final_row + 2, 6).value == "CLOSED"

    source_ws = wb["Source Files"]
    source_names = {source_ws.cell(row_idx, 1).value for row_idx in range(4, source_ws.max_row + 1)}
    assert darden_report.source_file.name in source_names


def _legacy_boundary_week_missing_dates_outside_period_is_held_to_activity_totals(tmp_path: Path):
    activity_path = tmp_path / "06.07.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin="25-MAY-2026",
        end="07-JUN-2026",
        activation=Decimal("100.00"),
        redemption=Decimal("-300.00"),
    )
    micros_dir = tmp_path / "micros"
    micros_dir.mkdir()
    write_micros_exports(micros_dir, date(2026, 6, 1), [Decimal("10.00")] * 7, [Decimal("20.00")] * 7)

    activity = parse_activity_file(activity_path)
    daily_controls = parse_micros_daily_pos_controls(micros_dir)
    rows = build_weekly_pos_variances(
        [activity],
        set(),
        daily_controls,
        period_start=date(2026, 6, 1),
        period_end=date(2026, 6, 30),
    )

    assert rows[0].coverage_status == "Boundary week adjusted to activity totals"
    assert rows[0].pos_issue == Decimal("100.00")
    assert rows[0].pos_payment == Decimal("300.00")
    assert rows[0].payment_variance == Decimal("0.00")


def test_missing_micros_dates_inside_period_are_reported_as_partial(tmp_path: Path):
    activity_path = tmp_path / "06.07.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin="01-JUN-2026",
        end="07-JUN-2026",
        activation=Decimal("100.00"),
        redemption=Decimal("-300.00"),
    )
    micros_dir = tmp_path / "micros"
    micros_dir.mkdir()
    write_micros_exports(micros_dir, date(2026, 6, 1), [Decimal("10.00")] * 4, [Decimal("20.00")] * 4)

    activity = parse_activity_file(activity_path)
    daily_controls = parse_micros_daily_pos_controls(micros_dir)
    rows = build_weekly_pos_variances(
        [activity],
        set(),
        daily_controls,
        period_start=date(2026, 6, 1),
        period_end=date(2026, 6, 30),
    )

    assert rows[0].coverage_status == "Partial Micros POS coverage"
    assert rows[0].pos_issue == Decimal("40.00")
    assert rows[0].pos_payment == Decimal("80.00")
    assert rows[0].payment_variance == Decimal("-220.00")


def _legacy_missing_closed_mondays_inside_period_are_not_reported_as_partial(tmp_path: Path):
    activity_path = tmp_path / "06.07.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        activity_path,
        store="9355",
        begin="01-JUN-2026",
        end="07-JUN-2026",
        activation=Decimal("100.00"),
        redemption=Decimal("-300.00"),
        transaction_date="2026-06-02",
    )
    micros_dir = tmp_path / "micros"
    micros_dir.mkdir()
    write_micros_export_rows(
        micros_dir,
        [
            (date(2026, 6, 2), Decimal("10.00"), Decimal("50.00")),
            (date(2026, 6, 3), Decimal("20.00"), Decimal("50.00")),
            (date(2026, 6, 4), Decimal("15.00"), Decimal("50.00")),
            (date(2026, 6, 5), Decimal("15.00"), Decimal("50.00")),
            (date(2026, 6, 6), Decimal("20.00"), Decimal("50.00")),
            (date(2026, 6, 7), Decimal("20.00"), Decimal("50.00")),
        ],
    )

    activity = parse_activity_file(activity_path)
    daily_controls = parse_micros_daily_pos_controls(micros_dir)
    rows = build_weekly_pos_variances(
        [activity],
        set(),
        daily_controls,
        period_start=date(2026, 6, 1),
        period_end=date(2026, 6, 30),
    )

    assert rows[0].coverage_status == "Closed days omitted from Micros POS coverage"
    assert rows[0].pos_issue == Decimal("100.00")
    assert rows[0].pos_payment == Decimal("300.00")
    assert rows[0].payment_variance == Decimal("0.00")


def test_quoted_tender_names_are_matched_for_payment_validation(tmp_path: Path):
    micros_dir = tmp_path / "micros"
    micros_dir.mkdir()
    business_date = date(2026, 6, 1)
    system_row = ["0"] * 132
    system_row[0] = f"{business_date:%Y-%m-%d} 00:00:00.000"
    system_row[ISSUE_AMOUNT_INDEX] = "0.00"
    system_row[PAYMENT_AMOUNT_INDEX] = "12.00"
    (micros_dir / "DLYSYSTT.TXT").write_text(",".join(system_row), encoding="utf-8")
    (micros_dir / "TENDER_DETAIL.TXT").write_text(
        f"'{business_date:%Y-%m-%d}',10.00,350,'G C Payment','T'",
        encoding="utf-8",
    )

    daily_controls = parse_micros_daily_pos_controls(micros_dir)
    exceptions = validate_tender_payment_totals(micros_dir, daily_controls, {business_date})

    assert exceptions == [
        (
            "Review",
            "2026-06-01 G C Payment tender total 10.00 does not match DLYSYSTT.TXT column 103 12.00.",
        )
    ]


def test_monthly_close_preflight_stages_weekly_files_and_reports_missing_inputs(tmp_path: Path):
    input_root = tmp_path / "Monthly Close"
    archive_dir = tmp_path / "9355 - Weekly" / "archive" / "2026-W25"
    current_dir = tmp_path / "9355 - Weekly" / "activity"
    archive_dir.mkdir(parents=True)
    current_dir.mkdir(parents=True)
    micros_dir = tmp_path / "micros"
    micros_dir.mkdir()

    create_activity(
        archive_dir / "06.21.2026 9355 Gift Card Activity.xlsx",
        store="9355",
        begin="15-JUN-2026",
        end="21-JUN-2026",
        activation=Decimal("100.00"),
        redemption=Decimal("-300.00"),
    )
    create_activity(
        current_dir / "06.28.2026 9355 Gift Card Activity.xlsx",
        store="9355",
        begin="22-JUN-2026",
        end="28-JUN-2026",
        activation=Decimal("100.00"),
        redemption=Decimal("-300.00"),
    )
    write_micros_exports(micros_dir, date(2026, 7, 5), [Decimal("0.00")], [Decimal("0.00")])
    fiscal_period = fiscal_period_for_label("FY27-M01")

    preflight = prepare_monthly_close_inputs(
        store="9355",
        period="FY27-M01",
        fiscal_period=fiscal_period,
        period_start=date(2026, 6, 1),
        period_end=date(2026, 7, 5),
        input_root=input_root,
        input_dir=input_root / "9355" / "FY27 M01 - Fiscal June",
        micros_path=micros_dir,
        micros_work_dir=tmp_path / "extract",
    )

    assert monthly_activity_week_endings(date(2026, 6, 1), date(2026, 7, 5)) == [
        date(2026, 6, 7),
        date(2026, 6, 14),
        date(2026, 6, 21),
        date(2026, 6, 28),
        date(2026, 7, 5),
    ]
    assert (input_root / "9355" / "FY27 M01 - Fiscal June" / "activity" / "06.21.2026 9355 Gift Card Activity.xlsx").exists()
    assert (input_root / "9355" / "FY27 M01 - Fiscal June" / "activity" / "06.28.2026 9355 Gift Card Activity.xlsx").exists()
    assert preflight.micros_ready
    assert preflight.missing_summary_path == input_root / "9355" / "FY27 M01 - Fiscal June" / "summary" / "07.05.2026 9355 Gift Card Summary.xlsx"
    assert preflight.missing_activity_paths == [
        input_root / "9355" / "FY27 M01 - Fiscal June" / "activity" / "06.07.2026 9355 Gift Card Activity.xls",
        input_root / "9355" / "FY27 M01 - Fiscal June" / "activity" / "06.14.2026 9355 Gift Card Activity.xls",
        input_root / "9355" / "FY27 M01 - Fiscal June" / "activity" / "07.05.2026 9355 Gift Card Activity.xls",
    ]
    report = format_monthly_close_preflight(preflight)
    assert "NOT READY" in report
    assert "07.05.2026 9355 Gift Card Summary.xlsx" in report
    assert "06.07.2026 9355 Gift Card Activity.xls" in report


def test_monthly_close_preflight_skips_symlinked_weekly_activity_candidates(tmp_path: Path):
    input_root = tmp_path / "Monthly Close"
    archive_dir = tmp_path / "9355 - Weekly" / "archive"
    archive_dir.mkdir(parents=True)
    outside_workbook = tmp_path / "outside" / "local_sensitive_workbook.xlsx"
    outside_workbook.parent.mkdir()
    wb = Workbook()
    ws = wb.active
    ws.title = "Not Activity"
    ws.append(["CONFIDENTIAL payroll workbook"])
    wb.save(outside_workbook)
    linked_activity = archive_dir / "06.07.2026 9355 Gift Card Activity.xlsx"
    try:
        linked_activity.symlink_to(outside_workbook)
    except OSError as exc:
        pytest.skip(f"Symlinks are not available in this test environment: {exc}")

    micros_dir = tmp_path / "micros"
    micros_dir.mkdir()
    write_micros_exports(micros_dir, date(2026, 7, 5), [Decimal("0.00")], [Decimal("0.00")])
    fiscal_period = fiscal_period_for_label("FY27-M01")

    preflight = prepare_monthly_close_inputs(
        store="9355",
        period=fiscal_period.period_key,
        fiscal_period=fiscal_period,
        period_start=fiscal_period.start_date,
        period_end=fiscal_period.end_date,
        input_root=input_root,
        input_dir=input_root / "9355" / fiscal_period.folder_name,
        micros_path=micros_dir,
        micros_work_dir=tmp_path / "extract",
    )

    destination = input_root / "9355" / fiscal_period.folder_name / "activity" / linked_activity.name
    assert preflight.staged_activity_paths == []
    assert not destination.exists()


@pytest.mark.parametrize("link_archive_root", [True, False])
def test_monthly_close_preflight_skips_linked_activity_directories(
    tmp_path: Path,
    link_archive_root: bool,
):
    input_root = tmp_path / "Monthly Close"
    weekly_root = tmp_path / "9355 - Weekly"
    weekly_root.mkdir()
    outside_dir = tmp_path / "outside"
    outside_dir.mkdir()
    linked_name = "06.07.2026 9355 Gift Card Activity.xlsx"
    create_activity(
        outside_dir / linked_name,
        store="9355",
        begin="01-JUN-2026",
        end="07-JUN-2026",
        activation=Decimal("100.00"),
        redemption=Decimal("-300.00"),
    )
    archive_dir = weekly_root / "archive"
    try:
        if link_archive_root:
            archive_dir.symlink_to(outside_dir, target_is_directory=True)
        else:
            archive_dir.mkdir()
            (archive_dir / "linked-week").symlink_to(outside_dir, target_is_directory=True)
    except OSError as exc:
        pytest.skip(f"Directory links are not available in this test environment: {exc}")

    micros_dir = tmp_path / "micros"
    micros_dir.mkdir()
    write_micros_exports(micros_dir, date(2026, 7, 5), [Decimal("0.00")], [Decimal("0.00")])
    fiscal_period = fiscal_period_for_label("FY27-M01")

    preflight = prepare_monthly_close_inputs(
        store="9355",
        period=fiscal_period.period_key,
        fiscal_period=fiscal_period,
        period_start=fiscal_period.start_date,
        period_end=fiscal_period.end_date,
        input_root=input_root,
        input_dir=input_root / "9355" / fiscal_period.folder_name,
        micros_path=micros_dir,
        micros_work_dir=tmp_path / "extract",
    )

    destination = input_root / "9355" / fiscal_period.folder_name / "activity" / linked_name
    assert preflight.staged_activity_paths == []
    assert not destination.exists()


def test_darden_staging_rejects_linked_destination_directory(tmp_path: Path):
    source = tmp_path / "Darden Credit Memo.pdf"
    source.write_bytes(b"source evidence")
    outside_dir = tmp_path / "outside"
    outside_dir.mkdir()
    darden_dir = tmp_path / "monthly" / "darden"
    darden_dir.parent.mkdir()
    try:
        darden_dir.symlink_to(outside_dir, target_is_directory=True)
    except OSError as exc:
        pytest.skip(f"Directory links are not available in this test environment: {exc}")

    with pytest.raises(ParseError, match="linked or reparse-point destination"):
        stage_darden_credit_memo(source, darden_dir=darden_dir)

    assert not (outside_dir / source.name).exists()


def test_darden_staging_skips_broken_link_collision_destination(tmp_path: Path):
    source = tmp_path / "Darden Credit Memo.pdf"
    source.write_bytes(b"new source evidence")
    darden_dir = tmp_path / "monthly" / "darden"
    darden_dir.mkdir(parents=True)
    (darden_dir / source.name).write_bytes(b"existing evidence")
    redirected_target = tmp_path / "outside" / "redirected.pdf"
    broken_collision = darden_dir / "Darden Credit Memo_2.pdf"
    try:
        broken_collision.symlink_to(redirected_target)
    except OSError as exc:
        pytest.skip(f"File links are not available in this test environment: {exc}")

    staged = stage_darden_credit_memo(source, darden_dir=darden_dir)

    assert staged == darden_dir / "Darden Credit Memo_3.pdf"
    assert staged.read_bytes() == b"new source evidence"
    assert broken_collision.is_symlink()
    assert not redirected_target.exists()


def test_monthly_close_preflight_uses_darden_pdf_as_final_gate(tmp_path: Path, monkeypatch):
    fiscal_period = fiscal_period_for_label("FY27-M01")
    input_root = tmp_path / "Monthly Close"
    input_dir = input_root / "9355" / fiscal_period.folder_name
    summary_path = input_dir / "summary" / "07.05.2026 9355 Gift Card Summary.xlsx"
    darden_path = input_dir / "darden" / "memo.pdf"
    summary_path.parent.mkdir(parents=True)
    darden_path.parent.mkdir()
    create_summary(summary_path, store="9355", activations=Decimal("100.00"), redemptions=Decimal("-300.00"))
    darden_path.write_bytes(b"synthetic")
    micros_dir = tmp_path / "micros"
    micros_dir.mkdir()
    write_micros_exports(micros_dir, fiscal_period.end_date, [Decimal("0.00")], [Decimal("0.00")])
    parsed_report = DardenCreditMemo(
        source_file=darden_path,
        store="9355",
        period_start=fiscal_period.start_date,
        period_end=fiscal_period.end_date,
        total=Decimal("-200.00"),
    )
    monkeypatch.setattr("gift_card_recon.monthly_close.parse_darden_credit_memo", lambda _path: parsed_report)

    preflight = prepare_monthly_close_inputs(
        store="9355",
        period=fiscal_period.period_key,
        fiscal_period=fiscal_period,
        period_start=fiscal_period.start_date,
        period_end=fiscal_period.end_date,
        input_root=input_root,
        input_dir=input_dir,
        micros_path=micros_dir,
        micros_work_dir=tmp_path / "extract",
        stage_weekly=False,
    )

    assert preflight.darden_ready
    assert preflight.darden_certification is not None
    assert preflight.darden_certification.darden_matched
    assert preflight.darden_message.startswith("MATCH")
    assert "Darden final close: MATCH" in format_monthly_close_preflight(preflight)


def test_monthly_close_preflight_moves_loose_summary_into_fiscal_period(tmp_path: Path):
    input_root = tmp_path / "Monthly Close"
    store_root = input_root / "9355"
    input_dir = store_root / "FY27 M01 - Fiscal June"
    activity_dir = input_dir / "activity"
    micros_dir = tmp_path / "micros"
    store_root.mkdir(parents=True)
    activity_dir.mkdir(parents=True)
    micros_dir.mkdir()
    loose_summary = store_root / "07.05.2026 9355 Gift Card Summary.xlsx"
    create_summary(loose_summary, store="9355", activations=Decimal("100.00"), redemptions=Decimal("-300.00"))
    create_activity(
        activity_dir / "07.05.2026 9355 Gift Card Activity.xlsx",
        store="9355",
        begin="29-JUN-2026",
        end="05-JUL-2026",
        activation=Decimal("100.00"),
        redemption=Decimal("-300.00"),
    )
    write_micros_exports(micros_dir, date(2026, 7, 5), [Decimal("0.00")], [Decimal("0.00")])
    fiscal_period = fiscal_period_for_label("FY27-M01")

    preflight = prepare_monthly_close_inputs(
        store="9355",
        period=fiscal_period.period_key,
        fiscal_period=fiscal_period,
        period_start=fiscal_period.start_date,
        period_end=fiscal_period.end_date,
        input_root=input_root,
        input_dir=input_dir,
        micros_path=micros_dir,
        micros_work_dir=tmp_path / "extract",
        stage_weekly=False,
    )

    expected_summary = input_dir / "summary" / loose_summary.name
    assert not loose_summary.exists()
    assert expected_summary.exists()
    assert preflight.summary_paths == [expected_summary]


def test_monthly_close_preflight_does_not_stage_wrong_store_summary(tmp_path: Path):
    input_root = tmp_path / "Monthly Close"
    store_root = input_root / "9355"
    input_dir = store_root / "FY27 M01 - Fiscal June"
    micros_dir = tmp_path / "micros"
    store_root.mkdir(parents=True)
    micros_dir.mkdir()
    loose_summary = store_root / "07.05.2026 attacker 9355 Gift Card Summary.xlsx"
    create_summary(loose_summary, store="9999", activations=Decimal("100.00"), redemptions=Decimal("-300.00"))
    original_bytes = loose_summary.read_bytes()
    write_micros_exports(micros_dir, date(2026, 7, 5), [Decimal("0.00")], [Decimal("0.00")])
    fiscal_period = fiscal_period_for_label("FY27-M01")

    preflight = prepare_monthly_close_inputs(
        store="9355",
        period=fiscal_period.period_key,
        fiscal_period=fiscal_period,
        period_start=fiscal_period.start_date,
        period_end=fiscal_period.end_date,
        input_root=input_root,
        input_dir=input_dir,
        micros_path=micros_dir,
        micros_work_dir=tmp_path / "extract",
        stage_weekly=False,
    )

    assert loose_summary.exists()
    assert loose_summary.read_bytes() == original_bytes
    assert not (input_dir / "summary" / loose_summary.name).exists()
    assert preflight.summary_paths == []


def test_monthly_close_preflight_does_not_read_linked_loose_summary(tmp_path: Path):
    input_root = tmp_path / "Monthly Close"
    store_root = input_root / "9355"
    input_dir = store_root / "FY27 M01 - Fiscal June"
    micros_dir = tmp_path / "micros"
    outside_summary = tmp_path / "outside" / "sensitive.xlsx"
    store_root.mkdir(parents=True)
    micros_dir.mkdir()
    outside_summary.parent.mkdir()
    outside_summary.write_bytes(b"sensitive non-summary content")
    linked_summary = store_root / "07.05.2026 9355 Gift Card Summary.xlsx"
    try:
        linked_summary.symlink_to(outside_summary)
    except OSError as exc:
        pytest.skip(f"Symlinks are not available in this test environment: {exc}")
    write_micros_exports(micros_dir, date(2026, 7, 5), [Decimal("0.00")], [Decimal("0.00")])
    fiscal_period = fiscal_period_for_label("FY27-M01")

    preflight = prepare_monthly_close_inputs(
        store="9355",
        period=fiscal_period.period_key,
        fiscal_period=fiscal_period,
        period_start=fiscal_period.start_date,
        period_end=fiscal_period.end_date,
        input_root=input_root,
        input_dir=input_dir,
        micros_path=micros_dir,
        micros_work_dir=tmp_path / "extract",
        stage_weekly=False,
    )

    assert linked_summary.is_symlink()
    assert outside_summary.read_bytes() == b"sensitive non-summary content"
    assert not (input_dir / "summary" / linked_summary.name).exists()
    assert preflight.summary_paths == []


def test_monthly_close_preflight_surfaces_malformed_same_store_summary(tmp_path: Path):
    input_root = tmp_path / "Monthly Close"
    store_root = input_root / "9355"
    input_dir = store_root / "FY27 M01 - Fiscal June"
    darden_dir = input_dir / "darden"
    micros_dir = tmp_path / "micros"
    store_root.mkdir(parents=True)
    darden_dir.mkdir(parents=True)
    micros_dir.mkdir()
    loose_summary = store_root / "07.05.2026 9355 Gift Card Summary.xlsx"
    create_summary(loose_summary, store="9355", activations=Decimal("100.00"), redemptions=Decimal("-300.00"))
    workbook = load_workbook(loose_summary)
    workbook["Summary"]["H3"] = "malformed net settlement"
    workbook.save(loose_summary)
    (darden_dir / "Darden Credit Memo.pdf").write_bytes(b"synthetic test memo")
    write_micros_exports(micros_dir, date(2026, 7, 5), [Decimal("0.00")], [Decimal("0.00")])
    fiscal_period = fiscal_period_for_label("FY27-M01")

    preflight = prepare_monthly_close_inputs(
        store="9355",
        period=fiscal_period.period_key,
        fiscal_period=fiscal_period,
        period_start=fiscal_period.start_date,
        period_end=fiscal_period.end_date,
        input_root=input_root,
        input_dir=input_dir,
        micros_path=micros_dir,
        micros_work_dir=tmp_path / "extract",
        stage_weekly=False,
    )

    staged_summary = input_dir / "summary" / loose_summary.name
    assert not loose_summary.exists()
    assert preflight.summary_paths == [staged_summary]
    assert staged_summary.exists()
    assert not preflight.darden_ready
    assert preflight.darden_message.startswith("REVIEW REQUIRED -")
    assert "malformed net settlement" in preflight.darden_message
    assert preflight.missing_summary_path not in preflight.required_missing_paths


def test_monthly_close_preflight_preserves_different_same_name_summary(tmp_path: Path):
    input_root = tmp_path / "Monthly Close"
    store_root = input_root / "9355"
    input_dir = store_root / "FY27 M01 - Fiscal June"
    summary_dir = input_dir / "summary"
    micros_dir = tmp_path / "micros"
    store_root.mkdir(parents=True)
    summary_dir.mkdir(parents=True)
    micros_dir.mkdir()
    loose_summary = store_root / "07.05.2026 9355 Gift Card Summary.xlsx"
    canonical_summary = summary_dir / loose_summary.name
    create_summary(loose_summary, store="9355", activations=Decimal("111.00"), redemptions=Decimal("-300.00"))
    create_summary(canonical_summary, store="9355", activations=Decimal("222.00"), redemptions=Decimal("-300.00"))
    loose_bytes = loose_summary.read_bytes()
    canonical_bytes = canonical_summary.read_bytes()
    write_micros_exports(micros_dir, date(2026, 7, 5), [Decimal("0.00")], [Decimal("0.00")])
    fiscal_period = fiscal_period_for_label("FY27-M01")

    prepare_monthly_close_inputs(
        store="9355",
        period=fiscal_period.period_key,
        fiscal_period=fiscal_period,
        period_start=fiscal_period.start_date,
        period_end=fiscal_period.end_date,
        input_root=input_root,
        input_dir=input_dir,
        micros_path=micros_dir,
        micros_work_dir=tmp_path / "extract",
        stage_weekly=False,
    )

    assert not loose_summary.exists()
    staged_collision = summary_dir / "07.05.2026 9355 Gift Card Summary_2.xlsx"
    assert canonical_summary.read_bytes() == canonical_bytes
    assert staged_collision.read_bytes() == loose_bytes


def _legacy_darden_mismatch_blocks_close_and_preserves_sources(tmp_path: Path):
    fiscal_period = fiscal_period_for_label("FY27-M01")
    input_dir = tmp_path / "Monthly Close" / "9355" / fiscal_period.folder_name
    summary_path = input_dir / "summary" / "07.05.2026 9355 Gift Card Summary.xlsx"
    activity_path = input_dir / "activity" / "07.05.2026 9355 Gift Card Activity.xlsx"
    summary_path.parent.mkdir(parents=True)
    activity_path.parent.mkdir()
    create_summary(summary_path, store="9355", activations=Decimal("100.00"), redemptions=Decimal("-300.00"))
    create_activity(
        activity_path,
        store="9355",
        begin="29-JUN-2026",
        end="05-JUL-2026",
        activation=Decimal("100.00"),
        redemption=Decimal("-300.00"),
    )
    darden_report = create_darden_report(input_dir / "darden" / "memo.pdf", store="9355", total=Decimal("-199.99"))
    output_path = tmp_path / "Output" / "Gift_Card_Reconciliation_9355_FY27-M01.xlsx"

    with pytest.raises(ParseError, match="Darden final close mismatch"):
        run_monthly_close(
            store="9355",
            period=fiscal_period.period_key,
            period_start=fiscal_period.start_date,
            period_end=fiscal_period.end_date,
            input_dir=input_dir,
            output_path=output_path,
            micros_path=tmp_path / "micros-not-needed",
            micros_work_dir=tmp_path / "extract",
            cleanup_archive_root=tmp_path / "Archive - Old Files",
            fiscal_period=fiscal_period,
            darden_report=darden_report,
        )

    assert not output_path.exists()
    assert summary_path.exists()
    assert activity_path.exists()
    assert darden_report.source_file.exists()
    assert not (tmp_path / "Archive - Old Files" / "monthly-close").exists()


def _legacy_monthly_close_archives_sources_after_success(tmp_path: Path):
    fiscal_period = fiscal_period_for_label("FY27-M01")
    input_dir = tmp_path / "Monthly Close" / "9355" / fiscal_period.folder_name
    summary_dir = input_dir / "summary"
    activity_dir = input_dir / "activity"
    micros_dir = tmp_path / "micros"
    summary_dir.mkdir(parents=True)
    activity_dir.mkdir()
    micros_dir.mkdir()
    summary_path = summary_dir / "07.05.2026 9355 Gift Card Summary.xlsx"
    activity_path = activity_dir / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_summary(summary_path, store="9355", activations=Decimal("100.00"), redemptions=Decimal("-300.00"))
    create_activity(
        activity_path,
        store="9355",
        begin="29-JUN-2026",
        end="05-JUL-2026",
        activation=Decimal("100.00"),
        redemption=Decimal("-300.00"),
    )
    write_micros_exports(micros_dir, date(2026, 6, 29), [Decimal("15.00")] * 7, [Decimal("40.00")] * 6 + [Decimal("70.00")])

    output_path = tmp_path / "Output" / "Gift_Card_Reconciliation_9355_FY27-M01.xlsx"
    darden_report = create_darden_report(input_dir / "darden" / "memo.pdf", store="9355", total=Decimal("-200.00"))
    run_monthly_close(
        store="9355",
        period=fiscal_period.period_key,
        period_start=fiscal_period.start_date,
        period_end=fiscal_period.end_date,
        input_dir=input_dir,
        output_path=output_path,
        micros_path=micros_dir,
        micros_work_dir=tmp_path / "extract",
        cleanup_archive_root=tmp_path / "Archive - Old Files",
        fiscal_period=fiscal_period,
        darden_report=darden_report,
    )

    assert output_path.exists()
    assert not summary_path.exists()
    assert not activity_path.exists()
    assert not darden_report.source_file.exists()
    assert (tmp_path / "Archive - Old Files" / "monthly-close" / "9355" / fiscal_period.folder_name / "summary" / summary_path.name).exists()
    assert (tmp_path / "Archive - Old Files" / "monthly-close" / "9355" / fiscal_period.folder_name / "activity" / activity_path.name).exists()
    assert (tmp_path / "Archive - Old Files" / "monthly-close" / "9355" / fiscal_period.folder_name / "darden" / darden_report.source_file.name).exists()


def _legacy_monthly_close_can_rerun_from_archive_without_deleting_sources(tmp_path: Path):
    fiscal_period = fiscal_period_for_label("FY27-M01")
    input_dir = tmp_path / "Archive - Old Files" / "monthly-close" / "9355" / fiscal_period.folder_name
    summary_dir = input_dir / "summary"
    activity_dir = input_dir / "activity"
    micros_dir = tmp_path / "micros"
    summary_dir.mkdir(parents=True)
    activity_dir.mkdir()
    micros_dir.mkdir()
    summary_path = summary_dir / "07.05.2026 9355 Gift Card Summary.xlsx"
    activity_path = activity_dir / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_summary(summary_path, store="9355", activations=Decimal("100.00"), redemptions=Decimal("-300.00"))
    create_activity(
        activity_path,
        store="9355",
        begin="29-JUN-2026",
        end="05-JUL-2026",
        activation=Decimal("100.00"),
        redemption=Decimal("-300.00"),
    )
    write_micros_exports(micros_dir, date(2026, 6, 29), [Decimal("15.00")] * 7, [Decimal("40.00")] * 6 + [Decimal("70.00")])

    output_path = tmp_path / "Output" / "Gift_Card_Reconciliation_9355_FY27-M01_rerun.xlsx"
    darden_report = create_darden_report(input_dir / "darden" / "memo.pdf", store="9355", total=Decimal("-200.00"))
    run_monthly_close(
        store="9355",
        period=fiscal_period.period_key,
        period_start=fiscal_period.start_date,
        period_end=fiscal_period.end_date,
        input_dir=input_dir,
        output_path=output_path,
        micros_path=micros_dir,
        micros_work_dir=tmp_path / "extract",
        cleanup_archive_root=tmp_path / "Archive - Old Files",
        fiscal_period=fiscal_period,
        darden_report=darden_report,
    )

    assert output_path.exists()
    assert summary_path.exists()
    assert activity_path.exists()
    assert darden_report.source_file.exists()


def _legacy_monthly_close_does_not_archive_sources_when_run_fails(tmp_path: Path):
    fiscal_period = fiscal_period_for_label("FY27-M01")
    input_dir = tmp_path / "Monthly Close" / "9355" / fiscal_period.folder_name
    summary_dir = input_dir / "summary"
    activity_dir = input_dir / "activity"
    summary_dir.mkdir(parents=True)
    activity_dir.mkdir()
    summary_path = summary_dir / "07.05.2026 9355 Gift Card Summary.xlsx"
    activity_path = activity_dir / "07.05.2026 9355 Gift Card Activity.xlsx"
    create_summary(summary_path, store="9355", activations=Decimal("100.00"), redemptions=Decimal("-300.00"))
    create_activity(
        activity_path,
        store="9355",
        begin="29-JUN-2026",
        end="05-JUL-2026",
        activation=Decimal("100.00"),
        redemption=Decimal("-300.00"),
    )
    darden_report = create_darden_report(input_dir / "darden" / "memo.pdf", store="9355", total=Decimal("-200.00"))

    with pytest.raises(ParseError):
        run_monthly_close(
            store="9355",
            period=fiscal_period.period_key,
            period_start=fiscal_period.start_date,
            period_end=fiscal_period.end_date,
            input_dir=input_dir,
            output_path=tmp_path / "Output" / "Gift_Card_Reconciliation_9355_FY27-M01.xlsx",
            micros_path=tmp_path / "missing-micros",
            micros_work_dir=tmp_path / "extract",
            cleanup_archive_root=tmp_path / "Archive - Old Files",
            fiscal_period=fiscal_period,
            darden_report=darden_report,
        )

    assert summary_path.exists()
    assert activity_path.exists()
    assert darden_report.source_file.exists()
    assert not (tmp_path / "Archive - Old Files" / "monthly-close").exists()


def test_run_monthly_close_script_defaults_to_shared_inbox_scan():
    script = (REPO_ROOT / "_program" / "run_monthly_close.ps1").read_text(encoding="utf-8")
    assert '[string]$OperationsRoot = ""' in script
    assert '[string]$Store = ""' in script
    assert '[string]$Period = ""' in script
    assert '[string]$InputRoot = ""' in script
    assert '[string]$MicrosPath = ""' in script
    assert '[string]$DardenPath = ""' in script
    assert '"02 Monthly Close Inputs"' in script
    assert '"03 Finished Reports"' in script
    assert '"04 Archive"' in script
    assert 'Join-Path $DropboxRoot "GETLinkedData-VB"' in script
    assert 'Join-Path $DropboxRoot "micros_data\\RC-Richmond-current"' in script
    assert '[string]$MicrosWorkDir = ""' in script
    assert 'runtime.ps1' in script
    assert 'Initialize-GiftCardReconRuntime' in script
    assert '$MicrosWorkDir = $Runtime.MicrosExtractDir' in script
    assert '.venv\\Scripts\\python.exe' not in script
    assert 'pip install' not in script
    assert '"-m", "gift_card_recon.monthly_close"' in script
    assert '"--operations-root", $OperationsRoot' in script
    assert "Set-Location $OperationsRoot" in script
    assert '"--darden-path", $DardenPath' in script
    assert 'if ($Store -ne "")' in script
    assert 'if ($Period -ne "")' in script
    assert "exit $exitCode" in script

    click_script = (REPO_ROOT / "templates" / "Run Monthly Gift Card Close.cmd").read_text(encoding="utf-8")
    assert "run_monthly_close.ps1" in click_script
    assert '-OperationsRoot "%OPERATIONS_ROOT%"' in click_script
    assert "Review the exact message and diagnostic paths shown above" in click_script
    assert ".venv" not in click_script
    assert "exit /b %EXITCODE%" in click_script

    install_script = (REPO_ROOT / "_program" / "install.ps1").read_text(encoding="utf-8")
    assert "runtime.ps1" in install_script
    assert "Initialize-GiftCardReconRuntime" in install_script
    assert ".venv\\Scripts\\python.exe" not in install_script

    runtime_script = (REPO_ROOT / "_program" / "runtime.ps1").read_text(encoding="utf-8")
    assert '$env:LOCALAPPDATA' in runtime_script
    assert 'Join-Path $localAppData "GiftCardRecon"' in runtime_script
    assert "dependency-fingerprint.sha256" in runtime_script
    assert '$env:PIP_CACHE_DIR' in runtime_script
    assert '$env:PYTHONPYCACHEPREFIX' in runtime_script
    assert 'pyvenv.cfg' in runtime_script
    assert '-m pip check' in runtime_script
    assert 'Local\\GiftCardReconRuntimeInstall' in runtime_script
    assert '.WaitOne(' in runtime_script
    assert '.ReleaseMutex()' in runtime_script
    assert 'requirements.txt' in runtime_script
    assert 'pyproject.toml' in runtime_script

    weekly_click_script = (REPO_ROOT / "templates" / "Run Weekly Gift Card Reconciliation.cmd").read_text(encoding="utf-8")
    assert "run_weekly.ps1" in weekly_click_script
    assert '-OperationsRoot "%OPERATIONS_ROOT%"' in weekly_click_script
    assert ".venv" not in weekly_click_script
    assert "exit /b %EXITCODE%" in weekly_click_script

    installer = (REPO_ROOT / "_program" / "install_operator_assets.ps1").read_text(encoding="utf-8")
    assert "Gift Card Reconciliation Automation" in installer
    assert "Get-FileHash" in installer
    assert "SHA256" in installer
    assert "$PreparedAssets" in installer
    assert "$PreparedRetirements" in installer
    assert ".gcs-{0}-{1}.tmp" in installer
    assert ".gcb-{0}-{1}.tmp" in installer
    assert "Operator file refresh failed and rollback was incomplete" in installer
    assert "the prior operator asset set was restored" in installer
    assert "Check Gift Card Reconciliation Health.cmd" in installer
    assert "Preserving unrecognized same-name operator file" in installer
    assert "01 Weekly Gift Card Activity Reports\\9354 Richmond\\activity" in installer
    assert "01 Weekly Gift Card Activity Reports\\9355 Virginia Beach\\activity" in installer
    assert "02 Monthly Close Inputs\\9354 Richmond" in installer
    assert "02 Monthly Close Inputs\\9355 Virginia Beach" in installer
    assert "03 Finished Reports\\Monthly Close - Review Required" in installer
    assert '"03 Finished Reports\\Review Required"' not in installer
    assert "_automation_runs\\review" in installer
    assert not (REPO_ROOT / "Run-Gift-Card-Reconciliation.cmd").exists()
    assert not (REPO_ROOT / "Run-Monthly-Close.cmd").exists()

    operator_fixture = REPO_ROOT / "_program" / "maintenance" / "test_install_operator_assets.ps1"
    assert operator_fixture.is_file()
    fixture_text = operator_fixture.read_text(encoding="utf-8")
    assert "late-failure" in fixture_text
    assert "released stale health launcher is retired" in fixture_text

    migration_runbook = (REPO_ROOT / "docs" / "NUMBERED_DROPBOX_MIGRATION_RUNBOOK.md").read_text(
        encoding="utf-8"
    )
    assert "run_tests.ps1 -SkipInstall" not in migration_runbook
    assert "deployment-manifest.json" in migration_runbook
    assert "Move-Item -LiteralPath $program -Destination $snapshotBackup" in migration_runbook


def test_darden_staging_preserves_distinct_same_size_files(tmp_path: Path):
    first_source = tmp_path / "first" / "memo.pdf"
    second_source = tmp_path / "second" / "memo.pdf"
    darden_dir = tmp_path / "Monthly Close" / "9355" / "FY27 M01 - Fiscal June" / "darden"
    first_source.parent.mkdir()
    second_source.parent.mkdir()
    first_source.write_bytes(b"AAAA")
    second_source.write_bytes(b"BBBB")

    first_saved = stage_darden_credit_memo(first_source, darden_dir=darden_dir)
    second_saved = stage_darden_credit_memo(second_source, darden_dir=darden_dir)

    assert first_saved.name == "memo.pdf"
    assert second_saved.name == "memo_2.pdf"
    assert first_saved.read_bytes() == b"AAAA"
    assert second_saved.read_bytes() == b"BBBB"


def test_darden_archive_collision_preserves_distinct_same_size_evidence(tmp_path: Path):
    fiscal_period = fiscal_period_for_label("FY27-M01")
    input_dir = tmp_path / "Monthly Close" / "9355" / fiscal_period.folder_name
    source = input_dir / "darden" / "memo.pdf"
    archive_root = tmp_path / "Archive - Old Files"
    existing = archive_root / "Monthly Close" / "9355" / fiscal_period.folder_name / "darden" / "memo.pdf"
    source.parent.mkdir(parents=True)
    existing.parent.mkdir(parents=True)
    source.write_bytes(b"AAAA")
    existing.write_bytes(b"BBBB")

    moved = cleanup_monthly_close_sources(
        input_dir=input_dir,
        archive_root=archive_root,
        store="9355",
        fiscal_period=fiscal_period,
    )

    saved = existing.with_name("memo_2.pdf")
    assert moved == [saved]
    assert not source.exists()
    assert existing.read_bytes() == b"BBBB"
    assert saved.read_bytes() == b"AAAA"


def create_darden_report(path: Path, *, store: str, total: Decimal) -> DardenCreditMemo:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"Synthetic Darden credit memo for unit tests")
    return DardenCreditMemo(
        source_file=path,
        store=store,
        period_start=date(2026, 6, 1),
        period_end=date(2026, 7, 5),
        total=total,
        invoice_number="Synthetic FY27 M01",
        invoice_date=date(2026, 7, 6),
    )


def create_summary(path: Path, *, store: str, activations: Decimal, redemptions: Decimal) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["SUMMARY"])
    ws.append(["Franchise Partner", "Store Number", "Gift Card Franchise Fee Rate", "Total Activations", "Total Redemptions", "Payable Redemptions", "GCDR", "Net Settlement"])
    ws.append(["Sorensen", int(store), 0.1, float(activations), float(redemptions), float(redemptions), 0, float(activations + redemptions)])
    wb.save(path)


def create_activity(
    path: Path,
    *,
    store: str,
    begin: str,
    end: str,
    activation: Decimal,
    redemption: Decimal,
    transaction_date: str = "2026-06-01",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    ws.append([f"All GC Activity BY Rest Number and Date Range  BEGIN DATE: '{begin}', END DATE: '{end}', Rest Number Parameter 1: '{store}'"])
    ws.append(["Card No", "Request", "Request Code Listing", "Business Date", "Transaction No", "Amount SUM", "Promocode", "Authorization Code"])
    ws.append(["0001xxxx", 100, "Activation", transaction_date, 1, float(activation), None, 111111])
    ws.append(["0002xxxx", 202, "Redemption No Nsf", transaction_date, 2, float(redemption), None, 222222])
    wb.save(path)


def write_micros_exports(micros_dir: Path, first_date: date, issue_values: list[Decimal], payment_values: list[Decimal]) -> None:
    assert len(issue_values) == len(payment_values)
    write_micros_export_rows(
        micros_dir,
        [
            (first_date + timedelta(days=idx), issue, payment)
            for idx, (issue, payment) in enumerate(zip(issue_values, payment_values))
        ],
    )


def write_micros_export_rows(micros_dir: Path, rows: list[tuple[date, Decimal, Decimal]]) -> None:
    lines = []
    tender_lines = []
    for business_date, issue, payment in rows:
        row = ["0"] * 132
        row[0] = f"{business_date:%Y-%m-%d} 00:00:00.000"
        row[ISSUE_AMOUNT_INDEX] = f"{issue:.2f}"
        row[PAYMENT_AMOUNT_INDEX] = f"{payment:.2f}"
        lines.append(",".join(row))
        tender_lines.append(f"'{business_date:%Y-%m-%d}',{payment:.2f},350,'G C Payment','T'")
    (micros_dir / "DLYSYSTT.TXT").write_text("\n".join(lines), encoding="utf-8")
    (micros_dir / "TENDER_DETAIL.TXT").write_text("\n".join(tender_lines), encoding="utf-8")


def find_row(ws, value: str) -> int | None:
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == value:
            return row_idx
    return None
