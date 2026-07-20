from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from gift_card_recon.variance_explanation import (
    ACCOUNTING_FORMAT,
    DOCUMENT_TYPE,
    EXPLANATION_INPUT_CELL,
    EXPLANATION_SHEET,
    EXPLANATION_SHEET_NAME,
    IDENTITY_SHEET_NAME,
    MAX_EXPLANATION_LENGTH,
    SCHEMA_VERSION,
    WeeklyVarianceExplanation,
    read_variance_explanation_workbook,
    variance_control_mismatch_details,
    variance_explanation_path,
    write_variance_explanation_workbook,
)


WEEK_START = date(2026, 7, 13)
WEEK_END = date(2026, 7, 19)


def explanation_data(text: str = "") -> WeeklyVarianceExplanation:
    return WeeklyVarianceExplanation(
        store="9354",
        week_start=WEEK_START,
        week_end=WEEK_END,
        pos_issue_variance=Decimal("0.00"),
        pos_payment_variance=Decimal("200.00"),
        pos_net_variance=Decimal("-200.00"),
        tender_variance=Decimal("0.00"),
        explanation=text,
    )


def test_variance_explanation_path_is_deterministic_and_period_scoped(tmp_path: Path) -> None:
    monthly_period_dir = tmp_path / "9354 Richmond" / "FY27 M02 - Fiscal July"

    result = variance_explanation_path(monthly_period_dir, 9354, WEEK_END)

    assert result == (
        monthly_period_dir
        / "Variance Explanations"
        / "Weekly_Variance_9354_2026-07-19.xlsx"
    )
    with pytest.raises(ValueError, match="Sunday"):
        variance_explanation_path(monthly_period_dir, "9354", date(2026, 7, 18))
    with pytest.raises(ValueError, match="four-digit"):
        variance_explanation_path(monthly_period_dir, "../9354", WEEK_END)


def test_dataclass_normalizes_money_and_identifies_large_weekly_control() -> None:
    data = WeeklyVarianceExplanation(
        store=" 9354 ",
        week_start=WEEK_START,
        week_end=WEEK_END,
        pos_issue_variance="5.00",  # type: ignore[arg-type]
        pos_payment_variance="5.01",  # type: ignore[arg-type]
        pos_net_variance="-0.004",  # type: ignore[arg-type]
        tender_variance=0,  # type: ignore[arg-type]
        explanation="  Register correction documented.\r\nReviewed with accounting.  ",
    )

    assert data.store == "9354"
    assert data.pos_issue_variance == Decimal("5.00")
    assert data.pos_payment_variance == Decimal("5.01")
    assert data.pos_net_variance == Decimal("-0.00")
    assert data.explanation == "Register correction documented.\nReviewed with accounting."


def test_control_mismatch_details_use_exact_normalized_cents() -> None:
    data = explanation_data("Reviewed.")

    assert variance_control_mismatch_details(
        data,
        pos_issue_variance=Decimal("0.004"),
        pos_payment_variance=Decimal("201.00"),
        pos_net_variance=Decimal("-200.00"),
        tender_variance=Decimal("0.00"),
    ) == ("POS gift card payment: form +200.00, current +201.00",)
    assert data.requires_explanation is True


def test_writer_creates_professional_one_page_editable_companion(tmp_path: Path) -> None:
    output = tmp_path / "explanation.xlsx"

    written = write_variance_explanation_workbook(explanation_data(), output)

    assert written == output
    workbook = load_workbook(output, data_only=False)
    try:
        assert EXPLANATION_SHEET == EXPLANATION_SHEET_NAME
        assert workbook.sheetnames == [EXPLANATION_SHEET_NAME, IDENTITY_SHEET_NAME]
        visible = workbook[EXPLANATION_SHEET_NAME]
        identity = workbook[IDENTITY_SHEET_NAME]
        assert visible["A1"].value == "WEEKLY VARIANCE EXPLANATION"
        assert visible["B4"].value == "9354"
        assert visible["D4"].value.date() == WEEK_START
        assert visible["F4"].value.date() == WEEK_END
        assert visible["D9"].value == 200
        assert visible["D9"].number_format == ACCOUNTING_FORMAT
        assert visible["F9"].value == "YES"
        assert visible[EXPLANATION_INPUT_CELL].value is None
        assert visible[EXPLANATION_INPUT_CELL].fill.fgColor.rgb.endswith("FFFF00")
        assert visible[EXPLANATION_INPUT_CELL].alignment.wrap_text is True
        assert visible[EXPLANATION_INPUT_CELL].protection.locked is False
        assert visible.protection.sheet is True
        assert visible.print_area == "'Variance Explanation'!$A$1:$F$22"
        assert visible.page_setup.fitToWidth == 1
        assert visible.page_setup.fitToHeight == 1
        assert identity.sheet_state == "veryHidden"
        identity_values = {
            identity.cell(row, 1).value: identity.cell(row, 2).value
            for row in range(1, identity.max_row + 1)
        }
        assert identity_values["schema_version"] == SCHEMA_VERSION
        assert identity_values["document_type"] == DOCUMENT_TYPE
        assert identity_values["input_sheet"] == EXPLANATION_SHEET_NAME
        assert identity_values["input_cell"] == EXPLANATION_INPUT_CELL
    finally:
        workbook.close()


def test_reader_allows_blank_only_when_explicitly_requested(tmp_path: Path) -> None:
    output = write_variance_explanation_workbook(
        explanation_data(),
        tmp_path / "blank.xlsx",
    )

    parsed = read_variance_explanation_workbook(
        output,
        expected_store="9354",
        expected_week_start=WEEK_START,
        expected_week_end=WEEK_END,
        require_text=False,
    )

    assert parsed == explanation_data()
    with pytest.raises(ValueError, match="required.*B15"):
        read_variance_explanation_workbook(output)


def test_reader_round_trips_multiline_operator_explanation(tmp_path: Path) -> None:
    text = (
        "A $200 redemption was posted twice in the POS export.\n"
        "Accounting confirmed the correction will appear next week."
    )
    output = write_variance_explanation_workbook(
        explanation_data(text),
        tmp_path / "complete.xlsx",
    )

    parsed = read_variance_explanation_workbook(
        output,
        expected_store=9354,
        expected_week_end=WEEK_END,
    )

    assert parsed.explanation == text
    assert parsed.pos_payment_variance == Decimal("200.00")
    assert parsed.pos_net_variance == Decimal("-200.00")


def test_reader_rejects_formula_in_operator_input(tmp_path: Path) -> None:
    output = write_variance_explanation_workbook(
        explanation_data(),
        tmp_path / "formula.xlsx",
    )
    workbook = load_workbook(output)
    workbook[EXPLANATION_SHEET_NAME][EXPLANATION_INPUT_CELL] = (
        '=HYPERLINK("https://example.invalid","explanation")'
    )
    workbook.save(output)
    workbook.close()

    with pytest.raises(ValueError, match=r"Formulas are not allowed.*B15"):
        read_variance_explanation_workbook(output)


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("schema_version", 999, "schema version"),
        ("store", "9355", "Visible store/week values"),
        ("week_start", "2026-07-14", "Monday-Sunday"),
        ("week_end", "not-a-date", "valid ISO date"),
    ],
)
def test_reader_rejects_tampered_identity(
    tmp_path: Path,
    field: str,
    value: object,
    message: str,
) -> None:
    output = write_variance_explanation_workbook(
        explanation_data("Documented correction."),
        tmp_path / f"tampered-{field}.xlsx",
    )
    workbook = load_workbook(output)
    identity = workbook[IDENTITY_SHEET_NAME]
    row = next(
        row
        for row in range(1, identity.max_row + 1)
        if identity.cell(row, 1).value == field
    )
    identity.cell(row, 2, value)
    workbook.save(output)
    workbook.close()

    with pytest.raises(ValueError, match=message):
        read_variance_explanation_workbook(output)


def test_reader_rejects_visible_values_that_do_not_match_identity(tmp_path: Path) -> None:
    output = write_variance_explanation_workbook(
        explanation_data("Documented correction."),
        tmp_path / "visible-tamper.xlsx",
    )
    workbook = load_workbook(output)
    visible = workbook[EXPLANATION_SHEET_NAME]
    visible.protection.sheet = False
    visible["D9"] = 199
    workbook.save(output)
    workbook.close()

    with pytest.raises(ValueError, match="pos_payment_variance"):
        read_variance_explanation_workbook(output)


def test_explanation_length_limit_is_enforced_on_write_and_read(tmp_path: Path) -> None:
    accepted = "x" * MAX_EXPLANATION_LENGTH
    output = write_variance_explanation_workbook(
        explanation_data(accepted),
        tmp_path / "maximum.xlsx",
    )
    assert read_variance_explanation_workbook(output).explanation == accepted

    with pytest.raises(ValueError, match="character limit"):
        explanation_data("x" * (MAX_EXPLANATION_LENGTH + 1))

    workbook = load_workbook(output)
    workbook[EXPLANATION_SHEET_NAME][EXPLANATION_INPUT_CELL] = "y" * (
        MAX_EXPLANATION_LENGTH + 1
    )
    workbook.save(output)
    workbook.close()
    with pytest.raises(ValueError, match="character limit"):
        read_variance_explanation_workbook(output)


def test_writer_refuses_to_overwrite_operator_input_without_explicit_opt_in(
    tmp_path: Path,
) -> None:
    output = write_variance_explanation_workbook(
        explanation_data("Original explanation."),
        tmp_path / "existing.xlsx",
    )

    with pytest.raises(FileExistsError, match="already exists"):
        write_variance_explanation_workbook(explanation_data("Replacement."), output)

    assert read_variance_explanation_workbook(output).explanation == "Original explanation."


def test_invalid_week_and_formula_text_are_rejected_before_writing(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="Monday-Sunday"):
        WeeklyVarianceExplanation(
            store="9354",
            week_start=date(2026, 7, 14),
            week_end=WEEK_END,
            pos_issue_variance=Decimal("0.00"),
            pos_payment_variance=Decimal("200.00"),
            pos_net_variance=Decimal("-200.00"),
            tender_variance=Decimal("0.00"),
        )

    with pytest.raises(ValueError, match="Excel formula"):
        write_variance_explanation_workbook(
            explanation_data(" =1+1"),
            tmp_path / "unsafe.xlsx",
        )
